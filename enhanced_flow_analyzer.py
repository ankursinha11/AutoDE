#!/usr/bin/env python3
"""
ADF to Hadoop Pipeline Mapper
=============================

This script reads ADF pipeline names from an Excel file and maps them to 
Hadoop workflows/coordinators found in all Hadoop repositories.

Features:
- Reads ADF pipelines from Excel file
- Scans all Hadoop repositories for workflows/coordinators
- Maps based on name similarity and repository/category matching
- Generates comprehensive mapping report with Yes/No mapping status
"""

import os
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from difflib import SequenceMatcher

class ADFHadoopMapper:
    def __init__(self):
        self.adf_pipelines = []
        self.hadoop_workflows = []
        self.hadoop_coordinators = []
        self.mappings = []
        
    def read_adf_pipelines_from_excel(self, excel_file):
        """Read ADF pipeline list from Excel file"""
        try:
            # Check if file exists
            if not os.path.exists(excel_file):
                print(f"❌ Excel file not found: {excel_file}")
                return []
            
            # Check file extension
            if not excel_file.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
                print(f"❌ Invalid file format: {excel_file}")
                print("   Supported formats: .xlsx, .xlsm, .xltx, .xltm")
                return []
            
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            adf_pipelines = []
            
            # Read all rows, skip header
            for row_num in range(2, ws.max_row + 1):
                category = ws.cell(row=row_num, column=1).value
                pipeline_name = ws.cell(row=row_num, column=2).value
                
                if category and pipeline_name:
                    adf_pipelines.append({
                        'category': str(category).strip(),
                        'pipeline_name': str(pipeline_name).strip()
                    })
            
            print(f"📋 Read {len(adf_pipelines)} ADF pipelines from Excel file")
            return adf_pipelines
            
        except Exception as e:
            print(f"❌ Error reading ADF Excel file: {e}")
            print(f"   File: {excel_file}")
            print("   Please ensure the file is a valid Excel file (.xlsx format)")
            return []
    
    def find_hadoop_repositories(self, base_path):
        """Find all Hadoop repositories in the base path"""
        hadoop_repos = []
        base_path = Path(base_path)
        
        if not base_path.exists():
            print(f"❌ Base path not found: {base_path}")
            return []
        
        # Look for directories starting with 'app-'
        for item in base_path.iterdir():
            if item.is_dir() and item.name.startswith('app-'):
                hadoop_repos.append(item)
        
        print(f"🔍 Found {len(hadoop_repos)} Hadoop repositories:")
        for repo in hadoop_repos:
            print(f"  - {repo.name}")
        
        return hadoop_repos
    
    def scan_hadoop_repository(self, repo_path):
        """Scan a single Hadoop repository for workflows and coordinators"""
        repo_name = repo_path.name
        workflows = []
        coordinators = []
        
        print(f"📁 Scanning repository: {repo_name}")
        
        # Find all workflow.xml files
        workflow_patterns = [
            "**/*workflow.xml",
            "**/workflows/**/*workflow.xml",
            "**/oozie/**/*workflow.xml",
            "**/pipelines/**/*workflow.xml",
            "**/dev/**/*workflow.xml"
        ]
        
        for pattern in workflow_patterns:
            for workflow_file in repo_path.rglob(pattern):
                if workflow_file.is_file():
                    workflows.append({
                        'file_path': str(workflow_file),
                        'repo_name': repo_name,
                        'workflow_name': self.extract_workflow_name(workflow_file),
                        'relative_path': str(workflow_file.relative_to(repo_path))
                    })
        
        # Find all coordinator.xml files
        coordinator_patterns = [
            "**/*coordinator.xml",
            "**/coordinators/**/*coordinator.xml",
            "**/oozie/**/*coordinator.xml"
        ]
        
        for pattern in coordinator_patterns:
            for coordinator_file in repo_path.rglob(pattern):
                if coordinator_file.is_file():
                    coordinators.append({
                        'file_path': str(coordinator_file),
                        'repo_name': repo_name,
                        'coordinator_name': self.extract_coordinator_name(coordinator_file),
                        'relative_path': str(coordinator_file.relative_to(repo_path))
                    })
        
        print(f"  📄 Found {len(workflows)} workflows and {len(coordinators)} coordinators")
        return workflows, coordinators
    
    def extract_workflow_name(self, workflow_file):
        """Extract workflow name from XML file"""
        try:
            tree = ET.parse(workflow_file)
            root = tree.getroot()
            
            # Try to get name attribute
            name = root.get('name')
            if name:
                return name
            
            # Try to get name from different namespace patterns
            namespaces = [
                '{uri:oozie:workflow:0.5}',
                '{uri:oozie:workflow:0.2}',
                '{uri:oozie:workflow:0.1}',
                ''
            ]
            
            for ns in namespaces:
                name_elem = root.find(f'{ns}name')
                if name_elem is not None and name_elem.text:
                    return name_elem.text.strip()
            
            # Fallback to filename
            return workflow_file.stem
            
        except Exception as e:
            print(f"⚠️ Error parsing workflow {workflow_file}: {e}")
            return workflow_file.stem
    
    def extract_coordinator_name(self, coordinator_file):
        """Extract coordinator name from XML file"""
        try:
            tree = ET.parse(coordinator_file)
            root = tree.getroot()
            
            # Try to get name attribute
            name = root.get('name')
            if name:
                return name
            
            # Try to get name from different namespace patterns
            namespaces = [
                '{uri:oozie:coordinator:0.2}',
                '{uri:oozie:coordinator:0.1}',
                ''
            ]
            
            for ns in namespaces:
                name_elem = root.find(f'{ns}name')
                if name_elem is not None and name_elem.text:
                    return name_elem.text.strip()
            
            # Fallback to filename
            return coordinator_file.stem
            
        except Exception as e:
            print(f"⚠️ Error parsing coordinator {coordinator_file}: {e}")
            return coordinator_file.stem
    
    def calculate_similarity(self, adf_name, hadoop_name):
        """Calculate similarity between ADF and Hadoop pipeline names"""
        # Clean names for comparison
        adf_clean = adf_name.lower().replace('pl_', '').replace('pl ', '').replace('_', ' ')
        hadoop_clean = hadoop_name.lower().replace('_', ' ').replace('-', ' ')
        
        # Calculate similarity
        similarity = SequenceMatcher(None, adf_clean, hadoop_clean).ratio()
        
        # Boost similarity for keyword matches
        adf_keywords = set(adf_clean.split())
        hadoop_keywords = set(hadoop_clean.split())
        
        keyword_overlap = len(adf_keywords.intersection(hadoop_keywords))
        if keyword_overlap > 0:
            similarity += keyword_overlap * 0.1
        
        return min(similarity, 1.0)
    
    def map_adf_to_hadoop(self):
        """Map ADF pipelines to Hadoop workflows/coordinators"""
        print(f"\n🔗 Mapping {len(self.adf_pipelines)} ADF pipelines to Hadoop workflows...")
        
        mappings = []
        
        for adf_pipeline in self.adf_pipelines:
            adf_name = adf_pipeline['pipeline_name']
            adf_category = adf_pipeline['category']
            
            best_match = None
            best_score = 0
            match_type = None
            
            # Repository mapping based on category
            repo_mapping = {
                "dataingestion": "app-data-ingestion",
                "cdd": "app-cdd", 
                "gmrn": "app-globalmrn",
                "leaddiscovery": "app-lead-discovery",
                "leadrepository": "app-lead-repository",
                "leadservicebase": "app-lead-service-base",
                "ops hints": "app-ops-hints",
                "coverage helpers": "app-coverage-helpers",
                "audit": "app-data-ingestion",
                "maintainance": "app-data-ingestion",
                "hintsdiscovery": "app-ops-hints",
                "hfc": "app-lead-repository",
                "chc": "app-data-ingestion",
                "blacklisted": "app-data-ingestion",
                "copy": "app-data-ingestion",
                "test": "app-data-ingestion",
                "mh (lead repository)": "app-lead-repository",
                "chc (lead repository)": "app-lead-repository"
            }
            
            expected_repo = repo_mapping.get(adf_category.lower(), None)
            
            # If no exact match, try fuzzy matching
            if not expected_repo:
                for mapped_category, repo in repo_mapping.items():
                    if adf_category.lower() in mapped_category or mapped_category in adf_category.lower():
                        expected_repo = repo
                        break
            
            # If still no match, try to infer from pipeline name
            if not expected_repo:
                pipeline_lower = adf_name.lower()
                if 'dataingestion' in pipeline_lower or 'ingest' in pipeline_lower:
                    expected_repo = 'app-data-ingestion'
                elif 'cdd' in pipeline_lower:
                    expected_repo = 'app-cdd'
                elif 'gmrn' in pipeline_lower or 'globalmrn' in pipeline_lower:
                    expected_repo = 'app-globalmrn'
                elif 'lead' in pipeline_lower and 'discovery' in pipeline_lower:
                    expected_repo = 'app-lead-discovery'
                elif 'lead' in pipeline_lower and 'repository' in pipeline_lower:
                    expected_repo = 'app-lead-repository'
                elif 'lead' in pipeline_lower and 'service' in pipeline_lower:
                    expected_repo = 'app-lead-service-base'
                elif 'coverage' in pipeline_lower or 'helper' in pipeline_lower:
                    expected_repo = 'app-coverage-helpers'
                elif 'hints' in pipeline_lower or 'ops' in pipeline_lower:
                    expected_repo = 'app-ops-hints'
            
            # Search for matches in workflows
            for workflow in self.hadoop_workflows:
                if expected_repo and workflow['repo_name'] != expected_repo:
                    continue
                
                score = self.calculate_similarity(adf_name, workflow['workflow_name'])
                if score > best_score:
                    best_score = score
                    best_match = workflow
                    match_type = 'Workflow'
            
            # Search for matches in coordinators
            for coordinator in self.hadoop_coordinators:
                if expected_repo and coordinator['repo_name'] != expected_repo:
                    continue
                
                score = self.calculate_similarity(adf_name, coordinator['coordinator_name'])
                if score > best_score:
                    best_score = score
                    best_match = coordinator
                    match_type = 'Coordinator'
            
            # Determine mapping status
            mapping_found = "Yes" if best_score >= 0.4 else "No"
            
            if best_match:
                mappings.append({
                    'adf_category': adf_category,
                    'adf_pipeline_name': adf_name,
                    'expected_repo': expected_repo or 'Unknown',
                    'mapped_repo': best_match['repo_name'],
                    'mapped_name': best_match.get('workflow_name', best_match.get('coordinator_name')),
                    'mapped_type': match_type,
                    'similarity_score': best_score,
                    'mapping_found': mapping_found,
                    'relative_path': best_match['relative_path']
                })
            else:
                mappings.append({
                    'adf_category': adf_category,
                    'adf_pipeline_name': adf_name,
                    'expected_repo': expected_repo or 'Unknown',
                    'mapped_repo': 'N/A',
                    'mapped_name': 'N/A',
                    'mapped_type': 'N/A',
                    'similarity_score': 0.0,
                    'mapping_found': 'No',
                    'relative_path': 'N/A'
                })
        
        self.mappings = mappings
        return mappings
    
    def create_excel_report(self, output_file):
        """Create comprehensive Excel mapping report"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create main mapping sheet
        self.create_mapping_sheet(wb)
        
        # Create summary sheet
        self.create_summary_sheet(wb)
        
        # Create Hadoop inventory sheet
        self.create_hadoop_inventory_sheet(wb)
        
        # Save the workbook
        wb.save(output_file)
        print(f"📊 Excel report created: {output_file}")
    
    def create_mapping_sheet(self, wb):
        """Create the main ADF to Hadoop mapping sheet"""
        ws = wb.create_sheet("ADF_Hadoop_Mapping")
        
        # Add title
        ws['A1'] = "ADF to Hadoop Pipeline Mapping Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:J1')
        
        # Headers
        headers = [
            "ADF Category", "ADF Pipeline Name", "Expected Hadoop Repo", 
            "Mapped Hadoop Repo", "Mapped Name", "Type (Workflow/Coordinator)",
            "Similarity Score", "Mapping Found", "Relative Path", "Notes"
        ]
        
        ws.append([])  # Empty row
        ws.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add mapping data
        for mapping in self.mappings:
            notes = ""
            if mapping['mapping_found'] == 'Yes':
                if mapping['similarity_score'] >= 0.7:
                    notes = "Strong match"
                elif mapping['similarity_score'] >= 0.5:
                    notes = "Good match"
                else:
                    notes = "Weak match"
            else:
                notes = "No suitable match found"
            
            row = [
                mapping['adf_category'],
                mapping['adf_pipeline_name'],
                mapping['expected_repo'],
                mapping['mapped_repo'],
                mapping['mapped_name'],
                mapping['mapped_type'],
                f"{mapping['similarity_score']:.2f}",
                mapping['mapping_found'],
                mapping['relative_path'],
                notes
            ]
            ws.append(row)
        
        # Style the data rows
        for row_num in range(4, ws.max_row + 1):
            mapping_found_cell = ws.cell(row=row_num, column=8)
            if mapping_found_cell.value == "Yes":
                mapping_found_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                mapping_found_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, wb):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Mapping_Summary")
        
        # Add title
        ws['A1'] = "ADF to Hadoop Mapping Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Calculate statistics
        total_pipelines = len(self.mappings)
        mapped_pipelines = len([m for m in self.mappings if m['mapping_found'] == 'Yes'])
        unmapped_pipelines = total_pipelines - mapped_pipelines
        success_rate = (mapped_pipelines / total_pipelines * 100) if total_pipelines > 0 else 0
        
        # Add summary data
        ws.append([])
        ws.append(["SUMMARY STATISTICS"])
        ws.append([f"Total ADF Pipelines: {total_pipelines}"])
        ws.append([f"Successfully Mapped: {mapped_pipelines}"])
        ws.append([f"Not Mapped: {unmapped_pipelines}"])
        ws.append([f"Mapping Success Rate: {success_rate:.1f}%"])
        
        # Add repository breakdown
        ws.append([])
        ws.append(["REPOSITORY BREAKDOWN"])
        repo_counts = {}
        for mapping in self.mappings:
            repo = mapping['expected_repo']
            repo_counts[repo] = repo_counts.get(repo, 0) + 1
        
        for repo, count in sorted(repo_counts.items()):
            ws.append([f"{repo}: {count} ADF pipelines"])
        
        # Add category breakdown
        ws.append([])
        ws.append(["CATEGORY BREAKDOWN"])
        category_counts = {}
        for mapping in self.mappings:
            category = mapping['adf_category']
            category_counts[category] = category_counts.get(category, 0) + 1
        
        for category, count in sorted(category_counts.items()):
            ws.append([f"{category}: {count} ADF pipelines"])
    
    def create_hadoop_inventory_sheet(self, wb):
        """Create Hadoop workflows/coordinators inventory sheet"""
        ws = wb.create_sheet("Hadoop_Inventory")
        
        # Add title
        ws['A1'] = "Hadoop Workflows and Coordinators Inventory"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Add workflows
        ws.append([])
        ws.append(["WORKFLOWS"])
        ws.append(["Repository", "Workflow Name", "Relative Path"])
        
        for workflow in self.hadoop_workflows:
            ws.append([
                workflow['repo_name'],
                workflow['workflow_name'],
                workflow['relative_path']
            ])
        
        # Add coordinators
        ws.append([])
        ws.append(["COORDINATORS"])
        ws.append(["Repository", "Coordinator Name", "Relative Path"])
        
        for coordinator in self.hadoop_coordinators:
            ws.append([
                coordinator['repo_name'],
                coordinator['coordinator_name'],
                coordinator['relative_path']
            ])
        
        # Add totals
        ws.append([])
        ws.append(["TOTALS"])
        ws.append([f"Total Workflows: {len(self.hadoop_workflows)}"])
        ws.append([f"Total Coordinators: {len(self.hadoop_coordinators)}"])
        ws.append([f"Total Hadoop Pipelines: {len(self.hadoop_workflows) + len(self.hadoop_coordinators)}"])

def main():
    print("=" * 80)
    print("🚀 ADF TO HADOOP PIPELINE MAPPER")
    print("=" * 80)
    print("Maps ADF pipelines to Hadoop workflows/coordinators based on name similarity")
    print()
    
    # Get input parameters
    if len(sys.argv) > 2:
        base_path = sys.argv[1]
        adf_excel_file = sys.argv[2]
    else:
        base_path = input("Enter the path containing all Hadoop repositories: ").strip()
        adf_excel_file = input("Enter the path to ADF pipeline Excel file: ").strip()
    
    # Initialize mapper
    mapper = ADFHadoopMapper()
    
    # Read ADF pipelines
    print(f"\n📋 Reading ADF pipelines from: {adf_excel_file}")
    adf_pipelines = mapper.read_adf_pipelines_from_excel(adf_excel_file)
    if not adf_pipelines:
        print("❌ No ADF pipelines found. Exiting.")
        return
    
    mapper.adf_pipelines = adf_pipelines
    
    # Find Hadoop repositories
    print(f"\n🔍 Scanning for Hadoop repositories in: {base_path}")
    hadoop_repos = mapper.find_hadoop_repositories(base_path)
    if not hadoop_repos:
        print("❌ No Hadoop repositories found. Exiting.")
        return
    
    # Scan all repositories
    print(f"\n📊 Scanning all Hadoop repositories...")
    all_workflows = []
    all_coordinators = []
    
    for repo in hadoop_repos:
        workflows, coordinators = mapper.scan_hadoop_repository(repo)
        all_workflows.extend(workflows)
        all_coordinators.extend(coordinators)
    
    mapper.hadoop_workflows = all_workflows
    mapper.hadoop_coordinators = all_coordinators
    
    print(f"\n📄 Total Hadoop pipelines found:")
    print(f"  - Workflows: {len(all_workflows)}")
    print(f"  - Coordinators: {len(all_coordinators)}")
    print(f"  - Total: {len(all_workflows) + len(all_coordinators)}")
    
    # Perform mapping
    mappings = mapper.map_adf_to_hadoop()
    
    # Create Excel report
    output_file = "ADF_HADOOP_MAPPING_REPORT.xlsx"
    mapper.create_excel_report(output_file)
    
    # Print summary
    total_pipelines = len(mappings)
    mapped_pipelines = len([m for m in mappings if m['mapping_found'] == 'Yes'])
    success_rate = (mapped_pipelines / total_pipelines * 100) if total_pipelines > 0 else 0
    
    print("\n" + "=" * 80)
    print("🎯 MAPPING SUMMARY")
    print("=" * 80)
    print(f"📋 Total ADF Pipelines: {total_pipelines}")
    print(f"✅ Successfully Mapped: {mapped_pipelines}")
    print(f"❌ Not Mapped: {total_pipelines - mapped_pipelines}")
    print(f"📊 Mapping Success Rate: {success_rate:.1f}%")
    print(f"\n📄 Report saved as: {output_file}")
    print("=" * 80)

if __name__ == "__main__":
    main()
