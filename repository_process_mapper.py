import os
import re
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import ast
import sqlparse
from collections import defaultdict

class RepositoryAnalyzer:
    """
    Enhanced repository analyzer that automatically pairs Hadoop and Azure Databricks repositories
    and generates separate Excel mappings for each pair
    """
    
    def __init__(self):
        self.hadoop_processes = []
        self.databricks_processes = []
        self.mappings = []
        self.business_logic_patterns = {
            'permid': ['permid', 'person_id', 'personid', 'tu_id', 'transunion'],
            'coverage': ['coverage', 'insurance', 'policy', 'payer', 'benefit'],
            'patient': ['patient', 'demographic', 'demographics', 'patientacct'],
            'address': ['address', 'addr', 'location', 'zip', 'city', 'state'],
            'phone': ['phone', 'telephone', 'contact'],
            'family': ['family', 'household', 'member', 'dependent', 'subscriber'],
            'lead': ['lead', 'discovery', 'generation', 'propagation'],
            'validation': ['validate', 'validation', 'check', 'verify'],
            'transformation': ['transform', 'convert', 'map', 'process'],
            'merge': ['merge', 'join', 'combine', 'union'],
            'filter': ['filter', 'where', 'condition', 'criteria'],
            'group': ['group', 'aggregate', 'sum', 'count', 'avg'],
            'sort': ['sort', 'order', 'rank'],
            'dedupe': ['dedupe', 'distinct', 'unique', 'duplicate']
        }
    
    def find_repository_pairs(self, base_path):
        """Find all Hadoop-Databricks repository pairs"""
        base_path = Path(base_path)
        
        # Find all Hadoop repos (starting with app-)
        hadoop_repos = [d for d in base_path.iterdir() if d.is_dir() and d.name.startswith('app-')]
        
        # Find all Databricks repos (not starting with app-)
        databricks_repos = [d for d in base_path.iterdir() if d.is_dir() and not d.name.startswith('app-')]
        
        print(f"Found {len(hadoop_repos)} Hadoop repositories:")
        for repo in hadoop_repos:
            print(f"  - {repo.name}")
        
        print(f"\nFound {len(databricks_repos)} Databricks repositories:")
        for repo in databricks_repos:
            print(f"  - {repo.name}")
        
        # Create pairs based on naming patterns
        pairs = []
        
        for hadoop_repo in hadoop_repos:
            hadoop_name = hadoop_repo.name.replace('app-', '')  # Remove app- prefix
            
            # Try to find matching Databricks repo
            matching_databricks = None
            
            # Direct name match
            for databricks_repo in databricks_repos:
                if databricks_repo.name.lower() == hadoop_name.lower():
                    matching_databricks = databricks_repo
                    break
            
            # If no direct match, try partial matches
            if not matching_databricks:
                for databricks_repo in databricks_repos:
                    databricks_name = databricks_repo.name.lower()
                    if (hadoop_name.lower() in databricks_name or 
                        databricks_name in hadoop_name.lower() or
                        self._calculate_name_similarity(hadoop_name, databricks_name) > 0.6):
                        matching_databricks = databricks_repo
                        break
            
            if matching_databricks:
                pairs.append({
                    'hadoop_repo': hadoop_repo,
                    'databricks_repo': matching_databricks,
                    'hadoop_name': hadoop_repo.name,
                    'databricks_name': matching_databricks.name,
                    'pair_name': f"{hadoop_repo.name}_to_{matching_databricks.name}"
                })
                print(f"✅ Paired: {hadoop_repo.name} ↔ {matching_databricks.name}")
            else:
                print(f"⚠️ No match found for: {hadoop_repo.name}")
        
        return pairs
    
    def _calculate_name_similarity(self, name1, name2):
        """Calculate similarity between two names"""
        name1_lower = name1.lower()
        name2_lower = name2.lower()
        
        # Check for common substrings
        common_chars = 0
        for char in name1_lower:
            if char in name2_lower:
                common_chars += 1
        
        return common_chars / max(len(name1_lower), len(name2_lower))
    
    def analyze_repository_pair(self, hadoop_path, databricks_path):
        """Analyze a single Hadoop-Databricks repository pair"""
        print(f"\n{'='*80}")
        print(f"ANALYZING PAIR: {Path(hadoop_path).name} ↔ {Path(databricks_path).name}")
        print(f"{'='*80}")
        
        # Reset for this pair
        self.hadoop_processes = []
        self.databricks_processes = []
        self.mappings = []
        
        # Analyze Hadoop repository
        print(f"\n🔍 Analyzing Hadoop repository: {hadoop_path}")
        self.analyze_hadoop_repository(hadoop_path)
        
        # Analyze Databricks repository
        print(f"\n🔍 Analyzing Databricks repository: {databricks_path}")
        self.analyze_databricks_repository(databricks_path)
        
        # Create mappings
        print(f"\n🔗 Creating process mappings...")
        self.create_mappings()
        
        return {
            'hadoop_processes': len(self.hadoop_processes),
            'databricks_processes': len(self.databricks_processes),
            'mappings': len(self.mappings)
        }
    
    def analyze_hadoop_repository(self, hadoop_path):
        """Analyze Hadoop repository structure and content"""
        hadoop_path = Path(hadoop_path)
        
        # Analyze Pig scripts
        pig_files = list(hadoop_path.rglob("*.pig"))
        print(f"  📄 Found {len(pig_files)} Pig scripts")
        for pig_file in pig_files:
            process = self.analyze_pig_script(pig_file)
            if process:
                self.hadoop_processes.append(process)
        
        # Analyze PySpark scripts
        py_files = list(hadoop_path.rglob("*.py"))
        print(f"  📄 Found {len(py_files)} Python scripts")
        for py_file in py_files:
            process = self.analyze_pyspark_script(py_file)
            if process:
                self.hadoop_processes.append(process)
        
        # Analyze SQL files
        sql_files = list(hadoop_path.rglob("*.sql"))
        print(f"  📄 Found {len(sql_files)} SQL scripts")
        for sql_file in sql_files:
            process = self.analyze_sql_script(sql_file)
            if process:
                self.hadoop_processes.append(process)
        
        print(f"  ✅ Analyzed {len(self.hadoop_processes)} Hadoop processes")
    
    def analyze_databricks_repository(self, databricks_path):
        """Analyze Azure Databricks repository structure and content"""
        databricks_path = Path(databricks_path)
        
        # Analyze Python notebooks
        py_files = list(databricks_path.rglob("*.py"))
        print(f"  📄 Found {len(py_files)} Python notebooks")
        for py_file in py_files:
            process = self.analyze_databricks_notebook(py_file)
            if process:
                self.databricks_processes.append(process)
        
        # Analyze SQL notebooks
        sql_files = list(databricks_path.rglob("*.sql"))
        print(f"  📄 Found {len(sql_files)} SQL notebooks")
        for sql_file in sql_files:
            process = self.analyze_databricks_sql(sql_file)
            if process:
                self.databricks_processes.append(process)
        
        # Analyze Scala notebooks
        scala_files = list(databricks_path.rglob("*.scala"))
        print(f"  📄 Found {len(scala_files)} Scala notebooks")
        for scala_file in scala_files:
            process = self.analyze_scala_notebook(scala_file)
            if process:
                self.databricks_processes.append(process)
        
        print(f"  ✅ Analyzed {len(self.databricks_processes)} Databricks processes")
    
    def analyze_pig_script(self, pig_file):
        """Analyze Pig script content and extract business logic"""
        try:
            with open(pig_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract Pig-specific patterns
            loads = self.extract_pig_loads(content)
            transformations = self.extract_pig_transformations(content)
            stores = self.extract_pig_stores(content)
            functions = self.extract_pig_functions(content)
            
            # Analyze business logic
            business_logic = self.analyze_business_logic(content, pig_file.name)
            
            # Extract data flow
            data_flow = self.extract_pig_data_flow(content)
            
            process = {
                'type': 'Pig Script',
                'name': pig_file.name,
                'path': str(pig_file),
                'relative_path': str(pig_file.relative_to(pig_file.parents[2])),
                'loads': loads,
                'transformations': transformations,
                'stores': stores,
                'functions': functions,
                'business_logic': business_logic,
                'data_flow': data_flow,
                'content_snippet': content[:500] + "..." if len(content) > 500 else content
            }
            
            return process
            
        except Exception as e:
            print(f"    ⚠️ Error analyzing Pig script {pig_file}: {e}")
            return None
    
    def analyze_pyspark_script(self, py_file):
        """Analyze PySpark script content and extract business logic"""
        try:
            with open(py_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract Spark-specific patterns
            inputs = self.extract_spark_inputs(content)
            outputs = self.extract_spark_outputs(content)
            transformations = self.extract_spark_transformations(content)
            functions = self.extract_spark_functions(content)
            
            # Analyze business logic
            business_logic = self.analyze_business_logic(content, py_file.name)
            
            # Extract data flow
            data_flow = self.extract_spark_data_flow(content)
            
            process = {
                'type': 'PySpark Script',
                'name': py_file.name,
                'path': str(py_file),
                'relative_path': str(py_file.relative_to(py_file.parents[2])),
                'inputs': inputs,
                'outputs': outputs,
                'transformations': transformations,
                'functions': functions,
                'business_logic': business_logic,
                'data_flow': data_flow,
                'content_snippet': content[:500] + "..." if len(content) > 500 else content
            }
            
            return process
            
        except Exception as e:
            print(f"    ⚠️ Error analyzing PySpark script {py_file}: {e}")
            return None
    
    def analyze_databricks_notebook(self, py_file):
        """Analyze Databricks Python notebook content"""
        try:
            with open(py_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract Databricks-specific patterns
            inputs = self.extract_databricks_inputs(content)
            outputs = self.extract_databricks_outputs(content)
            transformations = self.extract_databricks_transformations(content)
            functions = self.extract_databricks_functions(content)
            
            # Analyze business logic
            business_logic = self.analyze_business_logic(content, py_file.name)
            
            # Extract data flow
            data_flow = self.extract_databricks_data_flow(content)
            
            process = {
                'type': 'Databricks Notebook',
                'name': py_file.name,
                'path': str(py_file),
                'relative_path': str(py_file.relative_to(py_file.parents[2])),
                'inputs': inputs,
                'outputs': outputs,
                'transformations': transformations,
                'functions': functions,
                'business_logic': business_logic,
                'data_flow': data_flow,
                'content_snippet': content[:500] + "..." if len(content) > 500 else content
            }
            
            return process
            
        except Exception as e:
            print(f"    ⚠️ Error analyzing Databricks notebook {py_file}: {e}")
            return None
    
    def analyze_sql_script(self, sql_file):
        """Analyze SQL script content"""
        try:
            with open(sql_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Parse SQL content
            parsed = sqlparse.parse(content)
            
            # Extract SQL patterns
            tables = self.extract_sql_tables(content)
            columns = self.extract_sql_columns(content)
            operations = self.extract_sql_operations(content)
            
            # Analyze business logic
            business_logic = self.analyze_business_logic(content, sql_file.name)
            
            process = {
                'type': 'SQL Script',
                'name': sql_file.name,
                'path': str(sql_file),
                'relative_path': str(sql_file.relative_to(sql_file.parents[2])),
                'tables': tables,
                'columns': columns,
                'operations': operations,
                'business_logic': business_logic,
                'content_snippet': content[:500] + "..." if len(content) > 500 else content
            }
            
            return process
            
        except Exception as e:
            print(f"    ⚠️ Error analyzing SQL script {sql_file}: {e}")
            return None
    
    def analyze_databricks_sql(self, sql_file):
        """Analyze Databricks SQL notebook"""
        return self.analyze_sql_script(sql_file)  # Same analysis for SQL
    
    def analyze_scala_notebook(self, scala_file):
        """Analyze Scala notebook content"""
        try:
            with open(scala_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract Scala-specific patterns
            inputs = self.extract_scala_inputs(content)
            outputs = self.extract_scala_outputs(content)
            transformations = self.extract_scala_transformations(content)
            
            # Analyze business logic
            business_logic = self.analyze_business_logic(content, scala_file.name)
            
            process = {
                'type': 'Scala Notebook',
                'name': scala_file.name,
                'path': str(scala_file),
                'relative_path': str(scala_file.relative_to(scala_file.parents[2])),
                'inputs': inputs,
                'outputs': outputs,
                'transformations': transformations,
                'business_logic': business_logic,
                'content_snippet': content[:500] + "..." if len(content) > 500 else content
            }
            
            return process
            
        except Exception as e:
            print(f"    ⚠️ Error analyzing Scala notebook {scala_file}: {e}")
            return None
    
    def extract_pig_loads(self, content):
        """Extract LOAD statements from Pig script"""
        loads = []
        pattern = r"(\w+)\s*=\s*LOAD\s+['\"]([^'\"]+)['\"]"
        matches = re.findall(pattern, content, re.IGNORECASE)
        return [{"alias": match[0], "path": match[1]} for match in matches]
    
    def extract_pig_transformations(self, content):
        """Extract transformations from Pig script"""
        transformations = []
        patterns = [
            r"(\w+)\s*=\s*FOREACH\s+(\w+)\s+GENERATE\s+([^;]+)",
            r"(\w+)\s*=\s*FILTER\s+(\w+)\s+BY\s+([^;]+)",
            r"(\w+)\s*=\s*GROUP\s+(\w+)\s+BY\s+([^;]+)",
            r"(\w+)\s*=\s*JOIN\s+(\w+)\s+BY\s+([^;]+)",
            r"(\w+)\s*=\s*UNION\s+(\w+)\s+(\w+)",
            r"(\w+)\s*=\s*DISTINCT\s+(\w+)"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content, re.IGNORECASE)
            transformations.extend(matches)
        
        return transformations
    
    def extract_pig_stores(self, content):
        """Extract STORE statements from Pig script"""
        stores = []
        pattern = r"STORE\s+(\w+)\s+INTO\s+['\"]([^'\"]+)['\"]"
        matches = re.findall(pattern, content, re.IGNORECASE)
        return [{"alias": match[0], "path": match[1]} for match in matches]
    
    def extract_pig_functions(self, content):
        """Extract function definitions from Pig script"""
        functions = []
        pattern = r"DEFINE\s+(\w+)"
        matches = re.findall(pattern, content, re.IGNORECASE)
        return matches
    
    def extract_spark_inputs(self, content):
        """Extract input sources from Spark script"""
        inputs = []
        patterns = [
            r"\.load\(['\"]([^'\"]+)['\"]",
            r"\.read\.format\(['\"]([^'\"]+)['\"]",
            r"\.read\.csv\(['\"]([^'\"]+)['\"]",
            r"\.read\.parquet\(['\"]([^'\"]+)['\"]",
            r"\.read\.json\(['\"]([^'\"]+)['\"]",
            r"spark\.read\.table\(['\"]([^'\"]+)['\"]"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content)
            inputs.extend(matches)
        
        return list(set(inputs))
    
    def extract_spark_outputs(self, content):
        """Extract output sources from Spark script"""
        outputs = []
        patterns = [
            r"\.save\(['\"]([^'\"]+)['\"]",
            r"\.write\.format\(['\"]([^'\"]+)['\"]",
            r"\.write\.csv\(['\"]([^'\"]+)['\"]",
            r"\.write\.parquet\(['\"]([^'\"]+)['\"]",
            r"\.write\.json\(['\"]([^'\"]+)['\"]",
            r"\.write\.saveAsTable\(['\"]([^'\"]+)['\"]"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content)
            outputs.extend(matches)
        
        return list(set(outputs))
    
    def extract_spark_transformations(self, content):
        """Extract data transformations from Spark script"""
        transformations = []
        patterns = [
            r"\.select\(([^)]+)\)",
            r"\.filter\(([^)]+)\)",
            r"\.groupBy\(([^)]+)\)",
            r"\.join\(([^)]+)\)",
            r"\.withColumn\(([^)]+)\)",
            r"\.agg\(([^)]+)\)",
            r"\.drop\(([^)]+)\)",
            r"\.fillna\(([^)]+)\)",
            r"\.replace\(([^)]+)\)",
            r"\.distinct\(\)",
            r"\.union\(([^)]+)\)"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content)
            transformations.extend(matches)
        
        return transformations
    
    def extract_spark_functions(self, content):
        """Extract function definitions from Spark script"""
        functions = []
        pattern = r"def\s+(\w+)\s*\("
        matches = re.findall(pattern, content)
        return matches
    
    def extract_databricks_inputs(self, content):
        """Extract input sources from Databricks notebook"""
        return self.extract_spark_inputs(content)  # Same as Spark
    
    def extract_databricks_outputs(self, content):
        """Extract output sources from Databricks notebook"""
        return self.extract_spark_outputs(content)  # Same as Spark
    
    def extract_databricks_transformations(self, content):
        """Extract transformations from Databricks notebook"""
        return self.extract_spark_transformations(content)  # Same as Spark
    
    def extract_databricks_functions(self, content):
        """Extract function definitions from Databricks notebook"""
        return self.extract_spark_functions(content)  # Same as Spark
    
    def extract_scala_inputs(self, content):
        """Extract input sources from Scala notebook"""
        inputs = []
        patterns = [
            r"\.load\(['\"]([^'\"]+)['\"]",
            r"\.read\.format\(['\"]([^'\"]+)['\"]",
            r"\.read\.csv\(['\"]([^'\"]+)['\"]",
            r"\.read\.parquet\(['\"]([^'\"]+)['\"]"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content)
            inputs.extend(matches)
        
        return list(set(inputs))
    
    def extract_scala_outputs(self, content):
        """Extract output sources from Scala notebook"""
        outputs = []
        patterns = [
            r"\.save\(['\"]([^'\"]+)['\"]",
            r"\.write\.format\(['\"]([^'\"]+)['\"]",
            r"\.write\.csv\(['\"]([^'\"]+)['\"]",
            r"\.write\.parquet\(['\"]([^'\"]+)['\"]"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content)
            outputs.extend(matches)
        
        return list(set(outputs))
    
    def extract_scala_transformations(self, content):
        """Extract transformations from Scala notebook"""
        transformations = []
        patterns = [
            r"\.select\(([^)]+)\)",
            r"\.filter\(([^)]+)\)",
            r"\.groupBy\(([^)]+)\)",
            r"\.join\(([^)]+)\)",
            r"\.withColumn\(([^)]+)\)"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content)
            transformations.extend(matches)
        
        return transformations
    
    def extract_sql_tables(self, content):
        """Extract table references from SQL"""
        tables = []
        patterns = [
            r"FROM\s+(\w+)",
            r"JOIN\s+(\w+)",
            r"INTO\s+(\w+)",
            r"UPDATE\s+(\w+)",
            r"DELETE\s+FROM\s+(\w+)"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content, re.IGNORECASE)
            tables.extend(matches)
        
        return list(set(tables))
    
    def extract_sql_columns(self, content):
        """Extract column references from SQL"""
        columns = []
        pattern = r"SELECT\s+([^FROM]+)"
        matches = re.findall(pattern, content, re.IGNORECASE)
        for match in matches:
            cols = [col.strip() for col in match.split(',')]
            columns.extend(cols)
        
        return list(set(columns))
    
    def extract_sql_operations(self, content):
        """Extract SQL operations"""
        operations = []
        patterns = [
            r"(SELECT)",
            r"(INSERT)",
            r"(UPDATE)",
            r"(DELETE)",
            r"(CREATE)",
            r"(ALTER)",
            r"(DROP)"
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, content, re.IGNORECASE)
            operations.extend(matches)
        
        return list(set(operations))
    
    def analyze_business_logic(self, content, filename):
        """Analyze business logic from content and filename"""
        logic_hints = []
        content_lower = content.lower()
        filename_lower = filename.lower()
        
        # Analyze filename patterns
        for key, patterns in self.business_logic_patterns.items():
            for pattern in patterns:
                if pattern in filename_lower:
                    logic_hints.append(f"{key}_processing")
                    break
        
        # Analyze content patterns
        for key, patterns in self.business_logic_patterns.items():
            for pattern in patterns:
                if pattern in content_lower:
                    logic_hints.append(f"{key}_logic")
                    break
        
        # Look for specific business terms
        business_terms = {
            'patient': 'patient_data_processing',
            'coverage': 'coverage_discovery',
            'insurance': 'insurance_processing',
            'demographic': 'demographic_processing',
            'lead': 'lead_generation',
            'edi': 'edi_processing',
            'hospital': 'hospital_data_processing',
            'policy': 'policy_processing',
            'permid': 'permid_processing',
            'gmrn': 'gmrn_processing'
        }
        
        for term, description in business_terms.items():
            if term in content_lower:
                logic_hints.append(description)
        
        return list(set(logic_hints))
    
    def extract_pig_data_flow(self, content):
        """Extract data flow from Pig script"""
        flow = []
        # Extract variable assignments and their relationships
        pattern = r"(\w+)\s*=\s*(\w+)\s+([^;]+)"
        matches = re.findall(pattern, content)
        for match in matches:
            flow.append(f"{match[0]} = {match[1]} {match[2]}")
        return flow
    
    def extract_spark_data_flow(self, content):
        """Extract data flow from Spark script"""
        flow = []
        # Extract DataFrame operations
        pattern = r"(\w+)\s*=\s*(\w+)\.[^=]+"
        matches = re.findall(pattern, content)
        for match in matches:
            flow.append(f"{match[0]} = {match[1]}.operation")
        return flow
    
    def extract_databricks_data_flow(self, content):
        """Extract data flow from Databricks notebook"""
        return self.extract_spark_data_flow(content)  # Same as Spark
    
    def create_mappings(self):
        """Create mappings between Hadoop and Databricks processes"""
        for hadoop_process in self.hadoop_processes:
            best_matches = []
            
            for databricks_process in self.databricks_processes:
                similarity_score = self.calculate_similarity(hadoop_process, databricks_process)
                
                if similarity_score > 0.3:  # Threshold for similarity
                    best_matches.append({
                        'databricks_process': databricks_process,
                        'similarity_score': similarity_score,
                        'matching_factors': self.get_matching_factors(hadoop_process, databricks_process)
                    })
            
            # Sort by similarity score
            best_matches.sort(key=lambda x: x['similarity_score'], reverse=True)
            
            if best_matches:
                mapping = {
                    'hadoop_process': hadoop_process,
                    'best_match': best_matches[0],
                    'all_matches': best_matches[:5]  # Top 5 matches
                }
                self.mappings.append(mapping)
    
    def calculate_similarity(self, hadoop_process, databricks_process):
        """Calculate similarity between Hadoop and Databricks processes"""
        score = 0.0
        
        # Filename similarity
        hadoop_name = hadoop_process['name'].lower()
        databricks_name = databricks_process['name'].lower()
        
        # Check for common keywords
        common_keywords = ['permid', 'policy', 'address', 'consumer', 'phone', 'merge', 'publish', 'validate', 'parse']
        for keyword in common_keywords:
            if keyword in hadoop_name and keyword in databricks_name:
                score += 0.2
        
        # Business logic similarity
        hadoop_logic = set(hadoop_process['business_logic'])
        databricks_logic = set(databricks_process['business_logic'])
        logic_intersection = hadoop_logic.intersection(databricks_logic)
        if hadoop_logic or databricks_logic:
            score += len(logic_intersection) / max(len(hadoop_logic), len(databricks_logic)) * 0.4
        
        # Data flow similarity
        if 'data_flow' in hadoop_process and 'data_flow' in databricks_process:
            hadoop_flow = set(hadoop_process['data_flow'])
            databricks_flow = set(databricks_process['data_flow'])
            flow_intersection = hadoop_flow.intersection(databricks_flow)
            if hadoop_flow or databricks_flow:
                score += len(flow_intersection) / max(len(hadoop_flow), len(databricks_flow)) * 0.2
        
        # Content similarity
        hadoop_content = hadoop_process['content_snippet'].lower()
        databricks_content = databricks_process['content_snippet'].lower()
        
        # Check for common patterns in content
        common_patterns = ['spark', 'dataframe', 'select', 'filter', 'join', 'groupby', 'agg']
        for pattern in common_patterns:
            if pattern in hadoop_content and pattern in databricks_content:
                score += 0.05
        
        return min(score, 1.0)  # Cap at 1.0
    
    def get_matching_factors(self, hadoop_process, databricks_process):
        """Get factors that contribute to the match"""
        factors = []
        
        # Filename similarity
        hadoop_name = hadoop_process['name'].lower()
        databricks_name = databricks_process['name'].lower()
        
        common_keywords = ['permid', 'policy', 'address', 'consumer', 'phone', 'merge', 'publish', 'validate', 'parse']
        matching_keywords = [kw for kw in common_keywords if kw in hadoop_name and kw in databricks_name]
        if matching_keywords:
            factors.append(f"Common keywords: {', '.join(matching_keywords)}")
        
        # Business logic similarity
        hadoop_logic = set(hadoop_process['business_logic'])
        databricks_logic = set(databricks_process['business_logic'])
        common_logic = hadoop_logic.intersection(databricks_logic)
        if common_logic:
            factors.append(f"Common business logic: {', '.join(common_logic)}")
        
        # Process type similarity
        if hadoop_process['type'] == 'PySpark Script' and databricks_process['type'] == 'Databricks Notebook':
            factors.append("Both use Spark/Python")
        elif hadoop_process['type'] == 'SQL Script' and databricks_process['type'] == 'SQL Script':
            factors.append("Both use SQL")
        
        return factors
    
    def create_excel_mapping(self, output_file, pair_name):
        """Create Excel file with process mappings"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create Process Mappings sheet
        self.create_process_mappings_sheet(wb, pair_name)
        
        # Create Hadoop Processes sheet
        self.create_hadoop_processes_sheet(wb)
        
        # Create Databricks Processes sheet
        self.create_databricks_processes_sheet(wb)
        
        # Create Similarity Analysis sheet
        self.create_similarity_analysis_sheet(wb)
        
        wb.save(output_file)
        print(f"  📊 Excel mapping file created: {output_file}")
    
    def create_process_mappings_sheet(self, wb, pair_name):
        """Create Process Mappings sheet"""
        ws = wb.create_sheet("Process_Mappings")
        
        # Add title
        ws['A1'] = f"Repository Process Mapping: {pair_name}"
        ws['A1'].font = Font(size=16, bold=True)
        
        headers = [
            "Hadoop Process", "Hadoop Type", "Hadoop Path", "Hadoop Business Logic",
            "Databricks Process", "Databricks Type", "Databricks Path", "Databricks Business Logic",
            "Similarity Score", "Matching Factors", "Confidence Level"
        ]
        ws.append(headers)
        
        for mapping in self.mappings:
            hadoop_process = mapping['hadoop_process']
            best_match = mapping['best_match']
            databricks_process = best_match['databricks_process']
            
            row = [
                hadoop_process['name'],
                hadoop_process['type'],
                hadoop_process['relative_path'],
                ', '.join(hadoop_process['business_logic']),
                databricks_process['name'],
                databricks_process['type'],
                databricks_process['relative_path'],
                ', '.join(databricks_process['business_logic']),
                f"{best_match['similarity_score']:.2f}",
                '; '.join(best_match['matching_factors']),
                self.get_confidence_level(best_match['similarity_score'])
            ]
            ws.append(row)
        
        self.format_sheet(ws)
    
    def create_hadoop_processes_sheet(self, wb):
        """Create Hadoop Processes sheet"""
        ws = wb.create_sheet("Hadoop_Processes")
        
        headers = [
            "Process Name", "Type", "Path", "Business Logic", "Data Flow", "Content Snippet"
        ]
        ws.append(headers)
        
        for process in self.hadoop_processes:
            row = [
                process['name'],
                process['type'],
                process['relative_path'],
                ', '.join(process['business_logic']),
                '; '.join(process.get('data_flow', [])),
                process['content_snippet']
            ]
            ws.append(row)
        
        self.format_sheet(ws)
    
    def create_databricks_processes_sheet(self, wb):
        """Create Databricks Processes sheet"""
        ws = wb.create_sheet("Databricks_Processes")
        
        headers = [
            "Process Name", "Type", "Path", "Business Logic", "Data Flow", "Content Snippet"
        ]
        ws.append(headers)
        
        for process in self.databricks_processes:
            row = [
                process['name'],
                process['type'],
                process['relative_path'],
                ', '.join(process['business_logic']),
                '; '.join(process.get('data_flow', [])),
                process['content_snippet']
            ]
            ws.append(row)
        
        self.format_sheet(ws)
    
    def create_similarity_analysis_sheet(self, wb):
        """Create Similarity Analysis sheet"""
        ws = wb.create_sheet("Similarity_Analysis")
        
        headers = [
            "Hadoop Process", "Databricks Process", "Similarity Score", "Matching Factors", "Recommendation"
        ]
        ws.append(headers)
        
        for mapping in self.mappings:
            hadoop_process = mapping['hadoop_process']
            
            for match in mapping['all_matches']:
                databricks_process = match['databricks_process']
                
                row = [
                    hadoop_process['name'],
                    databricks_process['name'],
                    f"{match['similarity_score']:.2f}",
                    '; '.join(match['matching_factors']),
                    self.get_recommendation(match['similarity_score'])
                ]
                ws.append(row)
        
        self.format_sheet(ws)
    
    def get_confidence_level(self, score):
        """Get confidence level based on similarity score"""
        if score >= 0.8:
            return "High"
        elif score >= 0.6:
            return "Medium"
        elif score >= 0.4:
            return "Low"
        else:
            return "Very Low"
    
    def get_recommendation(self, score):
        """Get recommendation based on similarity score"""
        if score >= 0.8:
            return "Strong Match - Likely Equivalent"
        elif score >= 0.6:
            return "Good Match - Probably Equivalent"
        elif score >= 0.4:
            return "Possible Match - Review Required"
        else:
            return "Weak Match - Manual Review"
    
    def format_sheet(self, ws):
        """Format Excel sheet with headers and styling - simplified to avoid MergedCell issues"""
        try:
            # Just set basic column widths without complex formatting
            for col_idx in range(1, ws.max_column + 1):
                try:
                    column_letter = ws.cell(row=1, column=col_idx).column_letter
                    ws.column_dimensions[column_letter].width = 20
                except:
                    pass
        except Exception as e:
            print(f"    ⚠️ Warning: Could not format sheet: {e}")

def main():
    """Main function to run the enhanced repository analyzer"""
    import sys
    
    print("=" * 80)
    print("🚀 ENHANCED REPOSITORY PROCESS MAPPING TOOL")
    print("=" * 80)
    print("This tool automatically pairs Hadoop and Azure Databricks repositories")
    print("and generates separate Excel mappings for each pair.")
    print()
    
    # Get base path from command line arguments or prompt
    if len(sys.argv) > 1:
        base_path = sys.argv[1]
    else:
        base_path = input("Enter the path containing all repositories: ").strip()
    
    if not base_path:
        print("Error: Repository path is required!")
        print("Usage: python repository_analyzer_enhanced.py <path_to_repositories>")
        print("Example: python repository_analyzer_enhanced.py OneDrive_1_7-25-2025/Hadoop")
        return
    
    if not os.path.exists(base_path):
        print(f"Error: Repository path does not exist: {base_path}")
        return
    
    # Initialize analyzer
    analyzer = RepositoryAnalyzer()
    
    # Find repository pairs
    print(f"\n🔍 Scanning for repository pairs in: {base_path}")
    pairs = analyzer.find_repository_pairs(base_path)
    
    if not pairs:
        print("❌ No repository pairs found!")
        return
    
    print(f"\n✅ Found {len(pairs)} repository pairs to analyze")
    
    # Analyze each pair and generate Excel files
    results = []
    
    for i, pair in enumerate(pairs, 1):
        print(f"\n{'='*80}")
        print(f"📊 PROCESSING PAIR {i}/{len(pairs)}: {pair['pair_name']}")
        print(f"{'='*80}")
        
        try:
            # Analyze the pair
            stats = analyzer.analyze_repository_pair(
                pair['hadoop_repo'], 
                pair['databricks_repo']
            )
            
            # Generate Excel file
            output_file = f"REPOSITORY_MAPPING_{pair['pair_name']}.xlsx"
            analyzer.create_excel_mapping(output_file, pair['pair_name'])
            
            results.append({
                'pair_name': pair['pair_name'],
                'hadoop_name': pair['hadoop_name'],
                'databricks_name': pair['databricks_name'],
                'output_file': output_file,
                'hadoop_processes': stats['hadoop_processes'],
                'databricks_processes': stats['databricks_processes'],
                'mappings': stats['mappings'],
                'status': 'Success'
            })
            
            print(f"✅ Completed: {pair['pair_name']}")
            
        except Exception as e:
            print(f"❌ Error processing {pair['pair_name']}: {e}")
            results.append({
                'pair_name': pair['pair_name'],
                'hadoop_name': pair['hadoop_name'],
                'databricks_name': pair['databricks_name'],
                'output_file': 'N/A',
                'hadoop_processes': 0,
                'databricks_processes': 0,
                'mappings': 0,
                'status': f'Error: {e}'
            })
    
    # Print final summary
    print(f"\n{'='*80}")
    print("🎯 FINAL SUMMARY")
    print(f"{'='*80}")
    
    successful = [r for r in results if r['status'] == 'Success']
    failed = [r for r in results if r['status'] != 'Success']
    
    print(f"✅ Successful: {len(successful)}")
    print(f"❌ Failed: {len(failed)}")
    
    if successful:
        print(f"\n📊 Generated Excel Files:")
        total_hadoop = sum(r['hadoop_processes'] for r in successful)
        total_databricks = sum(r['databricks_processes'] for r in successful)
        total_mappings = sum(r['mappings'] for r in successful)
        
        for result in successful:
            print(f"  📄 {result['output_file']}")
            print(f"     Hadoop: {result['hadoop_processes']} processes")
            print(f"     Databricks: {result['databricks_processes']} processes")
            print(f"     Mappings: {result['mappings']} matches")
        
        print(f"\n📈 TOTALS:")
        print(f"   Total Hadoop Processes: {total_hadoop}")
        print(f"   Total Databricks Processes: {total_databricks}")
        print(f"   Total Mappings: {total_mappings}")
    
    if failed:
        print(f"\n⚠️ Failed Analyses:")
        for result in failed:
            print(f"   ❌ {result['pair_name']}: {result['status']}")
    
    print(f"\n🎉 Analysis complete! Check the generated Excel files for detailed mappings.")

if __name__ == "__main__":
    main()
