"""
Fast STAG Comparison Generator - Uses VM_FAWN Excel Directly
============================================================

Instead of AI-analyzing each component (slow, rate limits), this reads
VM_FAWN Excel which already has all transformation logic extracted.

For Ab Initio ↔ Databricks comparison:
1. Read VM_FAWN Excel Component&Fields (has all transform rules)
2. Read Databricks parsed data (workflows, notebooks)
3. Use AI to find semantic matches (not per-component, but batched)
4. Generate comparison Excel with red highlights for unmatched logic

Speed: Minutes instead of hours!
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
from typing import Dict, List, Any, Optional
import json
from datetime import datetime
from loguru import logger


class FastSTAGComparison:
    """Fast STAG comparison using VM_FAWN Excel data directly"""

    def __init__(self, ai_analyzer=None):
        """
        Args:
            ai_analyzer: Optional AI analyzer for semantic matching
        """
        self.ai_analyzer = ai_analyzer
        self.red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    def generate_abinitio_databricks_comparison(
        self,
        abinitio_vm_fawn_excel: str,
        databricks_parsed_json: str,
        output_excel: str
    ) -> Dict[str, Any]:
        """
        Generate STAG comparison between Ab Initio and Databricks

        Args:
            abinitio_vm_fawn_excel: Path to VM_FAWN generated Excel
            databricks_parsed_json: Path to Databricks parsed JSON
            output_excel: Output Excel path

        Returns:
            Dict with comparison statistics
        """
        logger.info(f"🚀 Starting Fast STAG Comparison")
        logger.info(f"  Ab Initio Excel: {abinitio_vm_fawn_excel}")
        logger.info(f"  Databricks JSON: {databricks_parsed_json}")

        # Step 1: Read Ab Initio VM_FAWN Excel
        logger.info("📊 Step 1: Reading Ab Initio VM_FAWN Excel...")
        abinitio_components = self._read_vm_fawn_excel(abinitio_vm_fawn_excel)
        logger.info(f"  Found {len(abinitio_components)} Ab Initio components")

        # Step 2: Read Databricks data
        logger.info("📊 Step 2: Reading Databricks data...")
        databricks_logic = self._read_databricks_json(databricks_parsed_json)
        logger.info(f"  Found {len(databricks_logic)} Databricks transformations")

        # Step 3: Group Ab Initio logic by type (for batched AI analysis)
        logger.info("📊 Step 3: Grouping Ab Initio logic by type...")
        abinitio_grouped = self._group_by_transform_type(abinitio_components)
        logger.info(f"  Grouped into {len(abinitio_grouped)} unique transform types")

        # Step 4: Find matches using AI (batched, not per-component!)
        logger.info("📊 Step 4: Finding semantic matches (batched AI analysis)...")
        matches = self._find_semantic_matches_batched(
            abinitio_grouped,
            databricks_logic
        )
        logger.info(f"  Found {matches['matched_count']} matches")
        logger.info(f"  Unmatched Ab Initio logic: {matches['unmatched_count']}")

        # Step 5: Generate comparison Excel
        logger.info("📊 Step 5: Generating comparison Excel...")
        self._generate_comparison_excel(
            abinitio_components,
            databricks_logic,
            matches,
            output_excel
        )

        logger.info(f"✅ STAG Comparison complete: {output_excel}")

        return {
            'success': True,
            'output_file': output_excel,
            'abinitio_components': len(abinitio_components),
            'databricks_transformations': len(databricks_logic),
            'matched': matches['matched_count'],
            'unmatched': matches['unmatched_count'],
            'match_rate': matches['matched_count'] / len(abinitio_components) * 100 if abinitio_components else 0
        }

    def _read_vm_fawn_excel(self, excel_path: str) -> List[Dict[str, Any]]:
        """
        Read VM_FAWN Excel Component&Fields sheet

        Returns:
            List of component dicts with transform_rule, keys, etc.
        """
        df = pd.read_excel(excel_path, sheet_name='Component&Fields')

        components = []
        for idx, row in df.iterrows():
            component = {
                'id': f"comp_{idx}",
                'name': row.get('Component Name', ''),
                'type': row.get('Transform Component', ''),
                'transform_rule': row.get('Transform Rule', ''),
                'location': row.get('TransLocation', ''),
                'keys': row.get('Keys', ''),
                'subgraph': row.get('subgraph_hierarchy', ''),
                'phase': row.get('component_phase', ''),
                'matched': False,  # Will be updated
                'databricks_match': None
            }
            components.append(component)

        return components

    def _read_databricks_json(self, json_path: str) -> List[Dict[str, Any]]:
        """
        Read Databricks parsed JSON

        Returns:
            List of transformation dicts
        """
        with open(json_path, 'r') as f:
            data = json.load(f)

        # Extract transformations from Databricks data
        transformations = []

        # Assuming structure: {notebooks: [...], workflows: [...]}
        notebooks = data.get('notebooks', [])
        for notebook in notebooks:
            trans = {
                'name': notebook.get('name', ''),
                'type': 'notebook',
                'logic': notebook.get('content', ''),
                'path': notebook.get('path', ''),
                'language': notebook.get('language', '')
            }
            transformations.append(trans)

        workflows = data.get('workflows', [])
        for workflow in workflows:
            trans = {
                'name': workflow.get('name', ''),
                'type': 'workflow',
                'logic': workflow.get('tasks', []),
                'path': workflow.get('path', '')
            }
            transformations.append(trans)

        return transformations

    def _group_by_transform_type(self, components: List[Dict]) -> Dict[str, List[Dict]]:
        """Group components by transform type for batched analysis"""
        grouped = {}

        for comp in components:
            comp_type = comp['type']
            if comp_type not in grouped:
                grouped[comp_type] = []
            grouped[comp_type].append(comp)

        return grouped

    def _find_semantic_matches_batched(
        self,
        abinitio_grouped: Dict[str, List[Dict]],
        databricks_logic: List[Dict]
    ) -> Dict[str, Any]:
        """
        Find semantic matches between Ab Initio and Databricks using batched AI analysis

        Instead of analyzing each component individually (slow):
        1. Analyze each Ab Initio transform TYPE once
        2. Match against Databricks logic
        3. Apply matches to all components of that type

        Speed: ~20-30 AI calls instead of 1000+
        """
        matches = {
            'matched_count': 0,
            'unmatched_count': 0,
            'matches': []
        }

        if not self.ai_analyzer or not self.ai_analyzer.enabled:
            logger.warning("AI analyzer not available - using basic keyword matching")
            return self._find_keyword_matches(abinitio_grouped, databricks_logic)

        # For each Ab Initio transform type
        for transform_type, components in abinitio_grouped.items():
            logger.info(f"  Analyzing transform type: {transform_type} ({len(components)} instances)")

            # Get sample transform rules for this type (not all!)
            sample_rules = list(set([c['transform_rule'] for c in components[:5] if c['transform_rule']]))

            if not sample_rules:
                # No transform rules, mark as unmatched
                matches['unmatched_count'] += len(components)
                continue

            # Ask AI to find Databricks equivalent (ONE call per type!)
            try:
                ai_match = self._ai_find_databricks_match(
                    transform_type,
                    sample_rules,
                    databricks_logic
                )

                if ai_match:
                    # Apply match to ALL components of this type
                    for comp in components:
                        comp['matched'] = True
                        comp['databricks_match'] = ai_match
                        matches['matched_count'] += 1

                    matches['matches'].append({
                        'abinitio_type': transform_type,
                        'databricks_match': ai_match,
                        'component_count': len(components)
                    })
                else:
                    # No match found
                    matches['unmatched_count'] += len(components)

            except Exception as e:
                logger.error(f"AI matching error for {transform_type}: {e}")
                matches['unmatched_count'] += len(components)

        return matches

    def _ai_find_databricks_match(
        self,
        transform_type: str,
        sample_rules: List[str],
        databricks_logic: List[Dict]
    ) -> Optional[Dict]:
        """
        Use AI to find Databricks equivalent for Ab Initio transform type

        Args:
            transform_type: Ab Initio component type (e.g., "Reformat", "Join")
            sample_rules: Sample transformation rules for this type
            databricks_logic: List of Databricks transformations

        Returns:
            Dict with Databricks match details, or None if no match
        """
        # Prepare context for AI
        databricks_summary = []
        for idx, db in enumerate(databricks_logic[:20], 1):  # Limit context
            summary = f"{idx}. {db['name']} ({db['type']})"
            databricks_summary.append(summary)

        # Ask AI
        prompt = f"""Ab Initio Transform Type: {transform_type}

Sample Transformation Rules:
{chr(10).join(sample_rules[:3])}

Available Databricks Transformations:
{chr(10).join(databricks_summary)}

Question: Which Databricks transformation(s) perform similar logic to this Ab Initio {transform_type} component?

Respond with JSON:
{{
  "has_match": true/false,
  "databricks_name": "name of matching Databricks transformation",
  "databricks_type": "notebook/workflow",
  "confidence": 0-100,
  "reasoning": "why these match"
}}
"""

        try:
            response = self.ai_analyzer.analyze_with_context(
                query=prompt,
                context=""
            )

            # Parse AI response
            import json
            match_data = json.loads(response.get('analysis', '{}'))

            if match_data.get('has_match') and match_data.get('confidence', 0) > 50:
                return {
                    'databricks_name': match_data['databricks_name'],
                    'databricks_type': match_data['databricks_type'],
                    'confidence': match_data['confidence'],
                    'reasoning': match_data['reasoning']
                }

        except Exception as e:
            logger.error(f"AI parsing error: {e}")

        return None

    def _find_keyword_matches(
        self,
        abinitio_grouped: Dict[str, List[Dict]],
        databricks_logic: List[Dict]
    ) -> Dict[str, Any]:
        """Fallback: keyword-based matching without AI"""
        matches = {
            'matched_count': 0,
            'unmatched_count': 0,
            'matches': []
        }

        # Simple keyword matching
        keyword_map = {
            'Reformat': ['transform', 'select', 'map'],
            'Join': ['join', 'merge'],
            'Filter': ['filter', 'where'],
            'Sort': ['sort', 'order by'],
            'Aggregate': ['group by', 'aggregate', 'sum', 'count']
        }

        for transform_type, components in abinitio_grouped.items():
            keywords = keyword_map.get(transform_type, [])

            # Search Databricks for matching keywords
            db_matches = []
            for db in databricks_logic:
                db_logic = str(db.get('logic', '')).lower()
                if any(kw in db_logic for kw in keywords):
                    db_matches.append(db)

            if db_matches:
                for comp in components:
                    comp['matched'] = True
                    comp['databricks_match'] = {
                        'databricks_name': db_matches[0]['name'],
                        'databricks_type': db_matches[0]['type'],
                        'confidence': 60,
                        'reasoning': f"Keyword match: {keywords}"
                    }
                matches['matched_count'] += len(components)
            else:
                matches['unmatched_count'] += len(components)

        return matches

    def _generate_comparison_excel(
        self,
        abinitio_components: List[Dict],
        databricks_logic: List[Dict],
        matches: Dict,
        output_excel: str
    ):
        """
        Generate comparison Excel with RED highlighting for unmatched Ab Initio logic

        Sheet structure:
        - Ab Initio Component | Transform Rule | Databricks Match | Status
        """
        # Create DataFrame
        rows = []
        for comp in abinitio_components:
            row = {
                'Ab Initio Component': comp['name'],
                'Type': comp['type'],
                'Transform Rule': comp['transform_rule'][:200] if comp['transform_rule'] else '',  # Truncate
                'Keys': comp['keys'],
                'Databricks Match': comp['databricks_match']['databricks_name'] if comp['matched'] else 'NO MATCH',
                'Match Confidence': comp['databricks_match']['confidence'] if comp['matched'] else 0,
                'Status': 'Matched' if comp['matched'] else 'UNMATCHED'
            }
            rows.append(row)

        df = pd.DataFrame(rows)

        # Write to Excel
        df.to_excel(output_excel, index=False, sheet_name='STAG Comparison')

        # Apply formatting (red for unmatched)
        wb = openpyxl.load_workbook(output_excel)
        ws = wb['STAG Comparison']

        # Header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data rows - RED for unmatched
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            status = ws.cell(row_idx, 7).value  # Status column

            if status == 'UNMATCHED':
                # Highlight entire row in RED
                for cell in row:
                    cell.fill = self.red_fill
                    cell.font = Font(color="CC0000", bold=True)
            else:
                # Highlight matched in GREEN
                for cell in row:
                    cell.fill = self.green_fill

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(output_excel)
        logger.info(f"✅ Comparison Excel saved with formatting")


# CLI usage
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 4:
        print("Usage: python fast_stag_comparison.py <vm_fawn_excel> <databricks_json> <output_excel>")
        print("\nExample:")
        print("  python fast_stag_comparison.py \\")
        print("    'Input Files/VM_FAWN/bbi_preprocessing_output/1500_CDD_*_auto.xlsx' \\")
        print("    'databricks_parsed.json' \\")
        print("    'outputs/1500_stag_comparison.xlsx'")
        sys.exit(1)

    vm_fawn_excel = sys.argv[1]
    databricks_json = sys.argv[2]
    output_excel = sys.argv[3]

    generator = FastSTAGComparison()
    result = generator.generate_abinitio_databricks_comparison(
        vm_fawn_excel,
        databricks_json,
        output_excel
    )

    print(f"\n✅ STAG Comparison Complete!")
    print(f"  Output: {result['output_file']}")
    print(f"  Ab Initio components: {result['abinitio_components']}")
    print(f"  Matched: {result['matched']} ({result['match_rate']:.1f}%)")
    print(f"  Unmatched (RED): {result['unmatched']}")
