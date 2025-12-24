"""
Excel Generator for STAG Comparison Reports

Generates professional Excel workbooks with 5 sheets matching the exact
format specification for system comparison reports.

Sheets:
1. Overview - Business stage comparison
2. Databricks Logic - Detailed Databricks implementation
3. [Source] Logic - Detailed source system implementation
4. Logic Comparison - Side-by-side technical comparison
5. STTM - Source-to-Target Mapping (column-level)

Features:
- Professional formatting (bold headers, colors, borders)
- Auto-adjusted column widths
- Code snippet formatting (monospace)
- Wrapped text for readability
- Frozen header rows
"""

import os
from pathlib import Path
from typing import Dict, List, Any, Optional
from loguru import logger
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel generation will be limited")


class ExcelGenerator:
    """Generate professional Excel comparison reports"""

    def __init__(self):
        """Initialize Excel Generator"""
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl is required for Excel generation")
            raise ImportError("Please install openpyxl: pip install openpyxl")

        # Define color scheme
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.similar_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.different_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_comparison_excel(
        self,
        source_system: str,
        source_workflow: str,
        databricks_pipeline: str,
        business_stages: List[Dict[str, Any]],
        source_logic: Dict[str, Any],
        databricks_logic: Dict[str, Any],
        sttm_mappings: List[Dict[str, Any]],
        output_folder: str = "outputs/stag_comparisons"
    ) -> str:
        """
        Generate complete comparison Excel file

        Args:
            source_system: "hadoop" or "abinitio"
            source_workflow: Source workflow name
            databricks_pipeline: Databricks pipeline name
            business_stages: Business stage comparison data
            source_logic: Source system logic
            databricks_logic: Databricks logic
            sttm_mappings: STTM column mappings
            output_folder: Output directory

        Returns:
            Path to generated Excel file
        """
        logger.info(f"📊 Generating Excel comparison: {source_workflow} → {databricks_pipeline}")

        # Create output folder
        os.makedirs(output_folder, exist_ok=True)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate each sheet
        logger.info("   Creating Overview sheet...")
        self._create_overview_sheet(wb, source_system, source_workflow, databricks_pipeline, business_stages)

        logger.info("   Creating Databricks Logic sheet...")
        self._create_databricks_logic_sheet(wb, databricks_logic)

        logger.info(f"   Creating {source_system.capitalize()} Logic sheet...")
        self._create_source_logic_sheet(wb, source_system, source_logic)

        logger.info("   Creating Logic Comparison sheet...")
        self._create_logic_comparison_sheet(wb, source_system, source_logic, databricks_logic)

        logger.info("   Creating STTM sheets (Source, Target, Comparison)...")
        self._create_sttm_sheets(wb, sttm_mappings, source_system, source_logic, databricks_logic)

        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Sanitize filenames to remove invalid characters for Windows/Linux
        safe_source = source_workflow.replace(':', '_').replace('/', '_').replace('\\', '_').replace('|', '_').replace('?', '_').replace('*', '_').replace('<', '_').replace('>', '_').replace('"', '_').strip()
        safe_databricks = databricks_pipeline.replace(':', '_').replace('/', '_').replace('\\', '_').replace('|', '_').replace('?', '_').replace('*', '_').replace('<', '_').replace('>', '_').replace('"', '_').strip()

        filename = f"{safe_source}_vs_{safe_databricks}_{timestamp}.xlsx"
        output_path = os.path.join(output_folder, filename)

        # Ensure output folder exists
        os.makedirs(output_folder, exist_ok=True)

        try:
            wb.save(output_path)
            logger.info(f"✅ Excel file saved: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"❌ Failed to save Excel file: {e}")
            # Try with even simpler filename
            simple_filename = f"STAG_comparison_{timestamp}.xlsx"
            fallback_path = os.path.join(output_folder, simple_filename)
            wb.save(fallback_path)
            logger.info(f"✅ Excel file saved (fallback): {fallback_path}")
            return fallback_path

    def generate_databricks_documentation_excel(
        self,
        pipeline_name: str,
        databricks_logic: Dict[str, Any],
        business_stages: List[Dict[str, Any]],
        output_folder: str = "outputs/databricks_documentation",
        timestamp: str = None
    ) -> str:
        """
        Generate single-system Databricks documentation Excel (no comparison)

        Args:
            pipeline_name: Databricks pipeline name
            databricks_logic: Databricks logic extraction result
            business_stages: Business stage abstractions
            output_folder: Output directory
            timestamp: Optional timestamp for filename

        Returns:
            Path to generated Excel file
        """
        logger.info(f"📊 Generating Databricks documentation: {pipeline_name}")

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Overview (business stages)
        logger.info("   Creating Overview sheet...")
        ws_overview = wb.create_sheet("Overview")
        self._create_databricks_overview_sheet(ws_overview, pipeline_name, business_stages)

        # Sheet 2: Logic (detailed activities) - use existing method which creates its own sheet
        logger.info("   Creating Logic sheet...")
        self._create_databricks_logic_sheet(wb, databricks_logic)

        # Sheet 3: STTM (column mappings from activities)
        logger.info("   Creating STTM sheet...")
        ws_sttm = wb.create_sheet("STTM")
        self._create_databricks_sttm_sheet(ws_sttm, databricks_logic, pipeline_name)

        # Save workbook
        if not timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Clean pipeline name for filename
        safe_pipeline = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in pipeline_name)[:50]
        filename = f"Databricks_Documentation_{safe_pipeline}_{timestamp}.xlsx"
        output_path = os.path.join(output_folder, filename)

        os.makedirs(output_folder, exist_ok=True)

        try:
            wb.save(output_path)
            logger.info(f"✅ Databricks documentation Excel saved: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"❌ Failed to save Excel: {e}")
            fallback_path = os.path.join(output_folder, f"Databricks_Doc_{timestamp}.xlsx")
            wb.save(fallback_path)
            return fallback_path

    def _create_databricks_overview_sheet(
        self,
        ws,
        pipeline_name: str,
        business_stages: List[Dict[str, Any]]
    ):
        """Create overview sheet for single-system Databricks documentation"""
        # Header
        ws.append(["Databricks Pipeline Documentation"])
        ws.merge_cells('A1:D1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.append([])
        ws.append(["Pipeline Name:", pipeline_name])
        ws['B3'].font = Font(bold=True)

        ws.append([])
        ws.append(["Business Stage", "Description", "Activities", "Purpose"])

        # Format header row
        for col in range(1, 5):
            cell = ws.cell(row=5, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Add business stages
        for stage in business_stages:
            ws.append([
                stage.get('stage_name', stage.get('stage', '')),
                stage.get('databricks_description', stage.get('description', '')),
                stage.get('databricks_description', '')[:50] + '...' if len(stage.get('databricks_description', '')) > 50 else stage.get('databricks_description', ''),
                stage.get('notes', stage.get('business_purpose', ''))
            ])

        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

        ws.freeze_panes = 'A6'

    def _create_databricks_sttm_sheet(
        self,
        ws,
        databricks_logic: Dict[str, Any],
        pipeline_name: str
    ):
        """Create STTM sheet for Databricks-only documentation with detailed column mappings"""
        # Header
        ws.append([f"Databricks Pipeline STTM: {pipeline_name}"])
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.append([])
        # Standard STTM columns as per user specification
        headers = [
            "Id", "Processing Order", "Schema", "Source Dataset Name",
            "Source Field Name", "Target Table/File Name", "Target Field Name",
            "Target Field Data Type", "pk?", "contains_pii", "Field Type",
            "Field Depends On", "Pre Processing Rules", "Field Definition"
        ]
        ws.append(headers)

        # Format header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Extract STTM from activities
        activities = databricks_logic.get('activities', [])
        row_count = 0
        field_id = 1

        for act_idx, activity in enumerate(activities, 1):
            # Safety check: ensure activity is a dict
            if not isinstance(activity, dict):
                logger.warning(f"   ⚠ Activity {act_idx} is not a dict, skipping: {activity}")
                continue

            activity_name = activity.get('name', 'Unknown')

            # Get notebook path from activity (NOT from details)
            notebook_path = activity.get('notebook', activity.get('path', 'N/A'))

            # PRIORITY 1: Use AI-extracted column lineage (most accurate)
            ai_column_lineage = activity.get('ai_column_lineage', [])

            if ai_column_lineage:
                # Use AI-extracted column lineage (like reference Excel format)
                logger.info(f"      Using AI column lineage for {activity_name}: {len(ai_column_lineage)} mappings")
                for mapping in ai_column_lineage:
                    source_table = mapping.get('source_table', 'Unknown')
                    source_column = mapping.get('source_column', 'Unknown')
                    target_table = mapping.get('target_table', 'Unknown')
                    target_column = mapping.get('target_column', 'Unknown')
                    transformation = mapping.get('transformation', '')
                    data_type = mapping.get('data_type', 'Unknown')
                    is_derived = mapping.get('is_derived', False)

                    # Extract schema from table name if available
                    if '.' in source_table:
                        schema = source_table.split('.')[0]
                    else:
                        schema = 'default'

                    ws.append([
                        field_id,  # Id
                        act_idx,  # Processing Order
                        schema,  # Schema
                        source_table,  # Source Dataset Name
                        source_column,  # Source Field Name
                        target_table,  # Target Table/File Name
                        target_column,  # Target Field Name
                        data_type,  # Target Field Data Type
                        '',  # pk?
                        '',  # contains_pii
                        'Derived' if is_derived else 'Direct',  # Field Type
                        '',  # Field Depends On
                        transformation,  # Pre Processing Rules
                        f'Activity: {activity_name}'  # Field Definition
                    ])
                    field_id += 1
                    row_count += 1

            else:
                # PRIORITY 2: Fall back to structural column schemas
                column_schemas = activity.get('column_schemas', [])
                input_tables = activity.get('inputs', [])
                output_tables = activity.get('outputs', [])

                if column_schemas:
                    # Activity has detailed column-level schema information
                    for schema in column_schemas:
                        table_name = schema.get('table_name', 'Unknown')
                        columns = schema.get('columns', [])
                        is_output = schema.get('is_output', False)

                        # Iterate through each column
                        for col in columns:
                            if isinstance(col, dict):
                                col_name = col.get('name', 'Unknown')
                                col_type = col.get('type', col.get('data_type', 'Unknown'))
                            else:
                                col_name = str(col)
                                col_type = 'Unknown'

                            # Create STTM row with all 14 columns
                            if is_output:
                                # Output table - this is a target
                                ws.append([
                                    field_id,  # Id
                                    act_idx,  # Processing Order
                                    table_name.split('.')[0] if '.' in table_name else 'default',  # Schema
                                    ', '.join(input_tables[:2]) if input_tables else 'Multiple',  # Source Dataset
                                    col_name,  # Source Field Name (assumed same in transform)
                                    table_name,  # Target Table/File Name
                                    col_name,  # Target Field Name
                                    col_type,  # Target Field Data Type
                                    '',  # pk?
                                    '',  # contains_pii
                                    'Derived',  # Field Type
                                    '',  # Field Depends On
                                    f'Transformed via {notebook_path}',  # Pre Processing Rules
                                    f'Field from activity: {activity_name}'  # Field Definition
                                ])
                            else:
                                # Input table - this is a source
                                ws.append([
                                    field_id,  # Id
                                    act_idx,  # Processing Order
                                    table_name.split('.')[0] if '.' in table_name else 'default',  # Schema
                                    table_name,  # Source Dataset Name
                                    col_name,  # Source Field Name
                                    ', '.join(output_tables[:2]) if output_tables else 'TBD',  # Target Table
                                    col_name,  # Target Field Name (assumed same)
                                    col_type,  # Target Field Data Type
                                    '',  # pk?
                                    '',  # contains_pii
                                    'Source',  # Field Type
                                    '',  # Field Depends On
                                    f'Loaded from {table_name}',  # Pre Processing Rules
                                    f'Source field from {activity_name}'  # Field Definition
                                ])

                            field_id += 1
                            row_count += 1
                elif input_tables or output_tables:
                    # Table-level mapping only (no column details)
                    for in_table in (input_tables if input_tables else ['N/A']):
                        for out_table in (output_tables if output_tables else ['N/A']):
                            ws.append([
                                field_id,  # Id
                                act_idx,  # Processing Order
                                in_table.split('.')[0] if '.' in in_table else 'default',  # Schema
                                in_table,  # Source Dataset Name
                                '*',  # Source Field Name (all fields)
                                out_table,  # Target Table/File Name
                                '*',  # Target Field Name (all fields)
                                'Various',  # Target Field Data Type
                                '',  # pk?
                                '',  # contains_pii
                                'Bulk Transform',  # Field Type
                                '',  # Field Depends On
                                f'Notebook: {notebook_path}',  # Pre Processing Rules
                                f'Activity: {activity_name}'  # Field Definition
                            ])
                            field_id += 1
                            row_count += 1

        # If no mappings found, add informative message
        if row_count == 0:
            ws.append([
                "No STTM data available",
                "Run with schema analysis to extract detailed column mappings",
                "",
                "",
                "",
                ""
            ])
            ws.merge_cells(f'A4:F4')
            ws['A4'].alignment = Alignment(horizontal='center')
            ws['A4'].font = Font(italic=True, color="666666")

        # Auto-adjust columns for all 14 STTM columns
        column_widths = [8, 15, 12, 30, 25, 30, 25, 20, 8, 12, 15, 20, 30, 30]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = 'A4'

    def _create_overview_sheet(
        self,
        wb: Workbook,
        source_system: str,
        source_workflow: str,
        databricks_pipeline: str,
        business_stages: List[Dict[str, Any]]
    ):
        """Create Overview sheet with business stage comparison"""
        ws = wb.create_sheet("Overview", 0)

        # Title row
        ws['A1'] = f"Workflow Comparison: {source_workflow} → {databricks_pipeline}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')

        # Metadata
        ws['A2'] = f"Source System: {source_system.capitalize()}"
        ws['A3'] = f"Target System: Databricks"
        ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Headers (row 6)
        headers = ['Stage Name', 'Databricks Description', f'{source_system.capitalize()} Description', 'Comparison', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # Data rows
        for row_idx, stage in enumerate(business_stages, 7):
            ws.cell(row=row_idx, column=1, value=stage.get('stage_name', ''))
            ws.cell(row=row_idx, column=2, value=stage.get('databricks_description', ''))
            ws.cell(row=row_idx, column=3, value=stage.get('source_description', ''))
            ws.cell(row=row_idx, column=4, value=stage.get('comparison', ''))
            ws.cell(row=row_idx, column=5, value=stage.get('notes', ''))

            # Apply coloring based on comparison
            comparison = stage.get('comparison', '')
            if comparison == 'Similar':
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = self.similar_fill
            elif comparison == 'Different':
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = self.different_fill

            # Apply borders and alignment
            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 35

        # Freeze header row
        ws.freeze_panes = 'A7'

    def _create_databricks_logic_sheet(self, wb: Workbook, databricks_logic: Dict[str, Any]):
        """Create Databricks Logic sheet with enhanced step-by-step logic and code snippets"""
        ws = wb.create_sheet("Databricks Logic")

        # Title
        ws['A1'] = f"Databricks Pipeline: {databricks_logic.get('pipeline_name', 'Unknown')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Headers (4 columns as per manual format)
        headers = ['Pipeline Step/Activity', 'Notebook', 'Purpose', 'Key Code Snippet / Logic']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # Data rows with step-by-step logic and code snippets
        current_row = 4
        activities = databricks_logic.get('activities', [])

        for activity in activities:
            activity_name = activity.get('name', '')
            notebook = activity.get('notebook', '')
            purpose = activity.get('purpose', '')

            # Main activity row
            ws.cell(row=current_row, column=1, value=activity_name)
            ws.cell(row=current_row, column=2, value=notebook)
            ws.cell(row=current_row, column=3, value=purpose)
            ws.cell(row=current_row, column=4, value='')

            # Bold activity name
            ws.cell(row=current_row, column=1).font = Font(bold=True)

            # Apply borders
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).border = self.border
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)

            current_row += 1

            # Add step-by-step logic items (if available)
            step_by_step = activity.get('step_by_step_logic', [])
            if step_by_step:
                for i, step in enumerate(step_by_step, 1):
                    ws.cell(row=current_row, column=1, value=f"  Step {i}")
                    ws.cell(row=current_row, column=2, value='')
                    ws.cell(row=current_row, column=3, value='')
                    ws.cell(row=current_row, column=4, value=step)

                    # Indent step
                    ws.cell(row=current_row, column=1).alignment = Alignment(indent=2, vertical='top', wrap_text=True)

                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).border = self.border
                        ws.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)

                    current_row += 1

            # Add code snippets (if available)
            code_snippets = activity.get('code_snippets', [])
            if code_snippets:
                for snippet in code_snippets:
                    # Extract code text from snippet dict or use string directly
                    if isinstance(snippet, dict):
                        code_text = snippet.get('code', '') or snippet.get('snippet', '')
                    else:
                        code_text = str(snippet)

                    if code_text:
                        ws.cell(row=current_row, column=1, value='')
                        ws.cell(row=current_row, column=2, value='')
                        ws.cell(row=current_row, column=3, value='')
                        ws.cell(row=current_row, column=4, value=code_text)

                        # Format as code (monospace, wrapped)
                        ws.cell(row=current_row, column=4).font = Font(name='Courier New', size=9)
                        ws.cell(row=current_row, column=4).alignment = Alignment(wrap_text=True, vertical='top')

                        for col in range(1, 5):
                            ws.cell(row=current_row, column=col).border = self.border

                        current_row += 1

            # Add blank row between activities
            current_row += 1

        # Column widths (4-column format)
        ws.column_dimensions['A'].width = 30   # Pipeline Step/Activity
        ws.column_dimensions['B'].width = 30   # Notebook
        ws.column_dimensions['C'].width = 40   # Purpose
        ws.column_dimensions['D'].width = 60   # Key Code Snippet / Logic

        ws.freeze_panes = 'A4'

    def _create_source_logic_sheet(self, wb: Workbook, source_system: str, source_logic: Dict[str, Any]):
        """Create Source Logic sheet (Hadoop or Ab Initio)"""
        sheet_name = f"{source_system.capitalize()} Logic"
        ws = wb.create_sheet(sheet_name)

        # Title
        workflow_name = source_logic.get('workflow_name') or source_logic.get('graph_name', 'Unknown')
        ws['A1'] = f"{source_system.capitalize()} Workflow: {workflow_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        if source_system == 'hadoop':
            self._populate_hadoop_logic(ws, source_logic)
        elif source_system == 'abinitio':
            self._populate_abinitio_logic(ws, source_logic)

        ws.freeze_panes = 'A4'

    def _populate_hadoop_logic(self, ws, hadoop_logic: Dict[str, Any]):
        """Populate Hadoop-specific logic with step-by-step breakdown"""
        # Headers (4 columns to match Databricks format)
        headers = ['Step', 'Script', 'Purpose', 'Logic']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # Data rows with step-by-step logic
        current_row = 4
        jobs = hadoop_logic.get('jobs', [])

        for job in jobs:
            job_name = job.get('name', '')
            script = job.get('script', '')
            purpose = job.get('purpose', '')

            # Main job row
            ws.cell(row=current_row, column=1, value=job_name)
            ws.cell(row=current_row, column=2, value=script)
            ws.cell(row=current_row, column=3, value=purpose)
            ws.cell(row=current_row, column=4, value='')

            # Bold job name
            ws.cell(row=current_row, column=1).font = Font(bold=True)

            # Apply borders
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).border = self.border
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)

            current_row += 1

            # Add step-by-step logic items
            step_by_step = job.get('step_by_step_logic', [])
            if step_by_step:
                for i, step in enumerate(step_by_step, 1):
                    ws.cell(row=current_row, column=1, value=f"  {i}")
                    ws.cell(row=current_row, column=2, value='')
                    ws.cell(row=current_row, column=3, value='')
                    ws.cell(row=current_row, column=4, value=step)

                    # Indent step number
                    ws.cell(row=current_row, column=1).alignment = Alignment(indent=2, vertical='top')

                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).border = self.border
                        ws.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)

                    current_row += 1

            # Add code snippets
            code_snippets = job.get('code_snippets', [])
            if code_snippets:
                for snippet in code_snippets:
                    # Extract code text
                    if isinstance(snippet, dict):
                        code_text = snippet.get('code', '') or snippet.get('snippet', '')
                    else:
                        code_text = str(snippet)

                    if code_text:
                        ws.cell(row=current_row, column=1, value='')
                        ws.cell(row=current_row, column=2, value='')
                        ws.cell(row=current_row, column=3, value='')
                        ws.cell(row=current_row, column=4, value=code_text)

                        # Format as code
                        ws.cell(row=current_row, column=4).font = Font(name='Courier New', size=9)
                        ws.cell(row=current_row, column=4).alignment = Alignment(wrap_text=True, vertical='top')

                        for col in range(1, 5):
                            ws.cell(row=current_row, column=col).border = self.border

                        current_row += 1

            # Blank row between jobs
            current_row += 1

        # Column widths (4-column format)
        ws.column_dimensions['A'].width = 30   # Step
        ws.column_dimensions['B'].width = 30   # Script
        ws.column_dimensions['C'].width = 40   # Purpose
        ws.column_dimensions['D'].width = 60   # Logic

    def _populate_abinitio_logic(self, ws, abinitio_logic: Dict[str, Any]):
        """Populate Ab Initio-specific logic with step-by-step breakdown"""
        # Headers (4 columns to match format)
        headers = ['Step', 'Component', 'Type', 'Transformation Logic']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # Data rows with detailed transformation logic
        current_row = 4
        steps = abinitio_logic.get('steps', [])

        for step in steps:
            # Handle both dict and string formats
            if isinstance(step, dict):
                step_num = step.get('step_number', '')
                component = step.get('dataset', '') or step.get('component', '')
                comp_type = step.get('component_type', '')
                transformation = step.get('transformation_rules', '')
                step_by_step = step.get('step_by_step_logic', [])
            elif isinstance(step, str):
                step_num = ''
                component = step
                comp_type = ''
                transformation = ''
                step_by_step = []
            else:
                step_num = ''
                component = 'Unknown step'
                comp_type = ''
                transformation = ''
                step_by_step = []

            # Main component row
            ws.cell(row=current_row, column=1, value=step_num)
            ws.cell(row=current_row, column=2, value=component)
            ws.cell(row=current_row, column=3, value=comp_type)
            ws.cell(row=current_row, column=4, value=transformation)

            # Bold component name
            ws.cell(row=current_row, column=2).font = Font(bold=True)

            # Apply borders
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).border = self.border
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)

            current_row += 1

            # Add step-by-step logic items (if available)
            if step_by_step:
                for i, logic_step in enumerate(step_by_step, 1):
                    ws.cell(row=current_row, column=1, value=f"  {i}")
                    ws.cell(row=current_row, column=2, value='')
                    ws.cell(row=current_row, column=3, value='')
                    ws.cell(row=current_row, column=4, value=logic_step)

                    # Indent step number
                    ws.cell(row=current_row, column=1).alignment = Alignment(indent=2, vertical='top')

                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).border = self.border
                        ws.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)

                    current_row += 1

            # Blank row between components
            current_row += 1

        # DML Files section
        dml_files = abinitio_logic.get('dml_files', [])
        if dml_files:
            current_row += 1
            ws.cell(row=current_row, column=1, value="DML Files:")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=', '.join(dml_files))
            ws.merge_cells(f'B{current_row}:D{current_row}')

        # Column widths (4-column format)
        ws.column_dimensions['A'].width = 15   # Step
        ws.column_dimensions['B'].width = 35   # Component
        ws.column_dimensions['C'].width = 25   # Type
        ws.column_dimensions['D'].width = 60   # Transformation Logic

    def _create_logic_comparison_sheet(
        self,
        wb: Workbook,
        source_system: str,
        source_logic: Dict[str, Any],
        databricks_logic: Dict[str, Any]
    ):
        """Create Logic Comparison sheet (side-by-side)"""
        ws = wb.create_sheet("Logic Comparison")

        # Title
        ws['A1'] = "Side-by-Side Logic Comparison"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')

        # Headers
        headers = [
            f'{source_system.capitalize()} Item',
            f'{source_system.capitalize()} Details',
            'Comparison',
            'Databricks Item',
            'Databricks Details',
            'Notes'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Build comparison rows
        comparison_rows = self._build_logic_comparison(source_system, source_logic, databricks_logic)

        for row_idx, comp_row in enumerate(comparison_rows, 4):
            ws.cell(row=row_idx, column=1, value=comp_row['source_item'])
            ws.cell(row=row_idx, column=2, value=comp_row['source_details'])
            ws.cell(row=row_idx, column=3, value=comp_row['comparison'])
            ws.cell(row=row_idx, column=4, value=comp_row['databricks_item'])
            ws.cell(row=row_idx, column=5, value=comp_row['databricks_details'])
            ws.cell(row=row_idx, column=6, value=comp_row['notes'])

            # Color coding
            if comp_row['comparison'] == '✓ Similar':
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = self.similar_fill
            elif comp_row['comparison'] == '✗ Different':
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = self.different_fill

            # Formatting
            for col in range(1, 7):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 30

        ws.freeze_panes = 'A4'

    def _build_logic_comparison(
        self,
        source_system: str,
        source_logic: Dict[str, Any],
        databricks_logic: Dict[str, Any]
    ) -> List[Dict[str, str]]:
        """
        Build INTELLIGENT side-by-side comparison with semantic mapping

        NEW APPROACH:
        - Semantic matching: understands purpose, not just position
        - Handles 1-to-many and many-to-1 mappings
        - Identifies implementation differences
        """
        rows = []

        # Get items from both systems
        source_items = self._get_logic_items(source_system, source_logic)
        databricks_items = self._get_logic_items('databricks', databricks_logic)

        logger.info(f"   Comparing {len(source_items)} {source_system} items with {len(databricks_items)} Databricks items")

        # INTELLIGENT semantic matching
        matched_pairs = self._intelligent_logic_matching(
            source_items,
            databricks_items,
            source_system
        )

        # Generate comparison rows from matched pairs
        for pair in matched_pairs:
            source_item = pair.get('source_item', {})
            databricks_item = pair.get('databricks_item', {})
            mapping_type = pair.get('mapping_type', 'unknown')
            comparison_note = pair.get('comparison_note', '')

            # Determine comparison symbol
            if mapping_type == 'equivalent':
                comparison = '✓ Equivalent'
            elif mapping_type == 'similar':
                comparison = '≈ Similar'
            elif mapping_type == 'new_in_databricks':
                comparison = '+ New'
            elif mapping_type == 'removed':
                comparison = '- Removed'
            elif mapping_type == 'different_implementation':
                comparison = '⚡ Different Impl'
            else:
                comparison = '? Unknown'

            rows.append({
                'source_item': source_item.get('name', '') if source_item else '(none)',
                'source_details': source_item.get('details', '') if source_item else '',
                'comparison': comparison,
                'databricks_item': databricks_item.get('name', '') if databricks_item else '(none)',
                'databricks_details': databricks_item.get('details', '') if databricks_item else '',
                'notes': comparison_note
            })

        logger.info(f"   ✅ Generated {len(rows)} comparison rows")
        return rows

    def _intelligent_logic_matching(
        self,
        source_items: List[Dict],
        databricks_items: List[Dict],
        source_system: str
    ) -> List[Dict[str, Any]]:
        """
        Intelligently match source scripts to Databricks notebooks using name/purpose similarity

        For now, uses simple heuristics. Can be enhanced with AI later.

        Returns:
            [
                {
                    'source_item': Dict or None,
                    'databricks_item': Dict or None,
                    'mapping_type': str,
                    'comparison_note': str
                }
            ]
        """
        matched_pairs = []
        used_source_indices = set()
        used_databricks_indices = set()

        # Step 1: Try PURPOSE-BASED matching first (primary signal)
        for i, source_item in enumerate(source_items):
            source_name = source_item.get('name', '').lower()
            source_details = source_item.get('details', '').lower()
            source_purpose = self._extract_purpose_text(source_details)

            best_score = 0
            best_match_idx = None
            best_reason = ''

            for j, databricks_item in enumerate(databricks_items):
                if j in used_databricks_indices:
                    continue

                databricks_name = databricks_item.get('name', '').lower()
                databricks_details = databricks_item.get('details', '').lower()
                databricks_purpose = self._extract_purpose_text(databricks_details)

                # Calculate PURPOSE similarity score (0-100)
                purpose_score = self._calculate_purpose_similarity(
                    source_name, source_purpose, source_details,
                    databricks_name, databricks_purpose, databricks_details
                )

                # Calculate NAME similarity score (0-100) as fallback
                name_score = self._calculate_name_similarity(source_name, databricks_name)

                # Weighted score: Purpose 70%, Name 30%
                total_score = (purpose_score * 0.7) + (name_score * 0.3)

                if total_score > best_score:
                    best_score = total_score
                    best_match_idx = j
                    best_reason = self._generate_match_reason(
                        source_name, databricks_name,
                        purpose_score, name_score, total_score
                    )

            # If good match found (threshold: 40), create pair
            if best_match_idx is not None and best_score >= 40:
                databricks_item = databricks_items[best_match_idx]
                used_source_indices.add(i)
                used_databricks_indices.add(best_match_idx)

                # Determine mapping type based on scores
                if best_score >= 85:
                    mapping_type = 'equivalent'
                elif best_score >= 60:
                    mapping_type = 'similar'
                else:
                    mapping_type = 'different_implementation'

                matched_pairs.append({
                    'source_item': source_item,
                    'databricks_item': databricks_item,
                    'mapping_type': mapping_type,
                    'comparison_note': best_reason
                })

        # Step 2: Add unmatched source items (removed in Databricks)
        for i, source_item in enumerate(source_items):
            if i not in used_source_indices:
                matched_pairs.append({
                    'source_item': source_item,
                    'databricks_item': None,
                    'mapping_type': 'removed',
                    'comparison_note': 'Logic removed or deprecated in Databricks migration'
                })

        # Step 3: Add unmatched Databricks items (new features)
        for j, databricks_item in enumerate(databricks_items):
            if j not in used_databricks_indices:
                matched_pairs.append({
                    'source_item': None,
                    'databricks_item': databricks_item,
                    'mapping_type': 'new_in_databricks',
                    'comparison_note': 'New functionality added in Databricks'
                })

        return matched_pairs

    def _extract_purpose_text(self, details: str) -> str:
        """
        Extract the purpose text from details field

        Example: "Purpose: Validates date format\nTransformations: xyz"
                 → "validates date format"
        """
        if not details:
            return ""

        # Look for "Purpose:" or "Type:" sections
        import re
        purpose_match = re.search(r'purpose:\s*([^\n]+)', details, re.IGNORECASE)
        if purpose_match:
            return purpose_match.group(1).strip()

        # Fallback: use first line or full details
        first_line = details.split('\n')[0]
        if ':' in first_line:
            return first_line.split(':', 1)[1].strip()
        return details[:200]  # Limit to 200 chars

    def _calculate_purpose_similarity(
        self,
        source_name: str,
        source_purpose: str,
        source_details: str,
        databricks_name: str,
        databricks_purpose: str,
        databricks_details: str
    ) -> float:
        """
        Calculate semantic similarity between purposes using keyword matching

        Returns score 0-100
        """
        score = 0.0

        # Keyword-based semantic similarity
        # Define business purpose keywords and their synonyms
        purpose_keywords = {
            'date': ['date', 'datetime', 'timestamp', 'time', 'breadcrumb', 'bc'],
            'validation': ['validate', 'validation', 'check', 'verify', 'audit', 'idempotency', 'duplicate'],
            'extraction': ['extract', 'unzip', 'download', 'parse', 'process', 'get'],
            'transformation': ['transform', 'merge', 'join', 'union', 'aggregate', 'group'],
            'output': ['write', 'save', 'store', 'publish', 'output'],
            'file_ops': ['file', 'zip', 'copy', 'move', 'archive'],
            'database': ['cosmos', 'cosmosdb', 'database', 'db', 'table', 'maprdbtable'],
            'matching': ['match', 'matching', 'permid', 'patient', 'hospital']
        }

        # Convert to lowercase for comparison
        source_text = (source_purpose + " " + source_details).lower()
        databricks_text = (databricks_purpose + " " + databricks_details).lower()

        # Count matching purpose categories
        matching_categories = 0
        total_categories = 0

        for category, keywords in purpose_keywords.items():
            source_has_category = any(kw in source_text for kw in keywords)
            databricks_has_category = any(kw in databricks_text for kw in keywords)

            if source_has_category or databricks_has_category:
                total_categories += 1
                if source_has_category and databricks_has_category:
                    matching_categories += 1

        if total_categories > 0:
            score = (matching_categories / total_categories) * 100

        # Bonus for specific known equivalences (business context)
        if self._is_known_business_equivalence(source_name, source_text, databricks_name, databricks_text):
            score = min(score + 20, 100)  # Bonus but cap at 100

        return score

    def _is_known_business_equivalence(
        self,
        source_name: str,
        source_text: str,
        databricks_name: str,
        databricks_text: str
    ) -> bool:
        """
        Check for known business equivalences between Hadoop and Databricks

        Examples:
        - get_date.sh ≈ ADF parameters + widgets (date handling moved to pipeline)
        - audit_bdf_swift.sh ≈ extra_check_bcs.py (file validation → CosmosDB idempotency)
        - process_bdf.sh ≈ process_bdf.py (same name, similar extraction logic)
        """
        # Known equivalences based on business analysis
        equivalences = [
            # Date handling: shell script → pipeline parameters + widgets
            (['get_date', 'datetime'], ['widget', 'parameter', 'getargument']),

            # Idempotency: file validation → CosmosDB duplicate check
            (['audit', 'validate', 'check', 'breadcrumb'], ['cosmos', 'extra_check', 'idempotency']),

            # File extraction: unzip + parse
            (['process', 'extract', 'unzip', 'breadcrumb'], ['process', 'extract', 'zip']),

            # Patient record transformation
            (['merge', 'patient', 'permid', 'match'], ['merge', 'patient', 'permid', 'match']),
        ]

        for source_keywords, databricks_keywords in equivalences:
            source_match = any(kw in source_name or kw in source_text for kw in source_keywords)
            databricks_match = any(kw in databricks_name or kw in databricks_text for kw in databricks_keywords)

            if source_match and databricks_match:
                return True

        return False

    def _calculate_name_similarity(self, source_name: str, databricks_name: str) -> float:
        """
        Calculate name similarity score (0-100)
        """
        # Remove extensions
        source_base = source_name.replace('.sh', '').replace('.pig', '').replace('.py', '')
        databricks_base = databricks_name.replace('.py', '').replace('.ipynb', '')

        if source_base == databricks_base:
            return 100.0
        elif source_base in databricks_base or databricks_base in source_base:
            return 80.0
        elif any(word in databricks_base for word in source_base.split('_') if len(word) > 3):
            return 50.0
        else:
            return 0.0

    def _generate_match_reason(
        self,
        source_name: str,
        databricks_name: str,
        purpose_score: float,
        name_score: float,
        total_score: float
    ) -> str:
        """
        Generate human-readable reason for the match
        """
        reasons = []

        # Determine primary matching factor
        if name_score >= 80:
            reasons.append(f"Same/similar naming ({source_name} → {databricks_name})")
        elif name_score >= 50:
            reasons.append(f"Partial name match")

        if purpose_score >= 70:
            reasons.append(f"Equivalent business purpose (semantic similarity: {purpose_score:.0f}%)")
        elif purpose_score >= 40:
            reasons.append(f"Similar business purpose (semantic similarity: {purpose_score:.0f}%)")

        if not reasons:
            reasons.append(f"Weak match (confidence: {total_score:.0f}%)")

        # Add context for specific known patterns
        if 'get_date' in source_name and 'widget' in databricks_name.lower():
            reasons.append("Date handling moved to pipeline parameters + widgets")
        elif 'audit' in source_name and ('extra_check' in databricks_name or 'cosmos' in databricks_name.lower()):
            reasons.append("File validation → CosmosDB idempotency check")
        elif 'process_bdf' in source_name and 'process_bdf' in databricks_name:
            reasons.append("Core file extraction logic preserved")

        return " | ".join(reasons)

    def _get_logic_items(self, system: str, logic: Dict[str, Any]) -> List[Dict[str, str]]:
        """Extract logic items for comparison"""
        items = []

        if system == 'hadoop':
            for job in logic.get('jobs', []):
                items.append({
                    'name': job.get('name', ''),
                    'details': f"Purpose: {job.get('purpose', '')}\nTransformations: {'; '.join(job.get('transformations', []))}"
                })
        elif system == 'abinitio':
            for step in logic.get('steps', []):
                # Handle both dict and string formats
                if isinstance(step, dict):
                    items.append({
                        'name': step.get('dataset', ''),
                        'details': f"Type: {step.get('component_type', '')}\nTransformation: {step.get('transformation_rules', '')}"
                    })
                elif isinstance(step, str):
                    items.append({
                        'name': step,
                        'details': ''
                    })
                else:
                    items.append({
                        'name': 'Unknown step',
                        'details': ''
                    })
        elif system == 'databricks':
            for activity in logic.get('activities', []):
                items.append({
                    'name': activity.get('name', ''),
                    'details': f"Type: {activity.get('type', '')}\nPurpose: {activity.get('purpose', '')}"
                })

        return items

    def _generate_comparison_note(self, source_item: Dict, databricks_item: Dict) -> str:
        """Generate comparison note"""
        if not source_item:
            return "New in Databricks"
        if not databricks_item:
            return "Removed from Databricks"
        return "Corresponding implementation"

    def _create_sttm_sheets(
        self,
        wb: Workbook,
        sttm_mappings: List[Dict[str, Any]],
        source_system: str,
        source_logic: Dict[str, Any],
        databricks_logic: Dict[str, Any]
    ):
        """
        Create SINGLE STTM sheet with 3 sections (matching manual gold standard):
        Section 1: Databricks (Target) column-level schema
        Section 2: Source System (Hadoop/Ab Initio) column-level schema
        Section 3: Side-by-side comparison
        """
        ws = wb.create_sheet("STTM")

        # Section 1: Databricks (Rows 1-N)
        current_row = self._create_databricks_sttm_section(
            ws,
            databricks_logic,
            start_row=1
        )

        # Blank row between sections
        current_row += 1

        # Section 2: Source System (Rows N+1 - M)
        current_row = self._create_source_sttm_section(
            ws,
            source_logic,
            source_system,
            start_row=current_row
        )

        # Blank row between sections
        current_row += 1

        # Section 3: Comparison (Rows M+1 - end)
        self._create_sttm_comparison_section(
            ws,
            sttm_mappings,
            source_system,
            source_logic,
            databricks_logic,
            start_row=current_row
        )

        # Column widths
        self._set_sttm_column_widths(ws)
        ws.freeze_panes = 'A3'  # Freeze first section header

    def _create_databricks_sttm_section(
        self,
        ws,
        databricks_logic: Dict[str, Any],
        start_row: int
    ) -> int:
        """
        Create Databricks (Target) STTM section with column-level schemas

        Returns:
            Next available row number
        """
        logger.info("Creating Databricks STTM section...")

        # Row 1: Section title
        ws.cell(row=start_row, column=1, value="Databricks")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(f'A{start_row}:M{start_row}')

        # Row 2: Headers (13-column format)
        header_row = start_row + 1
        headers = [
            'Processing Order',
            'Schema',
            'Target Table Name',
            'Target Field Name',
            'Target Field Data Type',
            'pk?',
            'contains_pii',
            'Field Type',
            'Field Depends On',
            'Pre Processing Rules',
            'Source Field Names',
            'Source Dataset Name',
            'Field Definition'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # Row 3+: Column-level data from Databricks activities
        current_row = header_row + 1
        order = 1

        activities = databricks_logic.get('activities', [])
        for activity in activities:
            activity_name = activity.get('name', 'Unknown')
            notebook_path = activity.get('notebook', '')
            column_schemas = activity.get('column_schemas', {})

            # Process each table and its columns
            for table_name, columns in column_schemas.items():
                for col_info in columns:
                    col_name = col_info.get('name', '')
                    col_type = col_info.get('type', 'STRING')
                    transformation = col_info.get('transformation', '')
                    source_line = col_info.get('source_line', '')

                    # Write row
                    ws.cell(row=current_row, column=1, value=order)
                    ws.cell(row=current_row, column=2, value='DATABRICKS_BDF')
                    ws.cell(row=current_row, column=3, value=table_name)
                    ws.cell(row=current_row, column=4, value=col_name)
                    ws.cell(row=current_row, column=5, value=col_type)
                    ws.cell(row=current_row, column=6, value='False')  # Would need PK detection
                    ws.cell(row=current_row, column=7, value='False')  # Would need PII detection
                    ws.cell(row=current_row, column=8, value='Data Field')
                    ws.cell(row=current_row, column=9, value='')  # Dependencies
                    ws.cell(row=current_row, column=10, value=f'ACTIVITY: {activity_name}. {transformation}')
                    ws.cell(row=current_row, column=11, value=transformation)
                    ws.cell(row=current_row, column=12, value=notebook_path)
                    ws.cell(row=current_row, column=13, value=f'Field from {activity_name}')

                    # Apply formatting
                    for col_idx in range(1, 14):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.border = self.border
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

                    current_row += 1
                    order += 1

        # If no column schemas found, show message
        if current_row == header_row + 1:
            ws.cell(row=current_row, column=1, value="No Databricks column schemas extracted")
            ws.merge_cells(f'A{current_row}:M{current_row}')
            current_row += 1

        logger.info(f"   Databricks STTM section: {current_row - header_row - 1} column rows")
        return current_row

    def _create_source_sttm_section(
        self,
        ws,
        source_logic: Dict[str, Any],
        source_system: str,
        start_row: int
    ) -> int:
        """
        Create Source System (Hadoop/Ab Initio) STTM section with column-level schemas

        Returns:
            Next available row number
        """
        logger.info(f"Creating {source_system} STTM section...")

        # Row 1: Section title
        ws.cell(row=start_row, column=1, value=source_system.upper())
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(f'A{start_row}:M{start_row}')

        # Row 2: Headers (same 13-column format)
        header_row = start_row + 1
        headers = [
            'Processing Order',
            'Schema',
            'Table Name',
            'Field Name',
            'Field Data Type',
            'pk?',
            'contains_pii',
            'Field Type',
            'Field Depends On',
            'Pre Processing Rules',
            'Source Field Names',
            'Source Script Name',
            'Field Definition'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # Row 3+: Column-level data from source scripts
        current_row = header_row + 1
        order = 1

        jobs = source_logic.get('jobs', [])
        for job in jobs:
            script_name = job.get('script_file', job.get('name', 'Unknown'))
            column_schemas = job.get('column_schemas', {})

            # Process each table and its columns
            for table_name, columns in column_schemas.items():
                for col_info in columns:
                    col_name = col_info.get('name', '')
                    col_type = col_info.get('type', 'chararray')
                    transformation = col_info.get('transformation', '')
                    source_line = col_info.get('source_line', '')

                    # Write row
                    ws.cell(row=current_row, column=1, value=order)
                    ws.cell(row=current_row, column=2, value=source_logic.get('workflow_name', 'ES_BDF'))
                    ws.cell(row=current_row, column=3, value=table_name)
                    ws.cell(row=current_row, column=4, value=col_name)
                    ws.cell(row=current_row, column=5, value=col_type)
                    ws.cell(row=current_row, column=6, value='False')  # Would need PK detection
                    ws.cell(row=current_row, column=7, value='False')  # Would need PII detection
                    ws.cell(row=current_row, column=8, value='Data Field')
                    ws.cell(row=current_row, column=9, value='')  # Dependencies
                    ws.cell(row=current_row, column=10, value=transformation)
                    ws.cell(row=current_row, column=11, value=col_name)
                    ws.cell(row=current_row, column=12, value=script_name)
                    ws.cell(row=current_row, column=13, value=f'Field from {script_name}')

                    # Apply formatting
                    for col_idx in range(1, 14):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.border = self.border
                        cell.alignment = Alignment(vertical='top', wrap_text=True)

                    current_row += 1
                    order += 1

        # If no column schemas found, show message
        if current_row == header_row + 1:
            ws.cell(row=current_row, column=1, value=f"No {source_system} column schemas extracted")
            ws.merge_cells(f'A{current_row}:M{current_row}')
            current_row += 1

        logger.info(f"   {source_system} STTM section: {current_row - header_row - 1} column rows")
        return current_row

    def _create_sttm_comparison_section(
        self,
        ws,
        sttm_mappings: List[Dict[str, Any]],
        source_system: str,
        source_logic: Dict[str, Any],
        databricks_logic: Dict[str, Any],
        start_row: int
    ):
        """
        Create STTM Comparison section with side-by-side field mappings

        This is a 4-column format (different from sections 1-2):
        - Feature
        - Databricks Transformation
        - Hadoop/Ab Initio Transformation
        - Comparison Summary
        """
        logger.info("Creating STTM Comparison section...")

        # Row 1: Section title (blank for spacing - comparison starts directly)
        header_row = start_row
        headers = [
            'Feature',
            'Databricks Transformation',
            f'{source_system} Transformation',
            'Comparison Summary'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # Clear extra columns (comparison only uses 4 columns)
        for col in range(5, 14):
            cell = ws.cell(row=header_row, column=col, value='')

        # Comparison rows
        current_row = header_row + 1

        # Generate comparison from column schemas
        # Match tables/columns between Databricks and Source
        comparisons = self._generate_column_comparisons(
            source_logic,
            databricks_logic,
            source_system
        )

        for comparison in comparisons:
            ws.cell(row=current_row, column=1, value=comparison.get('feature', ''))
            ws.cell(row=current_row, column=2, value=comparison.get('databricks_transformation', ''))
            ws.cell(row=current_row, column=3, value=comparison.get('source_transformation', ''))
            ws.cell(row=current_row, column=4, value=comparison.get('summary', ''))

            # Apply formatting
            for col_idx in range(1, 5):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

            current_row += 1

        # If no comparisons, show message
        if not comparisons:
            ws.cell(row=current_row, column=1, value="No column-level comparisons generated")
            ws.merge_cells(f'A{current_row}:D{current_row}')

        logger.info(f"   Comparison section: {len(comparisons)} comparison rows")

    def _generate_column_comparisons(
        self,
        source_logic: Dict[str, Any],
        databricks_logic: Dict[str, Any],
        source_system: str
    ) -> List[Dict[str, Any]]:
        """Generate field-level comparisons between source and Databricks"""
        comparisons = []

        # Build dictionaries of all column schemas
        source_columns = {}
        databricks_columns = {}

        # Extract source columns
        jobs = source_logic.get('jobs', [])
        for job in jobs:
            column_schemas = job.get('column_schemas', {})
            for table_name, columns in column_schemas.items():
                for col_info in columns:
                    col_name = col_info.get('name', '')
                    key = f"{table_name}.{col_name}"
                    source_columns[key] = {
                        'table': table_name,
                        'column': col_name,
                        'type': col_info.get('type', ''),
                        'transformation': col_info.get('transformation', ''),
                        'script': job.get('script_file', '')
                    }

        # Extract Databricks columns
        activities = databricks_logic.get('activities', [])
        for activity in activities:
            column_schemas = activity.get('column_schemas', {})
            for table_name, columns in column_schemas.items():
                for col_info in columns:
                    col_name = col_info.get('name', '')
                    key = f"{table_name}.{col_name}"
                    databricks_columns[key] = {
                        'table': table_name,
                        'column': col_name,
                        'type': col_info.get('type', ''),
                        'transformation': col_info.get('transformation', ''),
                        'activity': activity.get('name', '')
                    }

        # Match columns by name
        all_keys = set(list(source_columns.keys()) + list(databricks_columns.keys()))

        for key in sorted(all_keys):
            src_col = source_columns.get(key)
            db_col = databricks_columns.get(key)

            if src_col and db_col:
                # Both exist - compare
                comparisons.append({
                    'feature': f"{db_col['column']} ({db_col['table']})",
                    'databricks_transformation': f"Source: {db_col['transformation']}. Type: {db_col['type']}. Activity: {db_col['activity']}",
                    'source_transformation': f"Source: {src_col['transformation']}. Type: {src_col['type']}. Script: {src_col['script']}",
                    'summary': 'Column exists in both systems' if src_col['type'] == db_col['type'] else f"Type change: {src_col['type']} → {db_col['type']}"
                })
            elif db_col:
                # Only in Databricks
                comparisons.append({
                    'feature': f"{db_col['column']} ({db_col['table']})",
                    'databricks_transformation': f"Source: {db_col['transformation']}. Type: {db_col['type']}. Activity: {db_col['activity']}",
                    'source_transformation': 'NOT IN SOURCE',
                    'summary': 'New column added in Databricks migration'
                })
            elif src_col:
                # Only in source
                comparisons.append({
                    'feature': f"{src_col['column']} ({src_col['table']})",
                    'databricks_transformation': 'NOT IN DATABRICKS',
                    'source_transformation': f"Source: {src_col['transformation']}. Type: {src_col['type']}. Script: {src_col['script']}",
                    'summary': 'Column removed in Databricks migration'
                })

        return comparisons

    def _set_sttm_column_widths(self, ws):
        """Set consistent column widths for all STTM sheets"""
        ws.column_dimensions['A'].width = 12   # Order/Processing Order
        ws.column_dimensions['B'].width = 18   # Schema
        ws.column_dimensions['C'].width = 20   # Table
        ws.column_dimensions['D'].width = 22   # Field/Column
        ws.column_dimensions['E'].width = 12   # Data Type
        ws.column_dimensions['F'].width = 8    # pk?
        ws.column_dimensions['G'].width = 12   # PII
        ws.column_dimensions['H'].width = 15   # Field Type
        ws.column_dimensions['I'].width = 25   # Depends On
        ws.column_dimensions['J'].width = 50   # Transformation/Rules
        ws.column_dimensions['K'].width = 25   # Source Fields/Script
        ws.column_dimensions['L'].width = 25   # Source Dataset/Path
        ws.column_dimensions['M'].width = 40   # Description/Definition


# Example usage
if __name__ == "__main__":
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl not installed. Please run: pip install openpyxl")
        exit(1)

    print("\n" + "=" * 80)
    print("EXCEL GENERATOR TEST")
    print("=" * 80)

    generator = ExcelGenerator()

    # Sample data
    business_stages = [
        {
            'stage_name': 'Data Ingestion',
            'databricks_description': 'Load customer data from ADLS',
            'source_description': 'Load customer data from HDFS',
            'comparison': 'Similar',
            'notes': 'Both load from respective cloud storage'
        }
    ]

    source_logic = {
        'system': 'hadoop',
        'workflow_name': 'ES_BDF_DOWNLOAD',
        'jobs': [
            {'name': 'Load_Data', 'script': 'load.pig', 'purpose': 'Load data', 'inputs': ['raw_data'], 'outputs': ['processed'], 'transformations': ['Filter', 'Transform']}
        ]
    }

    databricks_logic = {
        'system': 'databricks',
        'pipeline_name': 'pl_cdd_bdf_download',
        'activities': [
            {'name': 'Ingest_Data', 'type': 'DatabricksNotebook', 'notebook': 'nb_load', 'purpose': 'Load data', 'inputs': ['adls_raw'], 'outputs': ['delta_table']}
        ]
    }

    sttm_mappings = [
        {
            'target_column': 'customer_id',
            'data_type': 'STRING',
            'source_columns': ['cust_id'],
            'transformation_logic': 'DIRECT',
            'dependencies': [],
            'confidence': 0.95
        }
    ]

    print("\nGenerating test Excel file...")
    output_path = generator.generate_comparison_excel(
        source_system='hadoop',
        source_workflow='ES_BDF_DOWNLOAD',
        databricks_pipeline='pl_cdd_bdf_download',
        business_stages=business_stages,
        source_logic=source_logic,
        databricks_logic=databricks_logic,
        sttm_mappings=sttm_mappings,
        output_folder='test_output'
    )

    print(f"\n✅ Test Excel file generated: {output_path}")
    print("\n" + "=" * 80)
