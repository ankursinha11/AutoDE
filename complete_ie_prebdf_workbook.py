#!/usr/bin/env python3
"""
Complete IE_PREBDF Excel Workbook with Databricks Information
Adds Databricks flow, logic, STTM, and comprehensive comparisons
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

def complete_overview_sheet(wb):
    """Add Databricks flow to Overview sheet"""
    ws = wb['Overview']
    
    # Find where to add Databricks flow (after Hadoop flow)
    # Hadoop flow ends around row 15, add Databricks starting at column H
    
    # Add Databricks header
    ws['H1'] = 'Databricks (pl_cdd_ie_prebdf) Flow'
    ws['H1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['H1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('H1:K1')
    
    # Add Databricks flow headers
    headers = ['Step', 'Notebook/Activity', 'Description', 'Input → Output']
    for col_idx, header in enumerate(headers, start=8):  # Column H=8
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Databricks flow data
    databricks_flow = [
        ('1', 'ieprebdf_sqoop.py', 'JDBC data ingestion', 'SQL Server (ICH2/3/4) → /sqoop/$bcdate/{db}'),
        ('2', 'ie-copy-to-input.py', 'Consolidate and filter sources', 'Filter: ClientId NOT IN (100515,100000,100523,0), UserId!=2344720; Combine all DBs → /input/ie/$bcdate'),
        ('3', 'Parse-271-final.py', 'Parse 271 XML responses', 'XPath extraction of 50+ fields from XML → /parsed_271/$bcdate'),
        ('4', 'cdd-parse-demographics.py', 'Extract demographics', 'Parse subscriber/dependent demographics from XML → /parsed_demographics/$bcdate'),
        ('5', 'cdd_cleanse_demographics.py', 'Comprehensive validation', 'Filter: 271 responses (St01=271), exclude VA (ClientId!=100515), ResponseResultStd NOT IN (0,3,4,5,7,8,10,15,16,17,312,313,314); Validate: State, Gender, SSN, Coverage IDs; Cross-populate addresses; Replace nulls → /parsed270/$bcdate'),
        ('6', 'ie-evaluate-parsed.py', 'Validate parse results', 'Evaluate quality of parsed data → /parsed/$bcdate'),
        ('7', 'ie-Parse-all.py', 'Consolidate parsed data', 'Merge all parsed records → /parsed_all/$bcdate'),
        ('8', 'ie-parse-eb.py', 'Extract EB segments', 'PySpark extraction of eligibility benefit segments from XML → /eb_segment/$bcdate'),
        ('9', 'ie-aaa-segment.py', 'Extract AAA segments', 'PySpark extraction of AAA segments from XML → /aaa_segment/$bcdate'),
        ('10', 'ie-parse-dtp.py', 'Extract DTP segments', 'PySpark extraction of date/time period segments → /dtp_segment/$bcdate'),
        ('11', 'ie-filter.py', 'Filter and separate records', 'Separate subscribers/dependents; Create 3 record types with Custom1 markers; Validate ID (numeric), filter bad IDs and multibyte characters → /filtered/$bcdate, /rejected/$bcdate, /wBadChars/$bcdate'),
        ('12', 'ie-bdf.py', 'BDF format conversion', 'Date: Validate DOB, format YYYY-MM-DD; SSN: Remove dashes, validate 9-digit; ZIP: Validate 5-digit; String: Clean non-ASCII, TRIM; Output 16-field pipe-delimited → /bdf_pipe/$bcdate'),
        ('13', 'ie-process-dupes.py', 'Deduplicate demographics', 'Convert blanks to NULL; Group by demographics; Aggregate: min(transactionkey), collect_list(transactionkey) → /bdf_parquet/$bcdate, /bdf_deduped/$bcdate'),
        ('14', 'ie_process_bdf.py', 'Final BDF processing', 'Final BDF file preparation → /bdf_final/$bcdate'),
        ('15', 'ie-upload-bdf.py', 'Upload BDF to destination', 'Upload final BDF files to target location'),
    ]
    
    # Add Databricks flow data
    for row_idx, (step, notebook, desc, io) in enumerate(databricks_flow, start=3):
        ws.cell(row=row_idx, column=8).value = step
        ws.cell(row=row_idx, column=9).value = notebook
        ws.cell(row=row_idx, column=10).value = desc
        ws.cell(row=row_idx, column=11).value = io
        
        # Apply formatting
        for col in range(8, 12):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Add Flow Comparison section
    comparison_start_row = len(databricks_flow) + 5
    ws.cell(row=comparison_start_row, column=1).value = 'Flow Comparison'
    ws.cell(row=comparison_start_row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(f'A{comparison_start_row}:D{comparison_start_row}')
    
    comparison_headers = ['Stage', 'Hadoop', 'Databricks', 'Match Status']
    for col_idx, header in enumerate(comparison_headers, start=1):
        cell = ws.cell(row=comparison_start_row+1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    flow_comparison = [
        ('Data Ingestion', 'Sqoop (ie_list_sqoop.xml)', 'JDBC (ieprebdf_sqoop.py)', 'Different technology, same purpose'),
        ('Source Filtering', 'At Sqoop query level', 'At consolidation level (ie-copy-to-input.py)', 'Different stage, same logic'),
        ('XML Parsing', 'Pig scripts', 'PySpark notebooks', 'Technology difference'),
        ('Demographics Cleansing', 'Spark (02_cdd_cleanse_demographics.py)', 'PySpark (cdd_cleanse_demographics.py)', '⚠️ GAP: ResponseResultStd filtering differs'),
        ('Record Separation', 'filter.pig', 'ie-filter.py', 'Matched'),
        ('BDF Formatting', 'bdf_olb.pig', 'ie-bdf.py', 'Matched'),
        ('Deduplication', 'process_ich_dupes.py', 'ie-process-dupes.py', 'Matched'),
    ]
    
    for row_idx, (stage, hadoop, databricks, status) in enumerate(flow_comparison, start=comparison_start_row+2):
        ws.cell(row=row_idx, column=1).value = stage
        ws.cell(row=row_idx, column=2).value = hadoop
        ws.cell(row=row_idx, column=3).value = databricks
        ws.cell(row=row_idx, column=4).value = status
        
        # Highlight gaps in red
        if '⚠️ GAP' in status or 'Different' in status:
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif 'Matched' in status:
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Auto-size columns
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 25


def add_sqoop_to_hadoop_logic(ws):
    """Add Sqoop ingestion to Hadoop Logic sheet if missing"""
    # Check if Sqoop is already there
    has_sqoop = False
    for row in ws.iter_rows(min_row=3, max_row=20, min_col=2, max_col=2):
        if row[0].value and 'sqoop' in str(row[0].value).lower():
            has_sqoop = True
            break
    
    if not has_sqoop:
        # Insert Sqoop as first step
        ws.insert_rows(3)
        ws.cell(row=3, column=1).value = '1'
        ws.cell(row=3, column=2).value = 'ie_list_sqoop.xml (ICH2/3/4.xml)'
        ws.cell(row=3, column=3).value = 'Sqoop data ingestion from SQL Server'
        ws.cell(row=3, column=4).value = 'JDBC query: SELECT tracenumber, message FROM oltp.MessageXml JOIN UserRequest JOIN Submission WHERE tracenumber > last_max AND clientid NOT IN (100515,100000,100523,0) AND userid != 2344720; Parallel import with 16 mappers, split by tracenumber; Update max tracenumber after each DB'
        ws.cell(row=3, column=5).value = 'SQL Server (ICH2, ICH3, ICH4)'
        ws.cell(row=3, column=6).value = '/CDD/intermediate_tmp/ie/$bcdate/{counter}'
        
        # Renumber subsequent rows
        for row_idx in range(4, ws.max_row + 1):
            current_num = ws.cell(row=row_idx, column=1).value
            if current_num and str(current_num).isdigit():
                ws.cell(row=row_idx, column=1).value = str(int(current_num) + 1)


def create_databricks_logic_sheet(wb):
    """Create Databricks Logic sheet"""
    if 'Databricks Logic' in wb.sheetnames:
        del wb['Databricks Logic']
    
    ws = wb.create_sheet('Databricks Logic')
    
    # Title
    ws['A1'] = 'Databricks (pl_cdd_ie_prebdf) - Complete Logic Flow'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Headers
    headers = ['Step', 'Notebook/Activity', 'Description', 'Detailed Logic', 'Input Path', 'Output Path']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Databricks logic data
    databricks_logic = [
        ('1', 'ieprebdf_sqoop.py', 'JDBC data ingestion', 'Loop through databases (ICH2, ICH3, ICH4); Read max received date from ADLS; Query: SELECT tracenumber, message FROM oltp.MessageXmlExtended WHERE received > last_max_date AND received <= max_date - 12 hours; PySpark JDBC with fetchsize=2000, numPartitions=16, partitionColumn=tracenumber; Remove newlines from message; Write as text files; Update max received date', 'SQL Server databases', '/sqoop/$bcdate/{db}'),
        ('2', 'ie-copy-to-input.py', 'Consolidate and filter', 'Read from all DB directories; Parse tracenumber and clientid from CSV; Filter: clientid NOT IN (100515,100000,100523,0) AND userid != 2344720; Union all databases; Write consolidated output', '/sqoop/$bcdate/{db}', '/input/ie/$bcdate'),
        ('3', 'Parse-271-final.py', 'Parse 271 XML', 'Parse XML using ElementTree; Extract 271 transaction set; XPath extraction of response codes, trace numbers, identifiers', '/input/ie/$bcdate', '/parsed_271/$bcdate'),
        ('4', 'cdd-parse-demographics.py', 'Extract demographics', 'XPath extraction of 50+ fields: St01, TraceNumber, PayerId, ClientId, ResponseResultStd, ResponseResultExt, subscriber/dependent demographics (name, address, DOB, SSN, gender, group numbers, coverage IDs); Parse from HL segments (HL01=3 for subscriber, HL01=4 for dependent)', '/parsed_271/$bcdate', '/parsed_demographics/$bcdate'),
        ('5', 'cdd_cleanse_demographics.py', 'Comprehensive validation', 'Filter: St01=271 only; Filter: ClientId != 100515 (VA Veteran); Filter: ResponseResultStd NOT IN (0,3,4,5,7,8,10,15,16,17,312,313,314); Trim all fields; Replace nulls/empty/null strings with XX/00000000/000000000; Validate State (2-letter codes from valid list); Validate Gender (M/F only, else X); Remove special chars from Coverage IDs, Group Numbers, Subscriber ID; Remove dashes from SSN; Cross-populate: If subscriber address is XX and dependent address exists, copy dependent address to subscriber; Cross-populate: If subscriber coverage ID is XX and subscriber ID exists, use subscriber ID; Filter: At least one valid demographic field (not all XX/00000000)', '/parsed_demographics/$bcdate', '/parsed270/$bcdate'),
        ('6', 'ie-evaluate-parsed.py', 'Validate parse results', 'Evaluate quality and completeness of parsed data; Check for required fields; Validate data formats', '/parsed270/$bcdate', '/parsed/$bcdate'),
        ('7', 'ie-Parse-all.py', 'Consolidate parsed', 'Read all parsed files; Union/merge all records; Remove duplicates; Write consolidated output', '/parsed/$bcdate', '/parsed_all/$bcdate'),
        ('8', 'ie-parse-eb.py', 'Extract EB segments', 'Parse XML for EB (Eligibility Benefit) segments; Extract benefit information, coverage details, service types; Map to output schema with ResponseResultStd', '/input/ie/$bcdate', '/eb_segment/$bcdate'),
        ('9', 'ie-aaa-segment.py', 'Extract AAA segments', 'Parse XML for AAA segments; Extract additional account information', '/input/ie/$bcdate', '/aaa_segment/$bcdate'),
        ('10', 'ie-parse-dtp.py', 'Extract DTP segments', 'Parse XML for DTP (Date/Time Period) segments; Extract date ranges, coverage periods', '/input/ie/$bcdate', '/dtp_segment/$bcdate'),
        ('11', 'ie-filter.py', 'Filter and separate', 'Define schema (28 fields); Separate subscribers (dependent fields are null/empty) and dependents (has dependent data); Create 3 record types: (1) Subscriber: ID=TraceNumber, Custom1=00DS00ie00SUB00; (2) Dependent Subscriber: ID=TraceNumber, Custom1=00DS00ie00DEPSUB00; (3) Dependent: ID=concat(D,TraceNumber), Custom1=00DS00ie00DEP00; Union all record types; Validate ID: Filter badID (non-numeric or empty using regex); Validate characters: Filter multibyte chars (ord <= 31 or ord >= 127); Write 3 outputs: filtered (good), rejected (bad ID), wBadChars (multibyte)', '/parsed/$bcdate', '/filtered/$bcdate, /rejected/$bcdate, /wBadChars/$bcdate'),
        ('12', 'ie-bdf.py', 'BDF format conversion', 'Read filtered data (16-field schema); Date validation: isValidDate UDF (YYYYMMDD or YYYY-MM-DD), format to YYYY-MM-DD, blank if year < 1900; SSN validation: Remove dashes, validate 9-digit numeric, blank if invalid patterns (111111111-999999999, 000000000, 123456789, 987654321); ZIP validation: Validate 5-digit numeric, blank if 00000, 00001, or <= 00500; String cleaning: remove_non_ascii UDF (ord < 127), getCleanString (trim + clean); Apply to FN, MN, LN, Address, Custom2; Output 16-field pipe-delimited: ID|FN|MN|LN|Gender|SSN|DOB|Address|Address2|City|State|Zip|Zip4|Phone|Custom1|Custom2', '/filtered/$bcdate', '/bdf_pipe/$bcdate'),
        ('13', 'ie-process-dupes.py', 'Deduplicate', 'Read pipe-delimited BDF; Convert blanks to NULL (trim, filter non-null/non-empty); Filter: At least one non-NULL demographic field; Group by all demographics (fn, mn, ln, gender, ssn, dob, address, address2, city, state, zip, zip4, phone, custom1, custom2); Aggregate: min(transactionkey) as primary key, collect_list(transactionkey) as all keys; Write deduplicated records and xref mapping', '/bdf_pipe/$bcdate', '/bdf_parquet/$bcdate (xref), /bdf_deduped/$bcdate'),
        ('14', 'ie_process_bdf.py', 'Final BDF processing', 'Final BDF file preparation and formatting; Apply any final transformations', '/bdf_deduped/$bcdate', '/bdf_final/$bcdate'),
        ('15', 'ie-upload-bdf.py', 'Upload BDF', 'Upload final BDF files to target location/storage', '/bdf_final/$bcdate', 'Target destination'),
    ]
    
    # Add data
    for row_idx, (step, notebook, desc, logic, input_path, output_path) in enumerate(databricks_logic, start=3):
        ws.cell(row=row_idx, column=1).value = step
        ws.cell(row=row_idx, column=2).value = notebook
        ws.cell(row=row_idx, column=3).value = desc
        ws.cell(row=row_idx, column=4).value = logic
        ws.cell(row=row_idx, column=5).value = input_path
        ws.cell(row=row_idx, column=6).value = output_path
        
        # Apply formatting
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Auto-size columns
    for col in range(1, 7):
        if col == 4:  # Detailed Logic column
            ws.column_dimensions[get_column_letter(col)].width = 60
        else:
            ws.column_dimensions[get_column_letter(col)].width = 25


def complete_logic_comparison_sheet(wb):
    """Complete the Logic Comparison sheet with Databricks logic"""
    ws = wb['Comparison']
    
    # Clear existing Databricks columns if any
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=4):
        for cell in row:
            cell.value = None
    
    # Add detailed comparison data
    comparison_data = [
        # Row 2: Data Ingestion
        ('Sqoop from databases → XML files\nie_list_sqoop.xml: Sqoop insurance eligibility data from SQL Server (ICH2, ICH3, ICH4) with filtering: clientid NOT IN (100515,100000,100523,0), userid != 2344720', 
         'JDBC from databases → Text files\nieprebdf_sqoop.py: PySpark JDBC ingestion from SQL Server (ICH2, ICH3, ICH4) using MessageXmlExtended table, incremental by received timestamp\nNOTE: Filtering happens later in ie-copy-to-input.py', 
         '⚠️ Different: Hadoop filters at ingestion, Databricks filters at consolidation. Different source tables (MessageXml vs MessageXmlExtended)'),
        
        # Row 4: XML Parsing
        ('ie_parse_271.pig + 01_cdd_parse_demographics.pig\nXPath extraction of 50+ fields from 271 eligibility XML responses', 
         'Parse-271-final.py + cdd-parse-demographics.py\nXPath extraction of 50+ fields from 271 eligibility XML responses using ElementTree', 
         'Matched: Same logic, different technology (Pig vs PySpark)'),
        
        # Row 6-10: Data Cleansing (expand existing rows)
        ('02_cdd_cleanse_demographics.py (Spark)\nFilter: 271 responses only (St01=271), exclude VA Veteran (ClientId!=100515)\nFilter: ResponseResultStd NOT IN (3, 0)\nValidate: State (2-letter codes), Gender (M/F only), SSN (remove dashes), Coverage IDs\nNull handling: Replace with XX/00000000/000000000\nCross-populate: Subscriber address from dependent if subscriber address is missing',
         'cdd_cleanse_demographics.py (PySpark)\nFilter: 271 responses only (St01=271), exclude VA Veteran (ClientId!=100515)\nFilter: ResponseResultStd NOT IN (0,3,4,5,7,8,10,15,16,17,312,313,314)\nValidate: State (2-letter codes), Gender (M/F only), SSN (remove dashes), Coverage IDs\nNull handling: Replace with XX/00000000/000000000\nCross-populate: Subscriber address from dependent if subscriber address is missing',
         '🚨 CRITICAL GAP: ResponseResultStd filtering differs!\nHadoop excludes: 3, 0 (2 codes)\nDatabricks excludes: 0,3,4,5,7,8,10,15,16,17,312,313,314 (13 codes)\nDatabricks will filter out 11 additional response codes, resulting in FEWER records'),
        
        # Row 11-14: Record Separation
        ('filter.pig: Separate into 3 record types\nSubscriber (no dependent): Custom1=00DS00ie00SUB00\nDependent Subscriber (has dependent): Custom1=00DS00ie00DEPSUB00\nDependent (separate record): Custom1=00DS00ie00DEP00, ID prefixed with D',
         'ie-filter.py: Separate into 3 record types\nSubscriber (no dependent): Custom1=00DS00ie00SUB00\nDependent Subscriber (has dependent): Custom1=00DS00ie00DEPSUB00\nDependent (separate record): Custom1=00DS00ie00DEP00, ID prefixed with D',
         'Matched: Identical logic'),
        
        # Row 15: ID Validation
        ('filter.pig: Regex validation (numeric only)\nFilter bad IDs and multibyte characters',
         'ie-filter.py: Regex validation (numeric only)\nFilter bad IDs (non-numeric) and multibyte characters (ord <= 31 or ord >= 127)',
         'Matched: Identical logic'),
        
        # Row 16-17: BDF Formatting
        ('bdf_olb.pig: BDF format conversion\nDate: Validate DOB, format to YYYY-MM-DD, blank bad dates\nSSN: Remove dashes, validate 9-digit, blank invalid patterns\nZIP: Validate 5-digit, blank if 00000/00001/<=500\nString: Clean non-ASCII, TRIM all fields\nOutput 16-field pipe-delimited',
         'ie-bdf.py: BDF format conversion\nDate: Validate DOB (isValidDate UDF), format to YYYY-MM-DD, blank if year < 1900\nSSN: Remove dashes, validate 9-digit, blank invalid patterns (111111111-999999999, etc.)\nZIP: Validate 5-digit, blank if 00000/00001/<=500\nString: Clean non-ASCII (remove_non_ascii UDF, ord < 127), TRIM all fields\nOutput 16-field pipe-delimited',
         'Matched: Identical validation logic'),
        
        # Row 18: Deduplication
        ('process_ich_dupes.py: Deduplicate demographics\nGroup by all demographics\nAggregate: min(transactionkey), collect_list(transactionkey)',
         'ie-process-dupes.py: Deduplicate demographics\nGroup by all demographics\nAggregate: min(transactionkey), collect_list(transactionkey)',
         'Matched: Identical logic'),
    ]
    
    # Update comparison data starting from row 2
    comparison_rows = [2, 4, 6, 11, 15, 16, 18]
    for row_idx, (hadoop, databricks, comparison) in zip(comparison_rows, comparison_data):
        ws.cell(row=row_idx, column=2).value = hadoop
        ws.cell(row=row_idx, column=3).value = databricks
        ws.cell(row=row_idx, column=4).value = comparison
        
        # Apply formatting
        for col in range(2, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Highlight gaps
        comparison_cell = ws.cell(row=row_idx, column=4)
        if '🚨 CRITICAL GAP' in comparison or '⚠️ Different' in comparison:
            comparison_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            comparison_cell.font = Font(bold=True)
        elif 'Matched' in comparison:
            comparison_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Auto-size columns
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 35


def create_databricks_sttm_sheet(wb):
    """Create Databricks STTM sheet matching Hadoop format"""
    if 'Databricks STTM' in wb.sheetnames:
        del wb['Databricks STTM']
    
    ws = wb.create_sheet('Databricks STTM')
    
    # Title
    ws['A1'] = 'Databricks (pl_cdd_ie_prebdf)'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:N1')
    
    # Headers (same as Hadoop STTM)
    headers = ['id', 'Processing Order', 'Schema', 'Source Field Name', 'Source Dataset Name', 
               'Target Table/File Name', 'Target Field Name', 'Target Field Data Type', 'pk?', 
               'contains_pii', 'Field Type', 'Field Depends On', 'Pre Processing Rules', 'Field Definition']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Databricks STTM data (matching Hadoop structure)
    sttm_data = [
        (1, 1, 'IE_PREBDF_PARSE', 'St01', 'Submission XML', '271_parsed/$bcdate', 'St01', 'StringType', 'False', 'False', 'Metadata', 'cdd-parse-demographics.py', 'ElementTree XPath extract from XML', 'Transaction set identifier (271 for eligibility)'),
        (2, 1, 'IE_PREBDF_PARSE', 'TraceNumber', 'Submission XML', '271_parsed/$bcdate', 'TraceNumber', 'StringType', 'True', 'False', 'Identifier', 'cdd-parse-demographics.py', 'ElementTree XPath extract from PayerRequest', 'Unique trace number for eligibility inquiry'),
        (3, 1, 'IE_PREBDF_PARSE', 'PayerId', 'Submission XML', '271_parsed/$bcdate', 'PayerId', 'StringType', 'False', 'False', 'Identifier', 'cdd-parse-demographics.py', 'ElementTree XPath extract from Identity', 'Payer identifier'),
        (4, 1, 'IE_PREBDF_PARSE', 'ClientId', 'Submission XML', '271_parsed/$bcdate', 'ClientId', 'StringType', 'False', 'False', 'Identifier', 'cdd-parse-demographics.py', 'ElementTree XPath extract from Submission/Identity', 'Client identifier'),
        (5, 1, 'IE_PREBDF_PARSE', 'ResponseResultStd', 'Submission XML', '271_parsed/$bcdate', 'ResponseResultStd', 'StringType', 'False', 'False', 'Metadata', 'cdd-parse-demographics.py', 'ElementTree XPath extract from PayerResponse', 'Standard response result code'),
        (6, 1, 'IE_PREBDF_PARSE', 'ResponseResultExt', 'Submission XML', '271_parsed/$bcdate', 'ResponseResultExt', 'StringType', 'False', 'False', 'Metadata', 'cdd-parse-demographics.py', 'ElementTree XPath extract from PayerResponse', 'Extended response result code'),
        (7, 1, 'IE_PREBDF_PARSE', 'subscriberGroupNumber', 'Submission XML', '271_parsed/$bcdate', 'subscriberGroupNumber', 'StringType', 'False', 'False', 'Identifier', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/REF[@REF01="IG"]/@REF02', 'Subscriber group number'),
        (8, 1, 'IE_PREBDF_PARSE', 'subscriberCoverageId', 'Submission XML', '271_parsed/$bcdate', 'subscriberCoverageId', 'StringType', 'False', 'False', 'Identifier', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1[@NM108="MI"]/@NM109', 'Subscriber coverage ID (member ID)'),
        (9, 1, 'IE_PREBDF_PARSE', 'subscriberLastName', 'Submission XML', '271_parsed/$bcdate', 'subscriberLastName', 'StringType', 'False', 'True', 'Demographic', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/@NM103', 'Subscriber last name'),
        (10, 1, 'IE_PREBDF_PARSE', 'subscriberFirstName', 'Submission XML', '271_parsed/$bcdate', 'subscriberFirstName', 'StringType', 'False', 'True', 'Demographic', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/@NM104', 'Subscriber first name'),
        (11, 1, 'IE_PREBDF_PARSE', 'subscriberMiddleName', 'Submission XML', '271_parsed/$bcdate', 'subscriberMiddleName', 'StringType', 'False', 'True', 'Demographic', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/@NM105', 'Subscriber middle name'),
        (12, 1, 'IE_PREBDF_PARSE', 'subscriberAddress', 'Submission XML', '271_parsed/$bcdate', 'subscriberAddress', 'StringType', 'False', 'True', 'Address', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/N3/@N301', 'Subscriber street address'),
        (13, 1, 'IE_PREBDF_PARSE', 'subscriberCity', 'Submission XML', '271_parsed/$bcdate', 'subscriberCity', 'StringType', 'False', 'True', 'Address', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/N4/@N401', 'Subscriber city'),
        (14, 1, 'IE_PREBDF_PARSE', 'subscriberState', 'Submission XML', '271_parsed/$bcdate', 'subscriberState', 'StringType', 'False', 'True', 'Address', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/N4/@N402', 'Subscriber state'),
        (15, 1, 'IE_PREBDF_PARSE', 'subscriberZip', 'Submission XML', '271_parsed/$bcdate', 'subscriberZip', 'StringType', 'False', 'True', 'Address', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/N4/@N403', 'Subscriber ZIP code'),
        (16, 1, 'IE_PREBDF_PARSE', 'subscriberDOB', 'Submission XML', '271_parsed/$bcdate', 'subscriberDOB', 'StringType', 'False', 'True', 'Demographic', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/DMG/@DMG02', 'Subscriber date of birth'),
        (17, 1, 'IE_PREBDF_PARSE', 'subscriberGender', 'Submission XML', '271_parsed/$bcdate', 'subscriberGender', 'StringType', 'False', 'True', 'Demographic', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/DMG/@DMG03', 'Subscriber gender'),
        (18, 1, 'IE_PREBDF_PARSE', 'subscriberSSN', 'Submission XML', '271_parsed/$bcdate', 'subscriberSSN', 'StringType', 'False', 'True', 'PII', 'cdd-parse-demographics.py', 'ElementTree XPath: HL[@HL01="3"]/NM1/REF[@REF01="SY"]/@REF02', 'Subscriber Social Security Number'),
        (19, 2, 'IE_PREBDF_CLEANSE', 'ResponseResultStd', '271_parsed/$bcdate', 'parsed270/$bcdate', 'ResponseResultStd', 'StringType', 'False', 'False', 'Metadata', 'cdd_cleanse_demographics.py', 'Filter: CAST(ResponseResultStd AS BIGINT) NOT IN (0,3,4,5,7,8,10,15,16,17,312,313,314)', 'Filtered response result code'),
        (20, 2, 'IE_PREBDF_CLEANSE', 'subscriberState', '271_parsed/$bcdate', 'parsed270/$bcdate', 'subscriberState', 'StringType', 'False', 'True', 'Address', 'cdd_cleanse_demographics.py', 'Validate: LENGTH=2 AND IN (valid state list), else XX', 'Validated subscriber state'),
        (21, 2, 'IE_PREBDF_CLEANSE', 'subscriberGender', '271_parsed/$bcdate', 'parsed270/$bcdate', 'subscriberGender', 'StringType', 'False', 'True', 'Demographic', 'cdd_cleanse_demographics.py', 'Validate: UPPER(TRIM) IN (M, F), else X', 'Validated subscriber gender'),
        (22, 2, 'IE_PREBDF_CLEANSE', 'subscriberSSN', '271_parsed/$bcdate', 'parsed270/$bcdate', 'subscriberSSN', 'StringType', 'False', 'True', 'PII', 'cdd_cleanse_demographics.py', 'REGEXP_REPLACE(TRIM, -, empty); Remove special chars; CAST validation', 'Cleansed SSN (dashes removed)'),
        (23, 3, 'IE_PREBDF_FILTER', 'ID', 'parsed/$bcdate', 'filtered/$bcdate', 'ID', 'StringType', 'True', 'False', 'Identifier', 'ie-filter.py', 'Subscriber: TraceNumber; Dependent: concat(D, TraceNumber); Validate numeric regex', 'Record identifier'),
        (24, 3, 'IE_PREBDF_FILTER', 'Custom1', 'parsed/$bcdate', 'filtered/$bcdate', 'Custom1', 'StringType', 'False', 'False', 'Metadata', 'ie-filter.py', 'Subscriber: 00DS00ie00SUB00; Dep Sub: 00DS00ie00DEPSUB00; Dep: 00DS00ie00DEP00', 'Record type marker'),
        (25, 4, 'IE_PREBDF_BDF', 'DOB', 'filtered/$bcdate', 'bdf_pipe/$bcdate', 'DOB', 'StringType', 'False', 'True', 'Demographic', 'ie-bdf.py', 'isValidDate UDF; Format YYYY-MM-DD; Blank if year < 1900', 'Validated and formatted DOB'),
        (26, 4, 'IE_PREBDF_BDF', 'SSN', 'filtered/$bcdate', 'bdf_pipe/$bcdate', 'SSN', 'StringType', 'False', 'True', 'PII', 'ie-bdf.py', 'Remove dashes; Validate 9-digit; Blank if test SSN (111111111-999999999, 000000000, 123456789, 987654321)', 'Validated SSN'),
        (27, 4, 'IE_PREBDF_BDF', 'Zip', 'filtered/$bcdate', 'bdf_pipe/$bcdate', 'Zip', 'StringType', 'False', 'True', 'Address', 'ie-bdf.py', 'Validate 5-digit; Blank if 00000, 00001, or <= 00500', 'Validated ZIP code'),
        (28, 4, 'IE_PREBDF_BDF', 'FN/MN/LN/Address', 'filtered/$bcdate', 'bdf_pipe/$bcdate', 'FN/MN/LN/Address', 'StringType', 'False', 'True', 'Demographic/Address', 'ie-bdf.py', 'remove_non_ascii UDF (ord < 127); getCleanString (trim + clean); Apply to all string fields', 'Cleaned string fields'),
    ]
    
    # Add data
    for row_idx, row_data in enumerate(sttm_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.cell(row=row_idx, column=col_idx).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Auto-size columns
    for col in range(1, 15):
        if col in [4, 13, 14]:  # Field name, rules, definition
            ws.column_dimensions[get_column_letter(col)].width = 35
        else:
            ws.column_dimensions[get_column_letter(col)].width = 18


def create_sttm_comparison_sheet(wb):
    """Create STTM Comparison sheet"""
    if 'STTM Comparison' in wb.sheetnames:
        del wb['STTM Comparison']
    
    ws = wb.create_sheet('STTM Comparison')
    
    # Title
    ws['A1'] = 'STTM Comparison - Hadoop vs Databricks'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Headers
    headers = ['Field Name', 'Hadoop Processing', 'Hadoop Rules', 'Databricks Processing', 
               'Databricks Rules', 'Match Status', 'Impact', 'Notes']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Comparison data
    comparison_data = [
        ('TraceNumber', 'XPath extract', 'Direct extraction from XML', 'ElementTree XPath', 'Direct extraction from XML', 'Matched', 'None', 'Identical logic'),
        ('ResponseResultStd', 'XPath extract + Filter', 'Filter: NOT IN (3, 0)', 'ElementTree XPath + Filter', 'Filter: NOT IN (0,3,4,5,7,8,10,15,16,17,312,313,314)', '🚨 CRITICAL GAP', 'HIGH - Databricks filters 11 additional codes', 'Databricks will have FEWER records due to additional filtering'),
        ('ClientId', 'XPath extract + Filter', 'Filter at Sqoop: != 100515, 100000, 100523, 0', 'ElementTree XPath + Filter', 'Filter at consolidation: NOT IN (100515, 100000, 100523, 0)', 'Matched', 'None', 'Same logic, different stage'),
        ('subscriberState', 'XPath + Validation', '2-letter validation from valid state list', 'ElementTree XPath + Validation', '2-letter validation from valid state list', 'Matched', 'None', 'Identical validation'),
        ('subscriberGender', 'XPath + Validation', 'Validate M/F only, else X', 'ElementTree XPath + Validation', 'Validate M/F only, else X', 'Matched', 'None', 'Identical validation'),
        ('subscriberSSN', 'XPath + Cleansing', 'Remove dashes, validate 9-digit numeric', 'ElementTree XPath + Cleansing', 'Remove dashes, validate 9-digit numeric', 'Matched', 'None', 'Identical cleansing'),
        ('subscriberCoverageId', 'XPath + Cleansing', 'Remove special chars; Cross-populate from subscriberId if XX', 'ElementTree XPath + Cleansing', 'Remove special chars; Cross-populate from subscriberId if XX', 'Matched', 'None', 'Identical logic'),
        ('subscriberAddress', 'XPath + Cross-populate', 'Copy from dependent if subscriber is XX', 'ElementTree XPath + Cross-populate', 'Copy from dependent if subscriber is XX', 'Matched', 'None', 'Identical cross-population'),
        ('ID (BDF)', 'Pig transformation', 'Subscriber: TraceNumber; Dependent: D+TraceNumber', 'PySpark transformation', 'Subscriber: TraceNumber; Dependent: concat(D, TraceNumber)', 'Matched', 'None', 'Identical logic'),
        ('Custom1', 'Pig transformation', 'SUB: 00DS00ie00SUB00; DEPSUB: 00DS00ie00DEPSUB00; DEP: 00DS00ie00DEP00', 'PySpark transformation', 'SUB: 00DS00ie00SUB00; DEPSUB: 00DS00ie00DEPSUB00; DEP: 00DS00ie00DEP00', 'Matched', 'None', 'Identical markers'),
        ('DOB', 'Pig UDF validation', 'Validate date, format YYYY-MM-DD, blank if year < 1900', 'PySpark UDF validation', 'isValidDate UDF, format YYYY-MM-DD, blank if year < 1900', 'Matched', 'None', 'Identical validation'),
        ('SSN (BDF)', 'Pig UDF validation', 'Remove dashes, 9-digit validation, blank test SSNs', 'PySpark UDF validation', 'Remove dashes, 9-digit validation, blank test SSNs (111111111-999999999, etc.)', 'Matched', 'None', 'Identical validation'),
        ('Zip', 'Pig validation', 'Validate 5-digit, blank if 00000/00001/<=500', 'PySpark validation', 'Validate 5-digit, blank if 00000/00001/<=500', 'Matched', 'None', 'Identical validation'),
        ('String Fields', 'Pig UDF cleaning', 'Remove non-ASCII, TRIM', 'PySpark UDF cleaning', 'remove_non_ascii (ord < 127), getCleanString, TRIM', 'Matched', 'None', 'Identical cleaning'),
    ]
    
    # Add data
    for row_idx, row_data in enumerate(comparison_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Highlight gaps
            if col_idx == 6:  # Match Status column
                if '🚨 CRITICAL GAP' in value:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(bold=True, color="C00000")
                elif 'Matched' in value:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            if col_idx == 7:  # Impact column
                if 'HIGH' in value:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(bold=True)
    
    # Auto-size columns
    for col in range(1, 9):
        if col in [3, 5, 8]:  # Rules and Notes columns
            ws.column_dimensions[get_column_letter(col)].width = 40
        else:
            ws.column_dimensions[get_column_letter(col)].width = 22


def main():
    """Main function to complete the workbook"""
    print("Loading IE_PREBDF_HADOOP.xlsx...")
    wb = openpyxl.load_workbook('/Users/ankurshome/Desktop/Hadoop_Parser/IE_PREBDF_HADOOP.xlsx')
    
    print(f"Found sheets: {wb.sheetnames}")
    
    print("\n1. Completing Overview_Flow sheet with Databricks flow and comparison...")
    # Rename function call to use correct sheet name
    ws_overview = wb['Overview_Flow']
    complete_overview_sheet_updated(wb, ws_overview)
    
    print("2. Adding Sqoop to Logic sheet if missing...")
    ws_logic = wb['Logic']
    add_sqoop_to_hadoop_logic(ws_logic)
    
    print("3. Creating Databricks Logic sheet...")
    create_databricks_logic_sheet(wb)
    
    print("4. Completing Comparison sheet...")
    complete_logic_comparison_sheet(wb)
    
    print("5. Creating Databricks STTM sheet...")
    create_databricks_sttm_sheet(wb)
    
    print("6. Creating STTM Comparison sheet...")
    create_sttm_comparison_sheet(wb)
    
    # Save the completed workbook
    output_file = '/Users/ankurshome/Desktop/Hadoop_Parser/IE_PREBDF_COMPLETE.xlsx'
    wb.save(output_file)
    print(f"\n✅ Completed workbook saved to: {output_file}")
    print("\nSummary:")
    print("- Overview_Flow: Added Databricks flow and flow comparison")
    print("- Logic: Verified Sqoop ingestion is included")
    print("- Databricks Logic: Created complete logic sheet")
    print("- Comparison: Completed with detailed Databricks logic and gap analysis")
    print("- Databricks STTM: Created matching Hadoop format")
    print("- STTM Comparison: Created with gap highlighting")
    print("\n🚨 CRITICAL GAP IDENTIFIED:")
    print("   ResponseResultStd filtering differs between Hadoop and Databricks")
    print("   Hadoop: Excludes 2 codes (3, 0)")
    print("   Databricks: Excludes 13 codes (0,3,4,5,7,8,10,15,16,17,312,313,314)")
    print("   Impact: Databricks will have FEWER records in output")


def complete_overview_sheet_updated(wb, ws):
    """Add Databricks flow to Overview_Flow sheet"""
    # Add Databricks header
    ws['H1'] = 'Databricks (pl_cdd_ie_prebdf) Flow'
    ws['H1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['H1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells('H1:K1')
    
    # Add Databricks flow headers
    headers = ['Step', 'Notebook/Activity', 'Description', 'Input → Output']
    for col_idx, header in enumerate(headers, start=8):  # Column H=8
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Databricks flow data
    databricks_flow = [
        ('1', 'ieprebdf_sqoop.py', 'JDBC data ingestion', 'SQL Server (ICH2/3/4) → /sqoop/$bcdate/{db}'),
        ('2', 'ie-copy-to-input.py', 'Consolidate and filter sources', 'Filter: ClientId NOT IN (100515,100000,100523,0), UserId!=2344720; Combine all DBs → /input/ie/$bcdate'),
        ('3', 'Parse-271-final.py', 'Parse 271 XML responses', 'XPath extraction of 50+ fields from XML → /parsed_271/$bcdate'),
        ('4', 'cdd-parse-demographics.py', 'Extract demographics', 'Parse subscriber/dependent demographics from XML → /parsed_demographics/$bcdate'),
        ('5', 'cdd_cleanse_demographics.py', 'Comprehensive validation', 'Filter: 271 responses (St01=271), exclude VA (ClientId!=100515), ResponseResultStd NOT IN (0,3,4,5,7,8,10,15,16,17,312,313,314); Validate: State, Gender, SSN, Coverage IDs; Cross-populate addresses; Replace nulls → /parsed270/$bcdate'),
        ('6', 'ie-evaluate-parsed.py', 'Validate parse results', 'Evaluate quality of parsed data → /parsed/$bcdate'),
        ('7', 'ie-Parse-all.py', 'Consolidate parsed data', 'Merge all parsed records → /parsed_all/$bcdate'),
        ('8', 'ie-parse-eb.py', 'Extract EB segments', 'PySpark extraction of eligibility benefit segments from XML → /eb_segment/$bcdate'),
        ('9', 'ie-aaa-segment.py', 'Extract AAA segments', 'PySpark extraction of AAA segments from XML → /aaa_segment/$bcdate'),
        ('10', 'ie-parse-dtp.py', 'Extract DTP segments', 'PySpark extraction of date/time period segments → /dtp_segment/$bcdate'),
        ('11', 'ie-filter.py', 'Filter and separate records', 'Separate subscribers/dependents; Create 3 record types with Custom1 markers; Validate ID (numeric), filter bad IDs and multibyte characters → /filtered/$bcdate, /rejected/$bcdate, /wBadChars/$bcdate'),
        ('12', 'ie-bdf.py', 'BDF format conversion', 'Date: Validate DOB, format YYYY-MM-DD; SSN: Remove dashes, validate 9-digit; ZIP: Validate 5-digit; String: Clean non-ASCII, TRIM; Output 16-field pipe-delimited → /bdf_pipe/$bcdate'),
        ('13', 'ie-process-dupes.py', 'Deduplicate demographics', 'Convert blanks to NULL; Group by demographics; Aggregate: min(transactionkey), collect_list(transactionkey) → /bdf_parquet/$bcdate, /bdf_deduped/$bcdate'),
        ('14', 'ie_process_bdf.py', 'Final BDF processing', 'Final BDF file preparation → /bdf_final/$bcdate'),
        ('15', 'ie-upload-bdf.py', 'Upload BDF to destination', 'Upload final BDF files to target location'),
    ]
    
    # Add Databricks flow data
    for row_idx, (step, notebook, desc, io) in enumerate(databricks_flow, start=3):
        ws.cell(row=row_idx, column=8).value = step
        ws.cell(row=row_idx, column=9).value = notebook
        ws.cell(row=row_idx, column=10).value = desc
        ws.cell(row=row_idx, column=11).value = io
        
        # Apply formatting
        for col in range(8, 12):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Add Flow Comparison section
    comparison_start_row = len(databricks_flow) + 5
    ws.cell(row=comparison_start_row, column=1).value = 'Flow Comparison'
    ws.cell(row=comparison_start_row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(f'A{comparison_start_row}:D{comparison_start_row}')
    
    comparison_headers = ['Stage', 'Hadoop', 'Databricks', 'Match Status']
    for col_idx, header in enumerate(comparison_headers, start=1):
        cell = ws.cell(row=comparison_start_row+1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    flow_comparison = [
        ('Data Ingestion', 'Sqoop (ie_list_sqoop.xml)', 'JDBC (ieprebdf_sqoop.py)', 'Different technology, same purpose'),
        ('Source Filtering', 'At Sqoop query level', 'At consolidation level (ie-copy-to-input.py)', 'Different stage, same logic'),
        ('XML Parsing', 'Pig scripts', 'PySpark notebooks', 'Technology difference'),
        ('Demographics Cleansing', 'Spark (02_cdd_cleanse_demographics.py)', 'PySpark (cdd_cleanse_demographics.py)', '⚠️ GAP: ResponseResultStd filtering differs'),
        ('Record Separation', 'filter.pig', 'ie-filter.py', 'Matched'),
        ('BDF Formatting', 'bdf_olb.pig', 'ie-bdf.py', 'Matched'),
        ('Deduplication', 'process_ich_dupes.py', 'ie-process-dupes.py', 'Matched'),
    ]
    
    for row_idx, (stage, hadoop, databricks, status) in enumerate(flow_comparison, start=comparison_start_row+2):
        ws.cell(row=row_idx, column=1).value = stage
        ws.cell(row=row_idx, column=2).value = hadoop
        ws.cell(row=row_idx, column=3).value = databricks
        ws.cell(row=row_idx, column=4).value = status
        
        # Highlight gaps in red
        if '⚠️ GAP' in status or 'Different' in status:
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif 'Matched' in status:
            ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Auto-size columns
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 25


if __name__ == "__main__":
    main()
