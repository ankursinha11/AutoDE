import os
import re
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import ast
import sqlparse
from collections import defaultdict
import xml.etree.ElementTree as ET

class MigrationValidationTool:
    """
    Tool to validate that Azure Databricks migration covers all Hadoop legacy functionality
    """
    
    def __init__(self):
        self.hadoop_processes = []
        self.databricks_processes = []
        self.hadoop_execution_flows = {}
        self.databricks_execution_flows = {}
        self.validation_results = {}
        
        self.business_logic_patterns = {
            'permid': ['permid', 'person_id', 'personid', 'tu_id', 'transunion'],
            'coverage': ['coverage', 'insurance', 'policy', 'payer', 'benefit'],
            'patient': ['patient', 'demographic', 'demographics', 'patientacct'],
            'address': ['address', 'addr', 'location', 'zip', 'city', 'state'],
            'phone': ['phone', 'telephone', 'contact'],
            'family': ['family', 'household', 'member', 'dependent', 'subscriber'],
            'lead': ['lead', 'discovery', 'generation', 'propagation'],
            'validation': ['validate', 'validation', 'check', 'verify'],
            'transformation': ['transform', 'convert', 'map', 'process'],
            'merge': ['merge', 'join', 'combine', 'union'],
            'filter': ['filter', 'where', 'condition', 'criteria'],
            'group': ['group', 'aggregate', 'sum', 'count', 'avg'],
            'sort': ['sort', 'order', 'rank'],
            'dedupe': ['dedupe', 'distinct', 'unique', 'duplicate'],
            'cm2': ['cm2', 'matching', 'pass', 'stage'],
            'bdf': ['bdf', 'batch', 'data', 'file'],
            'prebdf': ['prebdf', 'pre', 'before', 'preparation'],
            'postbdf': ['postbdf', 'post', 'after', 'processing'],
            'escan': ['escan', 'eligibility', 'enrollment'],
            'ich': ['ich', 'individual', 'enrollment'],
            'chc': ['chc', 'consumer', 'health', 'care'],
            'gmrn': ['gmrn', 'global', 'medical', 'record']
        }
    
    def analyze_hadoop_legacy(self, hadoop_path):
        """Analyze Hadoop legacy system execution flow and functionality"""
        print(f"🔍 Analyzing Hadoop Legacy System: {hadoop_path}")
        
        hadoop_path = Path(hadoop_path)
        
        # Analyze workflow files for execution order
        workflow_files = list(hadoop_path.rglob("*workflow*.xml"))
        for workflow_file in workflow_files:
            self.analyze_oozie_workflow(workflow_file)
        
        # Analyze coordinator files for scheduling
        coordinator_files = list(hadoop_path.rglob("*coordinator*.xml"))
        for coordinator_file in coordinator_files:
            self.analyze_oozie_coordinator(coordinator_file)
        
        # Analyze Pig scripts (main business logic)
        pig_files = list(hadoop_path.rglob("*.pig"))
        for pig_file in pig_files:
            process = self.analyze_pig_script_detailed(pig_file)
            if process:
                self.hadoop_processes.append(process)
        
        # Analyze PySpark scripts
        py_files = list(hadoop_path.rglob("*.py"))
        for py_file in py_files:
            process = self.analyze_pyspark_script_detailed(py_file)
            if process:
                self.hadoop_processes.append(process)
        
        # Build Hadoop execution flow
        self.build_hadoop_execution_flow()
        
        print(f"✅ Found {len(self.hadoop_processes)} Hadoop processes")
        print(f"✅ Analyzed {len(self.hadoop_execution_flows)} Hadoop execution flows")
    
    def analyze_azure_databricks_current(self, databricks_path):
        """Analyze Azure Databricks current system execution flow and functionality"""
        print(f"🔍 Analyzing Azure Databricks Current System: {databricks_path}")
        
        databricks_path = Path(databricks_path)
        
        # Look for pipeline definition files
        pipeline_files = list(databricks_path.rglob("*pipeline*.json"))
        for pipeline_file in pipeline_files:
            self.analyze_databricks_pipeline(pipeline_file)
        
        # Look for workflow files
        workflow_files = list(databricks_path.rglob("*workflow*.json"))
        for workflow_file in workflow_files:
            self.analyze_databricks_workflow(workflow_file)
        
        # Analyze notebooks
        py_files = list(databricks_path.rglob("*.py"))
        for py_file in py_files:
            process = self.analyze_databricks_notebook_detailed(py_file)
            if process:
                self.databricks_processes.append(process)
        
        sql_files = list(databricks_path.rglob("*.sql"))
        for sql_file in sql_files:
            process = self.analyze_databricks_sql_detailed(sql_file)
            if process:
                self.databricks_processes.append(process)
        
        # Build Databricks execution flow
        self.build_databricks_execution_flow()
        
        print(f"✅ Found {len(self.databricks_processes)} Databricks processes")
        print(f"✅ Analyzed {len(self.databricks_execution_flows)} Databricks execution flows")
    
    def analyze_oozie_workflow(self, workflow_file):
        """Analyze Oozie workflow XML to understand execution order"""
        try:
            tree = ET.parse(workflow_file)
            root = tree.getroot()
            
            workflow_name = root.get('name', 'Unknown')
            execution_order = []
            
            # Extract start node
            start_node = root.find('.//start')
            if start_node is not None:
                execution_order.append(start_node.get('to'))
            
            # Extract action sequence
            actions = root.findall('.//action')
            for action in actions:
                action_name = action.get('name')
                ok_node = action.find('ok')
                if ok_node is not None:
                    next_action = ok_node.get('to')
                    execution_order.append((action_name, next_action))
            
            # Extract forks and joins
            forks = root.findall('.//fork')
            for fork in forks:
                fork_name = fork.get('name')
                paths = [path.get('start') for path in fork.findall('.//path')]
                execution_order.append(('fork', fork_name, paths))
            
            joins = root.findall('.//join')
            for join in joins:
                join_name = join.get('name')
                execution_order.append(('join', join_name))
            
            self.hadoop_execution_flows[workflow_name] = {
                'type': 'Oozie Workflow',
                'file': str(workflow_file),
                'execution_order': execution_order
            }
            
        except Exception as e:
            print(f"⚠️ Error analyzing Oozie workflow {workflow_file}: {e}")
    
    def analyze_oozie_coordinator(self, coordinator_file):
        """Analyze Oozie coordinator XML to understand scheduling"""
        try:
            tree = ET.parse(coordinator_file)
            root = tree.getroot()
            
            coordinator_name = root.get('name', 'Unknown')
            frequency = root.get('frequency', 'Unknown')
            start_time = root.get('start', 'Unknown')
            end_time = root.get('end', 'Unknown')
            
            # Extract workflow path
            workflow_elem = root.find('.//workflow/app-path')
            workflow_path = workflow_elem.text if workflow_elem is not None else 'Unknown'
            
            self.hadoop_execution_flows[coordinator_name] = {
                'type': 'Oozie Coordinator',
                'file': str(coordinator_file),
                'frequency': frequency,
                'start_time': start_time,
                'end_time': end_time,
                'workflow_path': workflow_path
            }
            
        except Exception as e:
            print(f"⚠️ Error analyzing Oozie coordinator {coordinator_file}: {e}")
    
    def analyze_databricks_pipeline(self, pipeline_file):
        """Analyze Databricks pipeline JSON to understand execution order"""
        try:
            with open(pipeline_file, 'r', encoding='utf-8') as f:
                pipeline_data = json.load(f)
            
            pipeline_name = pipeline_data.get('name', 'Unknown')
            stages = pipeline_data.get('stages', [])
            
            execution_order = []
            for stage in stages:
                stage_name = stage.get('name', 'Unknown')
                notebooks = stage.get('notebooks', [])
                dependencies = stage.get('dependencies', [])
                
                execution_order.append({
                    'stage': stage_name,
                    'notebooks': notebooks,
                    'dependencies': dependencies
                })
            
            self.databricks_execution_flows[pipeline_name] = {
                'type': 'Databricks Pipeline',
                'file': str(pipeline_file),
                'execution_order': execution_order
            }
            
        except Exception as e:
            print(f"⚠️ Error analyzing Databricks pipeline {pipeline_file}: {e}")
    
    def analyze_databricks_workflow(self, workflow_file):
        """Analyze Databricks workflow JSON"""
        try:
            with open(workflow_file, 'r', encoding='utf-8') as f:
                workflow_data = json.load(f)
            
            workflow_name = workflow_data.get('name', 'Unknown')
            tasks = workflow_data.get('tasks', [])
            
            execution_order = []
            for task in tasks:
                task_name = task.get('task_key', 'Unknown')
                notebook_path = task.get('notebook_task', {}).get('notebook_path', 'Unknown')
                dependencies = task.get('depends_on', [])
                
                execution_order.append({
                    'task': task_name,
                    'notebook': notebook_path,
                    'dependencies': dependencies
                })
            
            self.databricks_execution_flows[workflow_name] = {
                'type': 'Databricks Workflow',
                'file': str(workflow_file),
                'execution_order': execution_order
            }
            
        except Exception as e:
            print(f"⚠️ Error analyzing Databricks workflow {workflow_file}: {e}")
    
    def analyze_pig_script_detailed(self, pig_file):
        """Analyze Pig script with detailed execution flow and comments"""
        try:
            with open(pig_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract comments and section headers
            comments = self.extract_pig_comments(content)
            
            # Extract execution steps
            execution_steps = self.extract_pig_execution_steps(content)
            
            # Extract data flow
            data_flow = self.extract_pig_data_flow_detailed(content)
            
            # Extract business logic from comments
            business_logic = self.extract_business_logic_from_comments(comments)
            
            # Analyze CM2 matching passes if present
            cm2_passes = self.extract_cm2_matching_passes(content)
            
            # Extract detailed functionality description
            functionality = self.extract_functionality_description(comments, content)
            
            process = {
                'type': 'Pig Script',
                'name': pig_file.name,
                'path': str(pig_file),
                'relative_path': str(pig_file.relative_to(pig_file.parents[2])),
                'comments': comments,
                'execution_steps': execution_steps,
                'data_flow': data_flow,
                'business_logic': business_logic,
                'cm2_passes': cm2_passes,
                'functionality': functionality,
                'content_snippet': content[:1000] + "..." if len(content) > 1000 else content
            }
            
            return process
            
        except Exception as e:
            print(f"⚠️ Error analyzing Pig script {pig_file}: {e}")
            return None
    
    def analyze_pyspark_script_detailed(self, py_file):
        """Analyze PySpark script with detailed execution flow"""
        try:
            with open(py_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract comments and docstrings
            comments = self.extract_python_comments(content)
            
            # Extract execution steps
            execution_steps = self.extract_pyspark_execution_steps(content)
            
            # Extract data flow
            data_flow = self.extract_pyspark_data_flow_detailed(content)
            
            # Extract business logic
            business_logic = self.extract_business_logic_from_comments(comments)
            
            # Extract function definitions and their purposes
            functions = self.extract_python_functions_detailed(content)
            
            # Extract detailed functionality description
            functionality = self.extract_functionality_description(comments, content)
            
            process = {
                'type': 'PySpark Script',
                'name': py_file.name,
                'path': str(py_file),
                'relative_path': str(py_file.relative_to(py_file.parents[2])),
                'comments': comments,
                'execution_steps': execution_steps,
                'data_flow': data_flow,
                'business_logic': business_logic,
                'functions': functions,
                'functionality': functionality,
                'content_snippet': content[:1000] + "..." if len(content) > 1000 else content
            }
            
            return process
            
        except Exception as e:
            print(f"⚠️ Error analyzing PySpark script {py_file}: {e}")
            return None
    
    def analyze_databricks_notebook_detailed(self, py_file):
        """Analyze Databricks notebook with detailed execution flow"""
        try:
            with open(py_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract comments and docstrings
            comments = self.extract_python_comments(content)
            
            # Extract execution steps
            execution_steps = self.extract_databricks_execution_steps(content)
            
            # Extract data flow
            data_flow = self.extract_databricks_data_flow_detailed(content)
            
            # Extract business logic
            business_logic = self.extract_business_logic_from_comments(comments)
            
            # Extract function definitions
            functions = self.extract_python_functions_detailed(content)
            
            # Extract detailed functionality description
            functionality = self.extract_functionality_description(comments, content)
            
            process = {
                'type': 'Databricks Notebook',
                'name': py_file.name,
                'path': str(py_file),
                'relative_path': str(py_file.relative_to(py_file.parents[2])),
                'comments': comments,
                'execution_steps': execution_steps,
                'data_flow': data_flow,
                'business_logic': business_logic,
                'functions': functions,
                'functionality': functionality,
                'content_snippet': content[:1000] + "..." if len(content) > 1000 else content
            }
            
            return process
            
        except Exception as e:
            print(f"⚠️ Error analyzing Databricks notebook {py_file}: {e}")
            return None
    
    def analyze_databricks_sql_detailed(self, sql_file):
        """Analyze Databricks SQL with detailed execution flow"""
        try:
            with open(sql_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Extract comments
            comments = self.extract_sql_comments(content)
            
            # Extract execution steps
            execution_steps = self.extract_sql_execution_steps(content)
            
            # Extract business logic
            business_logic = self.extract_business_logic_from_comments(comments)
            
            # Extract detailed functionality description
            functionality = self.extract_functionality_description(comments, content)
            
            process = {
                'type': 'Databricks SQL',
                'name': sql_file.name,
                'path': str(sql_file),
                'relative_path': str(sql_file.relative_to(sql_file.parents[2])),
                'comments': comments,
                'execution_steps': execution_steps,
                'business_logic': business_logic,
                'functionality': functionality,
                'content_snippet': content[:1000] + "..." if len(content) > 1000 else content
            }
            
            return process
            
        except Exception as e:
            print(f"⚠️ Error analyzing Databricks SQL {sql_file}: {e}")
            return None
    
    def extract_pig_comments(self, content):
        """Extract comments from Pig script"""
        comments = []
        lines = content.split('\n')
        
        for line in lines:
            line = line.strip()
            if line.startswith('--') or line.startswith('/*') or line.startswith('*'):
                comments.append(line)
        
        return comments
    
    def extract_python_comments(self, content):
        """Extract comments and docstrings from Python script"""
        comments = []
        lines = content.split('\n')
        
        in_docstring = False
        for line in lines:
            line_stripped = line.strip()
            
            # Check for docstrings
            if '"""' in line_stripped or "'''" in line_stripped:
                in_docstring = not in_docstring
            
            # Check for regular comments
            if line_stripped.startswith('#') and not in_docstring:
                comments.append(line_stripped)
            
            # Add docstring content
            if in_docstring and line_stripped:
                comments.append(line_stripped)
        
        return comments
    
    def extract_sql_comments(self, content):
        """Extract comments from SQL script"""
        comments = []
        lines = content.split('\n')
        
        for line in lines:
            line = line.strip()
            if line.startswith('--') or line.startswith('/*'):
                comments.append(line)
        
        return comments
    
    def extract_pig_execution_steps(self, content):
        """Extract execution steps from Pig script"""
        steps = []
        lines = content.split('\n')
        
        current_step = None
        for line in lines:
            line_stripped = line.strip()
            
            # Look for section headers
            if line_stripped.startswith('----'):
                current_step = line_stripped.replace('-', '').strip()
                steps.append({'step': current_step, 'operations': []})
            
            # Look for Pig operations
            elif current_step and ('=' in line_stripped and ('LOAD' in line_stripped or 'FOREACH' in line_stripped or 'FILTER' in line_stripped or 'GROUP' in line_stripped or 'JOIN' in line_stripped or 'STORE' in line_stripped)):
                if steps:
                    steps[-1]['operations'].append(line_stripped)
        
        return steps
    
    def extract_pyspark_execution_steps(self, content):
        """Extract execution steps from PySpark script"""
        steps = []
        lines = content.split('\n')
        
        current_step = None
        for line in lines:
            line_stripped = line.strip()
            
            # Look for section headers
            if line_stripped.startswith('#') and ('STEP' in line_stripped or 'PHASE' in line_stripped or 'PASS' in line_stripped):
                current_step = line_stripped.replace('#', '').strip()
                steps.append({'step': current_step, 'operations': []})
            
            # Look for Spark operations
            elif current_step and ('.' in line_stripped and ('load' in line_stripped or 'select' in line_stripped or 'filter' in line_stripped or 'groupBy' in line_stripped or 'join' in line_stripped or 'save' in line_stripped)):
                if steps:
                    steps[-1]['operations'].append(line_stripped)
        
        return steps
    
    def extract_databricks_execution_steps(self, content):
        """Extract execution steps from Databricks notebook"""
        return self.extract_pyspark_execution_steps(content)  # Same as PySpark
    
    def extract_sql_execution_steps(self, content):
        """Extract execution steps from SQL script"""
        steps = []
        lines = content.split('\n')
        
        current_step = None
        for line in lines:
            line_stripped = line.strip()
            
            # Look for section headers
            if line_stripped.startswith('--') and ('STEP' in line_stripped or 'PHASE' in line_stripped):
                current_step = line_stripped.replace('--', '').strip()
                steps.append({'step': current_step, 'operations': []})
            
            # Look for SQL operations
            elif current_step and (line_stripped.upper().startswith(('SELECT', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'ALTER', 'DROP'))):
                if steps:
                    steps[-1]['operations'].append(line_stripped)
        
        return steps
    
    def extract_cm2_matching_passes(self, content):
        """Extract CM2 matching passes from Pig script"""
        passes = []
        lines = content.split('\n')
        
        for line in lines:
            line_stripped = line.strip()
            if 'PASS' in line_stripped and ('FIRST' in line_stripped or 'SECOND' in line_stripped or 'THIRD' in line_stripped or 'FOURTH' in line_stripped or 'FIFTH' in line_stripped or 'SIXTH' in line_stripped):
                passes.append(line_stripped)
        
        return passes
    
    def extract_pig_data_flow_detailed(self, content):
        """Extract detailed data flow from Pig script"""
        flow = []
        lines = content.split('\n')
        
        for line in lines:
            line_stripped = line.strip()
            if '=' in line_stripped and ('LOAD' in line_stripped or 'FOREACH' in line_stripped or 'FILTER' in line_stripped or 'GROUP' in line_stripped or 'JOIN' in line_stripped or 'STORE' in line_stripped):
                flow.append(line_stripped)
        
        return flow
    
    def extract_pyspark_data_flow_detailed(self, content):
        """Extract detailed data flow from PySpark script"""
        flow = []
        lines = content.split('\n')
        
        for line in lines:
            line_stripped = line.strip()
            if '=' in line_stripped and ('.' in line_stripped and ('load' in line_stripped or 'select' in line_stripped or 'filter' in line_stripped or 'groupBy' in line_stripped or 'join' in line_stripped or 'save' in line_stripped)):
                flow.append(line_stripped)
        
        return flow
    
    def extract_databricks_data_flow_detailed(self, content):
        """Extract detailed data flow from Databricks notebook"""
        return self.extract_pyspark_data_flow_detailed(content)  # Same as PySpark
    
    def extract_business_logic_from_comments(self, comments):
        """Extract business logic from comments"""
        logic_hints = []
        
        for comment in comments:
            comment_lower = comment.lower()
            
            for key, patterns in self.business_logic_patterns.items():
                for pattern in patterns:
                    if pattern in comment_lower:
                        logic_hints.append(f"{key}_processing")
                        break
        
        return list(set(logic_hints))
    
    def extract_functionality_description(self, comments, content):
        """Extract detailed functionality description from comments and content"""
        functionality = []
        
        # Look for purpose/functionality comments
        for comment in comments:
            comment_lower = comment.lower()
            if any(keyword in comment_lower for keyword in ['purpose', 'function', 'does', 'process', 'handle', 'generate', 'create', 'validate', 'transform']):
                functionality.append(comment.strip())
        
        # Look for specific patterns in content
        content_lower = content.lower()
        
        # Check for specific functionality patterns
        if 'cm2' in content_lower:
            functionality.append("CM2 matching process")
        if 'permid' in content_lower:
            functionality.append("PERMID processing")
        if 'policy' in content_lower:
            functionality.append("Policy information processing")
        if 'address' in content_lower:
            functionality.append("Address processing")
        if 'demographic' in content_lower:
            functionality.append("Demographic processing")
        if 'validation' in content_lower:
            functionality.append("Data validation")
        if 'merge' in content_lower or 'join' in content_lower:
            functionality.append("Data merging/joining")
        
        return list(set(functionality))
    
    def extract_python_functions_detailed(self, content):
        """Extract function definitions with their purposes"""
        functions = []
        
        # Extract function definitions
        pattern = r"def\s+(\w+)\s*\([^)]*\):"
        matches = re.findall(pattern, content)
        
        for func_name in matches:
            # Try to find docstring or comment for the function
            func_pattern = rf"def\s+{func_name}\s*\([^)]*\):\s*\n\s*\"\"\"([^\"]*)\"\"\""
            docstring_match = re.search(func_pattern, content, re.DOTALL)
            
            if docstring_match:
                docstring = docstring_match.group(1).strip()
                functions.append({'name': func_name, 'purpose': docstring})
            else:
                functions.append({'name': func_name, 'purpose': 'No documentation found'})
        
        return functions
    
    def build_hadoop_execution_flow(self):
        """Build Hadoop execution flow from analyzed processes"""
        # Group processes by workflow
        workflow_groups = defaultdict(list)
        
        for process in self.hadoop_processes:
            # Extract workflow name from path
            path_parts = process['path'].split('/')
            if 'app-' in '/'.join(path_parts):
                workflow_name = [part for part in path_parts if part.startswith('app-')][0]
                workflow_groups[workflow_name].append(process)
        
        # Build execution order
        for workflow_name, processes in workflow_groups.items():
            # Sort by file name to get execution order
            processes.sort(key=lambda x: x['name'])
            
            self.hadoop_execution_flows[f"Hadoop_{workflow_name}"] = {
                'type': 'Hadoop Workflow',
                'processes': processes,
                'execution_order': [p['name'] for p in processes]
            }
    
    def build_databricks_execution_flow(self):
        """Build Databricks execution flow from analyzed processes"""
        # Group processes by folder
        folder_groups = defaultdict(list)
        
        for process in self.databricks_processes:
            # Extract folder name from path
            path_parts = process['path'].split('/')
            if len(path_parts) > 1:
                folder_name = path_parts[-2]  # Parent folder
                folder_groups[folder_name].append(process)
        
        # Build execution order
        for folder_name, processes in folder_groups.items():
            # Sort by file name to get execution order
            processes.sort(key=lambda x: x['name'])
            
            self.databricks_execution_flows[f"Databricks_{folder_name}"] = {
                'type': 'Databricks Workflow',
                'processes': processes,
                'execution_order': [p['name'] for p in processes]
            }
    
    def validate_migration_coverage(self):
        """Validate that Azure Databricks covers all Hadoop functionality"""
        print("🔍 Validating Migration Coverage...")
        
        validation_results = {
            'covered_processes': [],
            'missing_processes': [],
            'partial_coverage': [],
            'coverage_percentage': 0
        }
        
        total_hadoop_processes = len(self.hadoop_processes)
        covered_count = 0
        
        for hadoop_process in self.hadoop_processes:
            best_match = self.find_best_databricks_match(hadoop_process)
            
            if best_match and best_match['similarity'] > 0.7:
                validation_results['covered_processes'].append({
                    'hadoop_process': hadoop_process['name'],
                    'databricks_process': best_match['name'],
                    'similarity': best_match['similarity'],
                    'hadoop_functionality': hadoop_process.get('functionality', []),
                    'databricks_functionality': best_match.get('functionality', [])
                })
                covered_count += 1
            elif best_match and best_match['similarity'] > 0.4:
                validation_results['partial_coverage'].append({
                    'hadoop_process': hadoop_process['name'],
                    'databricks_process': best_match['name'],
                    'similarity': best_match['similarity'],
                    'hadoop_functionality': hadoop_process.get('functionality', []),
                    'databricks_functionality': best_match.get('functionality', []),
                    'gaps': best_match.get('differences', [])
                })
            else:
                validation_results['missing_processes'].append({
                    'hadoop_process': hadoop_process['name'],
                    'hadoop_functionality': hadoop_process.get('functionality', []),
                    'hadoop_business_logic': hadoop_process.get('business_logic', [])
                })
        
        validation_results['coverage_percentage'] = (covered_count / total_hadoop_processes * 100) if total_hadoop_processes > 0 else 0
        
        self.validation_results = validation_results
        
        print(f"✅ Migration Coverage: {validation_results['coverage_percentage']:.1f}%")
        print(f"✅ Fully Covered: {len(validation_results['covered_processes'])} processes")
        print(f"⚠️ Partial Coverage: {len(validation_results['partial_coverage'])} processes")
        print(f"❌ Missing Coverage: {len(validation_results['missing_processes'])} processes")
    
    def find_best_databricks_match(self, hadoop_process):
        """Find best Databricks match for Hadoop process"""
        best_match = None
        best_score = 0
        
        for databricks_process in self.databricks_processes:
            score = self.calculate_process_similarity(hadoop_process, databricks_process)
            if score > best_score:
                best_score = score
                best_match = {
                    'name': databricks_process['name'],
                    'similarity': score,
                    'differences': self.get_process_differences(hadoop_process, databricks_process),
                    'functionality': databricks_process.get('functionality', [])
                }
        
        return best_match if best_score > 0.3 else None
    
    def calculate_process_similarity(self, hadoop_process, databricks_process):
        """Calculate similarity between processes"""
        score = 0.0
        
        # Filename similarity
        hadoop_name = hadoop_process['name'].lower()
        databricks_name = databricks_process['name'].lower()
        
        common_keywords = ['permid', 'policy', 'address', 'consumer', 'phone', 'merge', 'publish', 'validate', 'parse', 'cm2', 'bdf', 'escan', 'ich', 'chc', 'gmrn']
        for keyword in common_keywords:
            if keyword in hadoop_name and keyword in databricks_name:
                score += 0.2
        
        # Business logic similarity
        hadoop_logic = set(hadoop_process['business_logic'])
        databricks_logic = set(databricks_process['business_logic'])
        logic_intersection = hadoop_logic.intersection(databricks_logic)
        if hadoop_logic or databricks_logic:
            score += len(logic_intersection) / max(len(hadoop_logic), len(databricks_logic)) * 0.4
        
        # Functionality similarity
        hadoop_functionality = set(hadoop_process.get('functionality', []))
        databricks_functionality = set(databricks_process.get('functionality', []))
        functionality_intersection = hadoop_functionality.intersection(databricks_functionality)
        if hadoop_functionality or databricks_functionality:
            score += len(functionality_intersection) / max(len(hadoop_functionality), len(databricks_functionality)) * 0.3
        
        # Comments similarity
        hadoop_comments = ' '.join(hadoop_process.get('comments', [])).lower()
        databricks_comments = ' '.join(databricks_process.get('comments', [])).lower()
        
        common_terms = ['permid', 'policy', 'address', 'consumer', 'phone', 'merge', 'publish', 'validate', 'parse', 'cm2', 'bdf']
        for term in common_terms:
            if term in hadoop_comments and term in databricks_comments:
                score += 0.05
        
        return min(score, 1.0)
    
    def get_process_differences(self, hadoop_process, databricks_process):
        """Get differences between processes"""
        differences = []
        
        # Type difference
        if hadoop_process['type'] != databricks_process['type']:
            differences.append(f"Type: {hadoop_process['type']} vs {databricks_process['type']}")
        
        # Business logic differences
        hadoop_logic = set(hadoop_process['business_logic'])
        databricks_logic = set(databricks_process['business_logic'])
        
        hadoop_only = hadoop_logic - databricks_logic
        databricks_only = databricks_logic - hadoop_logic
        
        if hadoop_only:
            differences.append(f"Hadoop only: {', '.join(hadoop_only)}")
        if databricks_only:
            differences.append(f"Databricks only: {', '.join(databricks_only)}")
        
        # Functionality differences
        hadoop_functionality = set(hadoop_process.get('functionality', []))
        databricks_functionality = set(databricks_process.get('functionality', []))
        
        hadoop_func_only = hadoop_functionality - databricks_functionality
        databricks_func_only = databricks_functionality - hadoop_functionality
        
        if hadoop_func_only:
            differences.append(f"Hadoop functionality only: {', '.join(hadoop_func_only)}")
        if databricks_func_only:
            differences.append(f"Databricks functionality only: {', '.join(databricks_func_only)}")
        
        return differences
    
    def create_migration_validation_excel(self, output_file="MIGRATION_VALIDATION_REPORT.xlsx"):
        """Create comprehensive migration validation Excel report"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create all sheets
        self.create_hadoop_legacy_sheet(wb)
        self.create_databricks_current_sheet(wb)
        self.create_execution_flow_comparison_sheet(wb)
        self.create_migration_coverage_sheet(wb)
        self.create_validation_summary_sheet(wb)
        
        wb.save(output_file)
        print(f"📊 Migration validation report created: {output_file}")
    
    def create_hadoop_legacy_sheet(self, wb):
        """Create Hadoop Legacy System sheet"""
        ws = wb.create_sheet("Hadoop_Legacy_System")
        
        headers = [
            "Process Name", "Type", "Path", "Execution Order", "Functionality", "CM2 Passes", 
            "Business Logic", "Comments Summary", "Data Flow Operations"
        ]
        ws.append(headers)
        
        for process in self.hadoop_processes:
            execution_steps = []
            for step in process.get('execution_steps', []):
                execution_steps.append(f"{step.get('step', 'Unknown')}: {len(step.get('operations', []))} operations")
            
            cm2_passes = process.get('cm2_passes', [])
            comments_summary = self.summarize_comments(process.get('comments', []))
            functionality = ', '.join(process.get('functionality', []))
            data_flow = ', '.join(process.get('data_flow', [])[:3])  # First 3 operations
            
            row = [
                process['name'],
                process['type'],
                process['relative_path'],
                '; '.join(execution_steps),
                functionality,
                '; '.join(cm2_passes),
                ', '.join(process['business_logic']),
                comments_summary,
                data_flow
            ]
            ws.append(row)
        
        self.format_sheet(ws)
    
    def create_databricks_current_sheet(self, wb):
        """Create Azure Databricks Current System sheet"""
        ws = wb.create_sheet("Azure_Databricks_Current")
        
        headers = [
            "Process Name", "Type", "Path", "Execution Order", "Functionality", "Functions", 
            "Business Logic", "Comments Summary", "Data Flow Operations"
        ]
        ws.append(headers)
        
        for process in self.databricks_processes:
            execution_steps = []
            for step in process.get('execution_steps', []):
                execution_steps.append(f"{step.get('step', 'Unknown')}: {len(step.get('operations', []))} operations")
            
            functions = [f['name'] for f in process.get('functions', [])]
            comments_summary = self.summarize_comments(process.get('comments', []))
            functionality = ', '.join(process.get('functionality', []))
            data_flow = ', '.join(process.get('data_flow', [])[:3])  # First 3 operations
            
            row = [
                process['name'],
                process['type'],
                process['relative_path'],
                '; '.join(execution_steps),
                functionality,
                ', '.join(functions),
                ', '.join(process['business_logic']),
                comments_summary,
                data_flow
            ]
            ws.append(row)
        
        self.format_sheet(ws)
    
    def create_execution_flow_comparison_sheet(self, wb):
        """Create Execution Flow Comparison sheet"""
        ws = wb.create_sheet("Execution_Flow_Comparison")
        
        headers = [
            "Environment", "Workflow/Process", "Type", "Execution Order", "Key Processes", "Purpose"
        ]
        ws.append(headers)
        
        # Add Hadoop flows
        for flow_name, flow_data in self.hadoop_execution_flows.items():
            processes = flow_data.get('processes', [])
            key_processes = [p['name'] for p in processes[:5]]  # First 5 processes
            purpose = self.infer_workflow_purpose(processes)
            
            row = [
                "Hadoop (Legacy)",
                flow_name,
                flow_data.get('type', 'Unknown'),
                ' → '.join(flow_data.get('execution_order', [])),
                ', '.join(key_processes),
                purpose
            ]
            ws.append(row)
        
        # Add Databricks flows
        for flow_name, flow_data in self.databricks_execution_flows.items():
            processes = flow_data.get('processes', [])
            key_processes = [p['name'] for p in processes[:5]]  # First 5 processes
            purpose = self.infer_workflow_purpose(processes)
            
            row = [
                "Azure Databricks (Current)",
                flow_name,
                flow_data.get('type', 'Unknown'),
                ' → '.join(flow_data.get('execution_order', [])),
                ', '.join(key_processes),
                purpose
            ]
            ws.append(row)
        
        self.format_sheet(ws)
    
    def create_migration_coverage_sheet(self, wb):
        """Create Migration Coverage sheet"""
        ws = wb.create_sheet("Migration_Coverage")
        
        headers = [
            "Hadoop Process", "Hadoop Functionality", "Databricks Process", "Databricks Functionality", 
            "Similarity Score", "Coverage Status", "Gaps/Differences"
        ]
        ws.append(headers)
        
        # Add covered processes
        for item in self.validation_results.get('covered_processes', []):
            row = [
                item['hadoop_process'],
                ', '.join(item['hadoop_functionality']),
                item['databricks_process'],
                ', '.join(item['databricks_functionality']),
                f"{item['similarity']:.2f}",
                "✅ FULLY COVERED",
                "None"
            ]
            ws.append(row)
        
        # Add partial coverage
        for item in self.validation_results.get('partial_coverage', []):
            row = [
                item['hadoop_process'],
                ', '.join(item['hadoop_functionality']),
                item['databricks_process'],
                ', '.join(item['databricks_functionality']),
                f"{item['similarity']:.2f}",
                "⚠️ PARTIAL COVERAGE",
                '; '.join(item.get('gaps', []))
            ]
            ws.append(row)
        
        # Add missing processes
        for item in self.validation_results.get('missing_processes', []):
            row = [
                item['hadoop_process'],
                ', '.join(item['hadoop_functionality']),
                "NOT FOUND",
                "N/A",
                "0.00",
                "❌ MISSING COVERAGE",
                ', '.join(item['hadoop_business_logic'])
            ]
            ws.append(row)
        
        self.format_sheet(ws)
    
    def create_validation_summary_sheet(self, wb):
        """Create Validation Summary sheet"""
        ws = wb.create_sheet("Validation_Summary")
        
        headers = ["Metric", "Value", "Details"]
        ws.append(headers)
        
        summary_data = [
            ["Total Hadoop Processes", len(self.hadoop_processes), "Legacy system processes analyzed"],
            ["Total Databricks Processes", len(self.databricks_processes), "Current system processes analyzed"],
            ["Fully Covered Processes", len(self.validation_results.get('covered_processes', [])), "Processes with >70% similarity"],
            ["Partially Covered Processes", len(self.validation_results.get('partial_coverage', [])), "Processes with 40-70% similarity"],
            ["Missing Processes", len(self.validation_results.get('missing_processes', [])), "Processes with <40% similarity"],
            ["Overall Coverage Percentage", f"{self.validation_results.get('coverage_percentage', 0):.1f}%", "Migration coverage score"],
            ["Hadoop Execution Flows", len(self.hadoop_execution_flows), "Legacy workflow flows"],
            ["Databricks Execution Flows", len(self.databricks_execution_flows), "Current workflow flows"]
        ]
        
        for row_data in summary_data:
            ws.append(row_data)
        
        self.format_sheet(ws)
    
    def infer_workflow_purpose(self, processes):
        """Infer workflow purpose from processes"""
        purposes = []
        for process in processes:
            purposes.extend(process.get('business_logic', []))
        
        # Count most common purposes
        purpose_counts = {}
        for purpose in purposes:
            purpose_counts[purpose] = purpose_counts.get(purpose, 0) + 1
        
        # Return top 3 purposes
        sorted_purposes = sorted(purpose_counts.items(), key=lambda x: x[1], reverse=True)
        return ', '.join([p[0] for p in sorted_purposes[:3]])
    
    def summarize_comments(self, comments):
        """Summarize comments"""
        if not comments:
            return "No comments"
        
        # Take first few meaningful comments
        meaningful_comments = []
        for comment in comments[:3]:
            if len(comment.strip()) > 10:  # Skip very short comments
                meaningful_comments.append(comment[:50] + "..." if len(comment) > 50 else comment)
        
        return '; '.join(meaningful_comments) if meaningful_comments else "No meaningful comments"
    
    def format_sheet(self, ws):
        """Format Excel sheet with headers and styling"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

def main():
    """Main function with hardcoded paths"""
    print("=== Migration Validation Tool (Hardcoded Paths) ===")
    print("This tool validates that Azure Databricks migration covers all Hadoop legacy functionality")
    print()
    
    # Hardcoded paths based on your environment
    hadoop_path = "app-cdd"
    databricks_path = "CDD"
    output_file = "MIGRATION_VALIDATION_REPORT.xlsx"
    
    print(f"🔍 Using hardcoded paths:")
    print(f"   Hadoop (Legacy): {hadoop_path}")
    print(f"   Databricks (Current): {databricks_path}")
    print(f"   Output file: {output_file}")
    print()
    
    # Check if paths exist
    if not os.path.exists(hadoop_path):
        print(f"❌ Error: Hadoop repository path does not exist: {hadoop_path}")
        print(f"   Current directory: {os.getcwd()}")
        print(f"   Available files: {os.listdir('.')}")
        return
    
    if not os.path.exists(databricks_path):
        print(f"❌ Error: Databricks repository path does not exist: {databricks_path}")
        print(f"   Current directory: {os.getcwd()}")
        print(f"   Available files: {os.listdir('.')}")
        return
    
    # Create validator and analyze both systems
    validator = MigrationValidationTool()
    
    # Analyze both systems
    validator.analyze_hadoop_legacy(hadoop_path)
    validator.analyze_azure_databricks_current(databricks_path)
    
    # Validate migration coverage
    validator.validate_migration_coverage()
    
    # Create comprehensive Excel report
    validator.create_migration_validation_excel(output_file)
    
    print(f"\n🎉 Migration validation complete!")
    print(f"📊 Results saved to: {output_file}")
    print(f"📈 Coverage: {validator.validation_results.get('coverage_percentage', 0):.1f}%")

if __name__ == "__main__":
    main()
