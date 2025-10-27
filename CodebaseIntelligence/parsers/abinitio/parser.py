"""
Enhanced Ab Initio Parser
Parses .mp files to extract:
- Components (DataSet sheet equivalent)
- Graph parameters
- GraphFlow (data lineage)
- DML schemas and field mappings
- Transformation logic
"""

import os
import re
import hashlib
from typing import List, Dict, Any, Optional
from pathlib import Path
from loguru import logger

from core.models import Component, Process, ComponentType, ProcessType, SystemType, DataFlow
from .mp_file_parser import MPFileParser
from .dml_parser import DMLParser
from .graph_flow_extractor import GraphFlowExtractor


class AbInitioParser:
    """Main Ab Initio parser"""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        self.config = config or {}
        self.mp_parser = MPFileParser()
        self.dml_parser = DMLParser()
        self.flow_extractor = GraphFlowExtractor()

        self.processes: List[Process] = []
        self.components: List[Component] = []
        self.data_flows: List[DataFlow] = []

    def parse_directory(self, directory: str) -> Dict[str, Any]:
        """Parse all Ab Initio files in a directory"""
        logger.info(f"Parsing Ab Initio directory: {directory}")

        mp_files = list(Path(directory).rglob("*.mp"))
        logger.info(f"Found {len(mp_files)} .mp files")

        for mp_file in mp_files:
            try:
                self.parse_file(str(mp_file))
            except Exception as e:
                logger.error(f"Error parsing {mp_file}: {e}")
                continue

        return {
            "processes": self.processes,
            "components": self.components,
            "data_flows": self.data_flows,
            "summary": {
                "total_processes": len(self.processes),
                "total_components": len(self.components),
                "total_flows": len(self.data_flows),
            },
        }

    def parse_file(self, file_path: str) -> Process:
        """Parse a single .mp file"""
        logger.info(f"Parsing file: {file_path}")

        # Read file content
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()

        # Parse .mp file
        mp_data = self.mp_parser.parse(content)

        # Create process
        process = self._create_process(file_path, mp_data)

        # Extract components
        components = self._extract_components(process, mp_data, content)

        # Extract graph flow
        data_flows = self.flow_extractor.extract_flows(mp_data, components)

        # Update process with component info
        process.component_ids = [c.id for c in components]
        process.component_count = len(components)

        # Store results
        self.processes.append(process)
        self.components.extend(components)
        self.data_flows.extend(data_flows)

        logger.info(
            f"Extracted {len(components)} components and {len(data_flows)} flows from {Path(file_path).name}"
        )

        return process

    def _create_process(self, file_path: str, mp_data: Dict[str, Any]) -> Process:
        """Create Process object from .mp file"""
        file_name = Path(file_path).stem
        process_id = self._generate_id(f"abinitio_process_{file_name}")

        # Extract graph parameters
        graph_params = mp_data.get("graph_parameters", {})

        # Try to infer business function
        business_function = self._infer_business_function(file_name, graph_params)

        return Process(
            id=process_id,
            name=file_name,
            system=SystemType.ABINITIO,
            process_type=ProcessType.GRAPH,
            file_path=file_path,
            repo_name=self._extract_repo_name(file_path),
            description=f"Ab Initio graph: {file_name}",
            business_function=business_function,
            parameters=graph_params,
            graph_parameters=graph_params,
            business_domain=self._infer_domain(file_name),
            tags=self._extract_tags(file_name),
        )

    def _extract_components(
        self, process: Process, mp_data: Dict[str, Any], content: str
    ) -> List[Component]:
        """Extract all components from .mp file"""
        components = []

        # Parse datasets (components)
        datasets = mp_data.get("datasets", [])

        for idx, dataset in enumerate(datasets):
            component = self._create_component(process, dataset, idx, content)
            components.append(component)

        return components

    def _create_component(
        self,
        process: Process,
        dataset: Dict[str, Any],
        index: int,
        full_content: str,
    ) -> Component:
        """Create Component object from dataset info"""
        comp_name = dataset.get("name", f"component_{index}")
        comp_id = self._generate_id(f"{process.id}_{comp_name}")

        # Map component type
        comp_type_str = dataset.get("type", "Unknown")
        comp_type = self._map_component_type(comp_type_str)

        # Extract schema from DML
        dml = dataset.get("dml", "")
        schema = self.dml_parser.parse_dml(dml) if dml else None

        # Extract transformation logic
        transformation_logic = self._extract_transformation_logic(
            dataset, full_content
        )

        # Extract file paths / dataset names
        dataset_names = dataset.get("dataset_names", "").split(",")
        input_datasets = [d.strip() for d in dataset_names if d.strip()]

        return Component(
            id=comp_id,
            name=comp_name,
            component_type=comp_type,
            system="abinitio",
            file_path=process.file_path,
            line_number=dataset.get("line_number"),
            process_id=process.id,
            process_name=process.name,
            input_datasets=input_datasets if comp_type != ComponentType.OUTPUT_FILE else [],
            output_datasets=input_datasets if comp_type == ComponentType.OUTPUT_FILE else [],
            input_schema=schema,
            output_schema=schema,
            dml_definition=dml,
            business_description=self._generate_component_description(
                comp_name, comp_type_str
            ),
            transformation_logic=transformation_logic,
            key_parameter_value=dataset.get("key_parameter_value"),
            parameters=dataset.get("parameters", {}),
            tags=self._extract_component_tags(comp_name, comp_type_str),
        )

    def _map_component_type(self, comp_type_str: str) -> ComponentType:
        """Map Ab Initio component type to ComponentType enum"""
        type_map = {
            "Input_File": ComponentType.INPUT_FILE,
            "Output_File": ComponentType.OUTPUT_FILE,
            "Lookup_File": ComponentType.LOOKUP_FILE,
            "Transform": ComponentType.TRANSFORM,
            "Join": ComponentType.JOIN,
            "Filter": ComponentType.FILTER,
            "Aggregate": ComponentType.AGGREGATE,
            "Sort": ComponentType.SORT,
            "Rollup": ComponentType.AGGREGATE,
            "Reformat": ComponentType.TRANSFORM,
            "Normalize": ComponentType.TRANSFORM,
            "Denormalize": ComponentType.TRANSFORM,
        }
        return type_map.get(comp_type_str, ComponentType.UNKNOWN)

    def _extract_transformation_logic(
        self, dataset: Dict[str, Any], full_content: str
    ) -> Optional[str]:
        """Extract transformation logic from component"""
        # Look for transform expressions, let expressions, etc.
        comp_name = dataset.get("name", "")

        # Search for transform blocks in the full content
        pattern = rf"{re.escape(comp_name)}.*?transform.*?\((.*?)\)"
        matches = re.findall(pattern, full_content, re.DOTALL | re.IGNORECASE)

        if matches:
            return matches[0][:500]  # Limit length

        return None

    def _infer_business_function(
        self, file_name: str, params: Dict[str, Any]
    ) -> Optional[str]:
        """Infer business function from file name and parameters"""
        name_lower = file_name.lower()

        if "lead" in name_lower or "ipa" in name_lower or "cpa" in name_lower:
            return "Lead Generation / Patient Account Processing"
        elif "cdd" in name_lower or "charlotte" in name_lower:
            return "CDD / ICH Data Processing"
        elif "gmrn" in name_lower or "globalmrn" in name_lower:
            return "Global MRN / Patient Matching"
        elif "coverage" in name_lower or "fc" in name_lower:
            return "Coverage Discovery"
        elif "edi" in name_lower:
            return "EDI Processing"
        elif "demo" in name_lower or "demographics" in name_lower:
            return "Demographics Processing"

        return None

    def _infer_domain(self, file_name: str) -> Optional[str]:
        """Infer business domain"""
        name_lower = file_name.lower()

        if any(x in name_lower for x in ["lead", "ipa", "cpa"]):
            return "Lead Discovery"
        elif any(x in name_lower for x in ["cdd", "charlotte", "ich"]):
            return "CDD"
        elif "gmrn" in name_lower:
            return "Patient Matching"
        elif "coverage" in name_lower:
            return "Coverage Discovery"

        return "General"

    def _extract_tags(self, file_name: str) -> List[str]:
        """Extract tags from file name"""
        tags = ["abinitio"]
        name_lower = file_name.lower()

        tag_keywords = {
            "lead": "lead_generation",
            "ipa": "patient_accounts",
            "cpa": "candidate_accounts",
            "cdd": "cdd_process",
            "gmrn": "patient_matching",
            "coverage": "coverage_discovery",
            "edi": "edi_processing",
        }

        for keyword, tag in tag_keywords.items():
            if keyword in name_lower:
                tags.append(tag)

        return tags

    def _extract_component_tags(self, comp_name: str, comp_type: str) -> List[str]:
        """Extract tags for component"""
        tags = [comp_type.lower()]

        name_lower = comp_name.lower()
        if "lookup" in name_lower or "lkp" in name_lower:
            tags.append("lookup")
        if "temp" in name_lower or "tmp" in name_lower:
            tags.append("temporary")
        if "output" in name_lower or "ofil" in name_lower:
            tags.append("output")
        if "input" in name_lower or "ifil" in name_lower:
            tags.append("input")

        return tags

    def _generate_component_description(
        self, comp_name: str, comp_type: str
    ) -> str:
        """Generate human-readable component description"""
        action_map = {
            "Input_File": "Reads",
            "Output_File": "Writes",
            "Lookup_File": "Looks up",
            "Transform": "Transforms",
            "Join": "Joins",
            "Filter": "Filters",
            "Aggregate": "Aggregates",
        }

        action = action_map.get(comp_type, "Processes")
        return f"{action} data for {comp_name}"

    def _extract_repo_name(self, file_path: str) -> Optional[str]:
        """Extract repository name from file path"""
        parts = Path(file_path).parts
        for part in parts:
            if "abinitio" in part.lower() or "abi" in part.lower():
                return part
        return None

    def _generate_id(self, base: str) -> str:
        """Generate unique ID"""
        return hashlib.md5(base.encode()).hexdigest()[:16]

    def export_to_excel(self, output_path: str):
        """Export parsed data to Excel format (like FAWN output)"""
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # DataSet sheet (Components)
        dataset_data = []
        for comp in self.components:
            dataset_data.append(
                {
                    "CompDisplay Label": comp.name,
                    "Component Type": comp.component_type.value,
                    "DataSet Names": ", ".join(comp.input_datasets or []),
                    "DML": comp.dml_definition or "",
                    "Key Parameter Value": comp.key_parameter_value or "",
                }
            )
        df_dataset = pd.DataFrame(dataset_data)
        ws_dataset = wb.create_sheet("DataSet")
        for r in dataframe_to_rows(df_dataset, index=False, header=True):
            ws_dataset.append(r)

        # GraphParameters sheet
        param_data = []
        for process in self.processes:
            for param_name, param_value in process.graph_parameters.items():
                param_data.append(
                    {
                        "Graph": process.name,
                        "Parameter": param_name,
                        "Value": str(param_value),
                    }
                )
        if param_data:
            df_params = pd.DataFrame(param_data)
            ws_params = wb.create_sheet("GraphParameters")
            for r in dataframe_to_rows(df_params, index=False, header=True):
                ws_params.append(r)

        # GraphFlow sheet
        flow_data = []
        for flow in self.data_flows:
            flow_data.append(
                {
                    "Source Component": flow.source_component_id,
                    "Target Component": flow.target_component_id,
                    "Dataset": flow.dataset_name or "",
                    "Flow Type": flow.flow_type,
                    "Transformation": flow.transformation_applied or "",
                }
            )
        if flow_data:
            df_flow = pd.DataFrame(flow_data)
            ws_flow = wb.create_sheet("GraphFlow")
            for r in dataframe_to_rows(df_flow, index=False, header=True):
                ws_flow.append(r)

        wb.save(output_path)
        logger.info(f"Exported to Excel: {output_path}")
